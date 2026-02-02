
import os
import io
import zipfile
import tempfile
import errno
import shutil
import smtplib
import secrets
import uuid
from typing import Optional
from email.message import EmailMessage
from datetime import datetime, date, timedelta
from flask import Flask, request, redirect, url_for, send_file, abort, flash, render_template_string
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, login_user, login_required, logout_user, current_user, UserMixin
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from sqlalchemy import text as sql_text, and_

APP_VERSION = "v30"


# Zdjęcia: kompresja i konwersja do JPEG przy zapisie
from PIL import Image
from PIL.ImageOps import exif_transpose

# HEIC/HEIF (iPhone) – opcjonalnie. Jeśli biblioteka nie będzie dostępna,
# takie pliki zostaną odrzucone z czytelnym komunikatem.
try:
    from pillow_heif import register_heif_opener  # type: ignore
    register_heif_opener()
    HEIF_SUPPORTED = True
except Exception:
    HEIF_SUPPORTED = False

# --- Flask & DB config ---
BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# Render: jeśli masz podpięty Persistent Disk, Render montuje go zwykle w /var/data.
# Trzymamy tam bazę i uploady, żeby:
# - zapisy działały (czasem katalog z kodem bywa niewygodny do zapisu),
# - pliki nie znikały po deployu/resecie.
DATA_DIR = "/var/data" if os.path.exists("/var/data") else BASE_DIR

DB_FILE = os.path.join(DATA_DIR, "app.db")

# Zdjęcia do wpisów (trzymamy obok bazy, nie w static)
UPLOAD_DIR = os.path.join(DATA_DIR, "uploads")
EXTRA_SIG_DIR = os.path.join(UPLOAD_DIR, "extra_signatures")

# Załączniki do raportów "Dodatki" (admin dodaje pliki/zdjęcia do wysyłki)
EXTRA_REPORT_ATTACH_DIR = os.path.join(UPLOAD_DIR, "extra_report_attachments")
MAX_ATTACH_MB = int(os.getenv("MAX_ATTACH_MB", "25"))
MAX_ATTACH_BYTES = MAX_ATTACH_MB * 1024 * 1024
MAX_ATTACH_COUNT = int(os.getenv("MAX_ATTACH_COUNT", "10"))
ALLOWED_ATTACH_EXTS = {".pdf", ".png", ".jpg", ".jpeg", ".webp", ".doc", ".docx", ".xls", ".xlsx", ".txt"}



# Plany (PDF) przypięte do projektów
PLANS_DIR = os.path.join(DATA_DIR, "plans")
ALLOWED_PLAN_EXTS = {".pdf"}
MAX_PLAN_MB = int(os.getenv("MAX_PLAN_MB", "25"))
MAX_PLAN_BYTES = MAX_PLAN_MB * 1024 * 1024



app = Flask(__name__, static_folder="static", static_url_path="/static")
app.secret_key = os.getenv("SECRET_KEY", "dev-key-change-me")

app.config["SQLALCHEMY_DATABASE_URI"] = f"sqlite:///{DB_FILE}"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"


# --- Models ---
class User(db.Model, UserMixin):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(120), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    is_active_u = db.Column(db.Boolean, default=True)

    def set_password(self, pw: str):
        self.password_hash = generate_password_hash(pw)

    def check_password(self, pw: str) -> bool:
        return check_password_hash(self.password_hash, pw)

    @property
    def is_active(self):
        return self.is_active_u


class Project(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, unique=True)
    is_active = db.Column(db.Boolean, default=True)


class Plan(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey("project.id"), nullable=False, index=True)
    title = db.Column(db.String(200), nullable=True)

    stored_filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=True)

    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    uploaded_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)

    project = db.relationship("Project", backref="plans")
    uploaded_by_user = db.relationship("User", foreign_keys=[uploaded_by])


class Entry(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    project_id = db.Column(db.Integer, db.ForeignKey("project.id"), nullable=False)
    work_date = db.Column(db.Date, nullable=False)
    minutes = db.Column(db.Integer, nullable=False, default=0)
    is_extra = db.Column(db.Boolean, default=False)
    is_overtime = db.Column(db.Boolean, default=False)
    note = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship("User", backref="entries")
    project = db.relationship("Project", backref="entries")

    images = db.relationship(
        "EntryImage",
        backref="entry",
        cascade="all, delete-orphan",
        lazy="select",
    )


class EntryImage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    entry_id = db.Column(db.Integer, db.ForeignKey("entry.id"), nullable=False, index=True)
    stored_filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class Cost(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)
    cost_date = db.Column(db.Date, nullable=False)
    amount = db.Column(db.String(50), nullable=False)
    description = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship("User", backref="costs")


class LeaveRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)

    date_from = db.Column(db.Date, nullable=False)
    date_to = db.Column(db.Date, nullable=False)
    reason = db.Column(db.Text, nullable=True)

    # DRAFT -> SUBMITTED -> APPROVED
    status = db.Column(db.String(20), nullable=False, default="DRAFT")
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    submitted_at = db.Column(db.DateTime, nullable=True)
    decided_at = db.Column(db.DateTime, nullable=True)
    decided_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)

    user = db.relationship("User", foreign_keys=[user_id], backref="leave_requests")
    decided_by_user = db.relationship("User", foreign_keys=[decided_by])


# --- Dodatki (extra godziny z akceptacją raportu) ---

class ProjectContact(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey("project.id"), nullable=False, index=True)
    email = db.Column(db.String(200), nullable=False)
    name = db.Column(db.String(200), nullable=True)
    is_default = db.Column(db.Boolean, default=True)

    project = db.relationship("Project", backref="contacts")


class ExtraRequest(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False, index=True)
    project_id = db.Column(db.Integer, db.ForeignKey("project.id"), nullable=False, index=True)

    work_date = db.Column(db.Date, nullable=False)
    minutes = db.Column(db.Integer, nullable=False, default=0)
    description = db.Column(db.Text, nullable=True)

    status = db.Column(db.String(30), default="NEW")  # NEW / INCLUDED / CANCELED
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    user = db.relationship("User", backref="extra_requests")
    project = db.relationship("Project", backref="extra_requests")

    images = db.relationship(
        "ExtraRequestImage",
        backref="request",
        cascade="all, delete-orphan",
        lazy="select",
    )


class ExtraRequestImage(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    request_id = db.Column(db.Integer, db.ForeignKey("extra_request.id"), nullable=False, index=True)
    stored_filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class ExtraReport(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    project_id = db.Column(db.Integer, db.ForeignKey("project.id"), nullable=False, index=True)
    created_by = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=True)

    # dane wysyłki
    recipient_email = db.Column(db.String(200), nullable=True)
    token = db.Column(db.String(64), unique=True, index=True, nullable=True)

    status = db.Column(db.String(30), default="DRAFT")  # DRAFT / SENT / APPROVED / REJECTED / COMMENTED / APPROVED_AUTO
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    sent_at = db.Column(db.DateTime, nullable=True)
    decided_at = db.Column(db.DateTime, nullable=True)
    decided_note = db.Column(db.Text, nullable=True)

    report_text = db.Column(db.Text, nullable=True)
    total_minutes_override = db.Column(db.Integer, nullable=True)  # jeśli admin chce ręcznie zmienić sumę

    project = db.relationship("Project", backref="extra_reports")
    created_by_user = db.relationship("User", foreign_keys=[created_by])

    items = db.relationship(
        "ExtraReportItem",
        backref="report",
        cascade="all, delete-orphan",
        lazy="select",
    )

    attachments = db.relationship(
        "ExtraReportAttachment",
        backref="report",
        cascade="all, delete-orphan",
        lazy="select",
    )


class ExtraReportItem(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("extra_report.id"), nullable=False, index=True)
    request_id = db.Column(db.Integer, db.ForeignKey("extra_request.id"), nullable=False, index=True)

    # snapshot, żeby można było edytować bez ruszania pierwotnego zgłoszenia
    user_name = db.Column(db.String(120), nullable=False)
    work_date = db.Column(db.Date, nullable=False)
    minutes = db.Column(db.Integer, nullable=False, default=0)
    description = db.Column(db.Text, nullable=True)

    request = db.relationship("ExtraRequest")




@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


# --- Helpers ---
def fmt_hhmm(minutes: int) -> str:
    minutes = minutes or 0
    h = minutes // 60
    m = minutes % 60
    return f"{h:02d}:{m:02d}"

def parse_hhmm(value: str) -> int:
    if not value:
        return 0
    v = value.strip().lower()
    if v.isdigit():
        return int(v)
    if 'h' in v:
        try:
            left, *rest = v.split('h')
            h = int(left) if left else 0
            m = int(rest[0]) if rest and rest[0] else 0
            return h*60 + m
        except Exception:
            pass
    if ':' in v:
        hh, mm = v.split(':', 1)
        h = int(hh) if hh else 0
        m = int(mm) if mm else 0
        return h*60 + m
    try:
        f = float(v.replace(',', '.'))
        return int(round(f*60))
    except Exception:
        return 0

def month_bounds(d: date):
    first = d.replace(day=1)
    if first.month == 12:
        nxt = first.replace(year=first.year + 1, month=1, day=1)
    else:
        nxt = first.replace(month=first.month + 1, day=1)
    last = nxt - timedelta(days=1)
    return first, last

def require_admin():
    if not current_user.is_authenticated or not current_user.is_admin:
        abort(403)

def ensure_db_file():
    os.makedirs(os.path.dirname(DB_FILE) or ".", exist_ok=True)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    os.makedirs(EXTRA_SIG_DIR, exist_ok=True)
    os.makedirs(EXTRA_REPORT_ATTACH_DIR, exist_ok=True)
    os.makedirs(PLANS_DIR, exist_ok=True)
    with app.app_context():
        db.create_all()
        _try_add_column('extra_requests', 'category', 'TEXT')
        try:
            db.session.execute(sql_text("SELECT 1"))
        except Exception:
            pass


# Telefony wysyłają głównie JPG/PNG. iPhone potrafi HEIC/HEIF – obsługujemy,
# jeśli pillow-heif jest dostępne.
ALLOWED_IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}
if HEIF_SUPPORTED:
    ALLOWED_IMAGE_EXTS.update({".heic", ".heif"})

# Ograniczenia, żeby backup i dysk nie puchły
MAX_IMAGE_MB = int(os.getenv("MAX_IMAGE_MB", "8"))  # limit na plik przed kompresją
MAX_IMAGE_BYTES = MAX_IMAGE_MB * 1024 * 1024
IMAGE_MAX_PX = int(os.getenv("IMAGE_MAX_PX", "1600"))  # dłuższy bok
IMAGE_JPEG_QUALITY = int(os.getenv("IMAGE_JPEG_QUALITY", "75"))


def _file_size_bytes(file_storage) -> Optional[int]:
    """Próbuje odczytać rozmiar pliku z FileStorage bez ładowania w pamięć."""
    try:
        if getattr(file_storage, "content_length", None):
            return int(file_storage.content_length)
    except Exception:
        pass
    try:
        stream = file_storage.stream
        pos = stream.tell()
        stream.seek(0, os.SEEK_END)
        size = stream.tell()
        stream.seek(pos)
        return int(size)
    except Exception:
        return None


def _save_compressed_image(file_storage, out_path: str) -> None:
    """Zmniejsza i kompresuje obraz do JPEG, zapisuje na dysku."""
    # Ważne: Image.open czyta ze strumienia, więc nie używamy file.save(...)
    img = Image.open(file_storage.stream)
    img = exif_transpose(img)

    # animowane GIF – bierzemy pierwszą klatkę
    try:
        if getattr(img, "is_animated", False):
            img.seek(0)
    except Exception:
        pass

    if img.mode in ("RGBA", "P"):
        img = img.convert("RGB")
    elif img.mode != "RGB":
        img = img.convert("RGB")

    img.thumbnail((IMAGE_MAX_PX, IMAGE_MAX_PX))
    img.save(out_path, format="JPEG", quality=IMAGE_JPEG_QUALITY, optimize=True, progressive=True)


def _safe_image_filename(original_name: str, entry_id: int) -> str:
    """Buduje bezpieczną, unikalną nazwę pliku dla zdjęcia wpisu."""
    # Zawsze zapisujemy jako JPEG (mniejszy plik + prostszy backup)
    return f"e{entry_id}_{uuid.uuid4().hex}.jpg"


def _safe_plan_filename(original_name: str, project_id: int) -> str:
    # zapisujemy zawsze jako pdf; unikalna nazwa na dysku
    return f"p{project_id}_{uuid.uuid4().hex}.pdf"



def _save_entry_images(entry, files):
    """Zapisuje zdjęcia do UPLOAD_DIR i tworzy rekordy w bazie."""
    if not files:
        return
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    for f in files:
        if not files or not any(getattr(x, "filename", "") for x in files):
            continue
        name = f.filename
        _, ext = os.path.splitext(name)
        ext = (ext or "").lower()
        if ext not in ALLOWED_IMAGE_EXTS:
            continue

        # HEIC/HEIF bez pillow-heif nie otworzymy – dajemy jasny komunikat
        if ext in {".heic", ".heif"} and not HEIF_SUPPORTED:
            flash("Ten telefon wysłał zdjęcie HEIC/HEIF. Zmień w iPhonie: Ustawienia > Aparat > Formaty > Najbardziej zgodne (JPG).")
            continue

        size = _file_size_bytes(f)
        if size is not None and size > MAX_IMAGE_BYTES:
            flash(f"Zdjęcie jest za duże ({MAX_IMAGE_MB} MB max). Zmniejsz je lub wyślij mniejsze.")
            continue

        stored = _safe_image_filename(name, entry.id)
        path = os.path.join(UPLOAD_DIR, stored)
        try:
            # upewnij się, że czytamy od początku strumienia
            try:
                f.stream.seek(0)
            except Exception:
                pass
            _save_compressed_image(f, path)
            db.session.add(EntryImage(entry_id=entry.id, stored_filename=stored, original_filename=name))
        except Exception:
            # Nie blokujemy dodawania godzin, ale informujemy
            flash("Nie udało się zapisać jednego ze zdjęć. Spróbuj ponownie lub wyślij inne zdjęcie.")


def _delete_entry_images_files(entry):
    """Kasuje z dysku zdjęcia przypisane do wpisu."""
    try:
        imgs = list(entry.images)
    except Exception:
        imgs = []
    for img in imgs:
        try:
            p = os.path.join(UPLOAD_DIR, img.stored_filename)
            if os.path.exists(p):
                os.remove(p)
        except Exception:
            pass


# --- Init DB (safe) ---


def _try_add_column(table: str, column: str, coltype: str = "TEXT"):
    """Best-effort SQLite schema tweak (no migrations)."""
    try:
        cols = [r[1] for r in db.session.execute(text(f"PRAGMA table_info({table})")).fetchall()]
        if column not in cols:
            db.session.execute(text(f"ALTER TABLE {table} ADD COLUMN {column} {coltype}"))
            db.session.commit()
    except Exception:
        db.session.rollback()
def init_db():
    ensure_db_file()
    with app.app_context():
        try:
            has_user = db.session.query(User.id).first()
        except Exception:
            db.create_all()
            has_user = None

        if not has_user:
            admin = User(name="Administrator", email="admin@local", is_admin=True)
            admin.set_password("admin123")
            db.session.add(admin)

        if not db.session.query(Project.id).first():
            db.session.add(Project(name="Projekt domyślny", is_active=True))

        db.session.commit()



# --- Base layout (jasny) ---
BASE = """
<!doctype html>
<html lang="pl">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>{{ title or 'EKKO NOR AS – Rejestrator czasu pracy' }}</title>
  <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
  <style>
    :root { color-scheme: light; }
    body{ background:#f5f7fb; color:#1f2937; }
    .navbar{ background:#ffffff; border-bottom:1px solid #e5e7eb; }
    .card{ background:#ffffff; border:1px solid #e5e7eb; border-radius:14px; }
    .form-control,.form-select,.form-check-input{ background:#ffffff; color:#111827; border:1px solid #d1d5db; }
    .btn-primary{ background:#2563eb; border-color:#2563eb; }
    .btn-outline-primary{ border-color:#2563eb; color:#2563eb; }
    .btn-outline-primary:hover{ background:#2563eb; color:white; }
    .table{ color:#111827; }
    .table thead{ background:#f3f4f6; }
    .badge-soft{ background:#eef2ff; border:1px solid #c7d2fe; color:#3730a3; }
    .brand-logo{ height:36px; }
    .brand-big{ max-width:180px; display:block; margin:0 auto 16px; }
    a{ color:#2563eb; }
    .container-narrow{ max-width:1100px; }
  
    @media (max-width: 576px){
      /* większe tap-targety i brak iOS zoom w polach */
      .btn{ min-height:44px; padding-top:.6rem; padding-bottom:.6rem; }
      .btn-sm{ min-height:44px; padding-top:.55rem; padding-bottom:.55rem; font-size:0.95rem; }
      .form-control,.form-select{ font-size:16px; min-height:44px; }
      .navbar .nav-link{ padding:.5rem 0; }
      .table{ display:block; overflow-x:auto; white-space:nowrap; -webkit-overflow-scrolling:touch; }
    }

  </style>
</head>
<body>
<nav class="navbar navbar-expand-lg navbar-light mb-4">
  <div class="container-fluid">
    <a class="navbar-brand d-flex align-items-center" href="{{ url_for('dashboard') if current_user.is_authenticated else url_for('login') }}">
      <img src="{{ url_for('static', filename='ekko_logo.png') }}" class="brand-logo me-2" alt="logo">
    </a>

    {% if current_user.is_authenticated %}
      <button class="navbar-toggler" type="button" data-bs-toggle="collapse" data-bs-target="#navMenu" aria-controls="navMenu" aria-expanded="false" aria-label="Menu">
        <span class="navbar-toggler-icon"></span>
      </button>

      <div class="collapse navbar-collapse" id="navMenu">
        <ul class="navbar-nav me-auto mb-2 mb-lg-0">
          {% if current_user.is_admin %}
            <li class="nav-item"><a class="nav-link" href="{{ url_for('admin_overview') }}">Admin</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('admin_users') }}">Pracownicy</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('admin_projects') }}">Projekty</a></li>
                        <li class="nav-item"><a class="nav-link" href="{{ url_for(\'admin_plans\') }}">Plany (PDF)</a></li>
<li class="nav-item"><a class="nav-link" href="{{ url_for('admin_entries') }}">Godziny (admin)</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('admin_reports') }}">Raporty</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('admin_extras') }}">Dodatki</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('leaves') }}">Urlopy</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('plans') }}">Plany</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('admin_costs') }}">Koszty</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('admin_plans') }}">Plany</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('admin_backup') }}">Backup</a></li>
          {% else %}
            <li class="nav-item"><a class="nav-link" href="{{ url_for('dashboard') }}">Godziny</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('extras') }}">Dodatki</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('user_summary') }}">Podsumowanie</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('user_costs') }}">Koszty</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('leaves') }}">Urlopy</a></li>
            <li class="nav-item"><a class="nav-link" href="{{ url_for('plans') }}">Plany</a></li>
          {% endif %}
        </ul>

        <div class="d-flex flex-column flex-lg-row gap-2 align-items-start align-items-lg-center">
          <span class="text-muted small">{{ current_user.name }}</span>
          <span class="badge bg-secondary">{{ app_version }}</span>
          <a class="btn btn-sm btn-danger" href="{{ url_for('logout') }}">Wyloguj</a>
        </div>
      </div>
    {% endif %}
  </div>
</nav>

<div class="container container-narrow mb-4">
  {% with messages = get_flashed_messages() %}
    {% if messages %}
      <div class="alert alert-warning">{{ messages[0] }}</div>
    {% endif %}
  {% endwith %}
  {{ body|safe }}
</div>

<div class="text-center mt-4 text-muted" style="font-size:12px; line-height:1.4;">Ekko Nor AS<br>Bruseveien 8A<br>1911 Flateby<br><br>Admin: dataconnect.no</div>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/js/bootstrap.bundle.min.js"></script>

<script>
function limitFiles(input, max){
  if (!input || !input.files) return;
  if (input.files.length > max) {
    alert('Możesz dodać maksymalnie ' + max + ' plików.');
    input.value = '';
  }
}
</script>

</body>
</html>
"""

def layout(title, body):
    return render_template_string(BASE, title=title, body=body, fmt=fmt_hhmm, app_version=APP_VERSION)




@app.route("/logout")
@login_required
def logout():
    logout_user()
    return redirect(url_for("login"))


EXTRA_CATEGORIES = [
    ("arbeid", "Arbeid"),
    ("material", "Materialer"),
    ("transport", "Transport"),
    ("leie", "Leie/utstyr"),
    ("annet", "Annet"),
]

# --- Dodatki: pomocnicze ---
def _default_project_contact_email(project_id: int) -> Optional[str]:
    c = ProjectContact.query.filter_by(project_id=project_id).order_by(ProjectContact.is_default.desc(), ProjectContact.id.asc()).first()
    return c.email if c else None

def _upsert_project_contact(project_id: int, email: str, name: Optional[str] = None):
    email = (email or "").strip()
    if not email:
        return
    c = ProjectContact.query.filter_by(project_id=project_id).order_by(ProjectContact.id.asc()).first()
    if not c:
        c = ProjectContact(project_id=project_id, email=email, name=name or None, is_default=True)
        db.session.add(c)
    else:
        c.email = email
        c.name = name or c.name
        c.is_default = True

def _send_smtp_email(to_email, subject, body):
    """
    Send a simple text email. Supports both STARTTLS (587) and implicit SSL (465).
    Required env vars:
      SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASSWORD
    Optional:
      SMTP_SSL=1  (force SSL)
      SMTP_STARTTLS=0 (disable starttls for non-SSL connections)
      SMTP_FROM (defaults to SMTP_USER)
      SMTP_FROM_NAME (defaults to "EKKO NOR AS")
    """
    smtp_host = os.getenv("SMTP_HOST", "").strip()
    smtp_port = int((os.getenv("SMTP_PORT", "587") or "587").strip())
    smtp_user = os.getenv("SMTP_USER", "").strip()
    smtp_pass = os.getenv("SMTP_PASSWORD", "").strip()

    if not smtp_host or not smtp_user or not smtp_pass:
        raise RuntimeError("Brak SMTP_HOST/SMTP_USER/SMTP_PASSWORD w zmiennych środowiskowych.")

    from_email = (os.getenv("SMTP_FROM", smtp_user) or smtp_user).strip()
    from_name = (os.getenv("SMTP_FROM_NAME", "EKKO NOR AS") or "EKKO NOR AS").strip()

    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = f"{from_name} <{from_email}>"
    msg["To"] = to_email
    msg.set_content(body)

    use_ssl = os.getenv("SMTP_SSL", "").lower() in ("1", "true", "yes") or smtp_port == 465
    use_starttls = os.getenv("SMTP_STARTTLS", "1").lower() not in ("0", "false", "no")

    if use_ssl:
        server = smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=30)
    else:
        server = smtplib.SMTP(smtp_host, smtp_port, timeout=30)

    try:
        if (not use_ssl) and use_starttls:
            server.ehlo()
            server.starttls()
            server.ehlo()
        server.login(smtp_user, smtp_pass)
        server.send_message(msg)
    finally:
        try:
            server.quit()
        except Exception:
            pass


def _gen_token() -> str:
    return uuid.uuid4().hex + uuid.uuid4().hex


def _extra_report_get_decisions(report_id: int):
    try:
        return ExtraReportDecision.query.filter_by(report_id=report_id).order_by(ExtraReportDecision.decided_at.asc()).all()
    except Exception:
        return []

def _extra_report_total_minutes(rep: ExtraReport) -> int:
    if rep.total_minutes_override is not None:
        return rep.total_minutes_override
    try:
        return sum(it.minutes for it in rep.items)
    except Exception:
        return 0


class ExtraReportAttachment(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("extra_report.id"), nullable=False, index=True)
    stored_filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class ExtraReportAudit(db.Model):
    __tablename__ = "extra_report_audit"
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, index=True, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    actor_type = db.Column(db.String(20), nullable=False)  # admin, public, system
    actor_name = db.Column(db.String(120), nullable=True)
    action = db.Column(db.String(40), nullable=False)  # sent, approved, rejected, commented, auto_approved, edited
    ip = db.Column(db.String(64), nullable=True)
    user_agent = db.Column(db.String(240), nullable=True)
    details = db.Column(db.Text, nullable=True)

class ExtraReportDecision(db.Model):
    __tablename__ = "extra_report_decisions"
    id = db.Column(db.Integer, primary_key=True)
    report_id = db.Column(db.Integer, db.ForeignKey("extra_report.id"), unique=True, index=True, nullable=False)
    decided_at = db.Column(db.DateTime, default=datetime.utcnow, index=True)
    work_date = db.Column("work_date", db.Date, nullable=False, default=date.today)
    decided_name = db.Column("user_name", db.String(120), nullable=False, default="")
    decided_note = db.Column(db.Text, nullable=True)
    minutes = db.Column("minutes", db.Integer, nullable=False, default=0)
    signature_png = db.Column(db.String(260), nullable=True)  # stored filename under uploads/extra_signatures
def _auto_accept_if_due(rep: ExtraReport) -> bool:
    # Auto akceptacja po 7 dniach od wysyłki
    if rep.status == "SENT" and rep.sent_at:
        if datetime.utcnow() >= (rep.sent_at + timedelta(days=7)):
            rep.status = "APPROVED_AUTO"
            try:
                _extra_audit(rep, "auto_approved", actor_type="system", actor_name=None, details="7 days elapsed")
            except Exception:
                pass
            rep.decided_at = datetime.utcnow()
            rep.decided_note = rep.decided_note or "Auto-zaakceptowano po 7 dniach."
            db.session.commit()
            try:
                _notify_extra_report_status(rep, "auto-zaakcept (7 dni)")
            except Exception:
                pass
            return True
    return False

def _save_extra_images(req_obj: ExtraRequest, files):
    if not files:
        return
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    # max 5
    safe_files = [f for f in files if f and getattr(f, "filename", "")]
    safe_files = safe_files[:5]
    for f in safe_files:
        name = secure_filename(f.filename)
        _, ext = os.path.splitext(name)
        ext = (ext or "").lower()

        # HEIC/HEIF bez pillow-heif nie otworzymy
        if ext in {".heic", ".heif"} and not HEIF_SUPPORTED:
            flash("Ten telefon wysłał zdjęcie HEIC/HEIF. Zmień w iPhonie: Ustawienia > Aparat > Formaty > Najbardziej zgodne (JPG).")
            continue

        size = _file_size_bytes(f)
        if size is not None and size > MAX_IMAGE_BYTES:
            flash(f"Zdjęcie jest za duże ({MAX_IMAGE_MB} MB max). Zmniejsz je lub wyślij mniejsze.")
            continue

        stored = _safe_image_filename(name, req_obj.id)
        path = os.path.join(UPLOAD_DIR, stored)
        try:
            try:
                f.stream.seek(0)
            except Exception:
                pass
            _save_compressed_image(f, path)
            db.session.add(ExtraRequestImage(request_id=req_obj.id, stored_filename=stored, original_filename=name))
        except Exception:
            flash("Nie udało się zapisać jednego ze zdjęć. Spróbuj ponownie lub wyślij inne zdjęcie.")

def extra_image_view_path(stored_filename: str) -> str:
    return os.path.join(UPLOAD_DIR, stored_filename)


# --- Auth ---
@app.route("/", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))

    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        pw = request.form.get("password", "")
        user = User.query.filter_by(email=email).first()
        if user and user.check_password(pw) and user.is_active:
            login_user(user)
            return redirect(url_for("dashboard"))
        flash("Nieprawidłowy login lub hasło albo konto nieaktywne.")
        return redirect(url_for("login"))

    body = render_template_string("""
<div class="row justify-content-center">
  <div class="col-md-5">
    <div class="text-center mb-3">
      <img src="{{ url_for('static', filename='ekko_logo.png') }}" class="brand-big" alt="logo">
      <h4 class="mb-0">EKKO NOR AS</h4>
      <div class="text-muted">Rejestrator czasu pracy</div>
    </div>
    <div class="card p-3">
      <form method="post" enctype="multipart/form-data">
        <div class="mb-3">
          <label class="form-label">E-mail</label>
          <input class="form-control" type="email" name="email" placeholder="np. imie@firma.no" required>
        </div>
        <div class="mb-3">
          <label class="form-label">Hasło</label>
          <input class="form-control" type="password" name="password" required>
        </div>
        <button class="btn btn-primary w-100">Zaloguj</button>
      </form>
    </div>
  </div>
</div>
""")
    return layout("Logowanie", body)





# --- Dashboard (user) ---
@app.route("/dashboard", methods=["GET", "POST"])
@login_required
def dashboard():
    if request.method == "POST":
        work_date_str = request.form.get("work_date")
        project_id = int(request.form.get("project_id"))
        hhmm = request.form.get("hhmm", "0")
        minutes = parse_hhmm(hhmm)
        is_extra = bool(request.form.get("is_extra"))
        is_overtime = bool(request.form.get("is_overtime"))
        note = request.form.get("note") or ""
        images_files = request.files.getlist("images")


        valid_images = [f for f in images_files if f and getattr(f, 'filename', '')]
        if len(valid_images) > 5:
            flash('Możesz dodać maksymalnie 5 zdjęć do jednego wpisu.')
            return redirect(url_for('admin_entries'))


        valid_images = [f for f in images_files if f and getattr(f, 'filename', '')]
        if len(valid_images) > 5:
            flash('Możesz dodać maksymalnie 5 zdjęć do jednego wpisu.')
            return redirect(url_for('dashboard'))

        # Konwersja daty z formularza
        try:
            work_date = datetime.strptime(work_date_str, "%Y-%m-%d").date()
        except ValueError:
            flash("Nieprawidłowa data.")
            return redirect(url_for("dashboard"))

        # Ograniczenie 48h tylko dla zwykłych użytkowników (nie adminów)
        if not getattr(current_user, "is_admin", False):
            now = datetime.now()
            # Traktujemy koniec dnia roboczego jako granicę (23:59:59 danego dnia)
            end_of_work_date = datetime.combine(work_date, datetime.max.time())
            if end_of_work_date < now - timedelta(hours=48):
                flash("Godziny zostaly zablokowane poniewaz mozesz dodawac je maksymalnie do 48h skontaktuj sie z Darkiem +4746572904.")
                return redirect(url_for("dashboard"))

        e = Entry(
            user_id=current_user.id,
            project_id=project_id,
            work_date=work_date,
            minutes=minutes,
            is_extra=is_extra,
            is_overtime=is_overtime,
            note=note,
        )
        db.session.add(e)
        db.session.commit()

        # zapis zdjęć (opcjonalnie)
        try:
            _save_entry_images(e, images_files)
            db.session.commit()
        except Exception:
            # nie blokujemy dodania wpisu, jeśli zdjęcie nie zapisze się z jakiegoś powodu
            db.session.rollback()
        flash("Dodano wpis.")
        return redirect(url_for("dashboard"))

    projects = Project.query.filter_by(is_active=True).order_by(Project.name).all()
    employees = User.query.order_by(User.name).all()
    today = date.today()
    m_from, m_to = month_bounds(today)
    entries = (
        Entry.query.filter(
            Entry.user_id == current_user.id,
            Entry.work_date >= m_from,
            Entry.work_date <= m_to,
        )
        .order_by(Entry.work_date.desc(), Entry.id.desc())
        .all()
    )
    tot = sum(e.minutes for e in entries)
    tot_extra = sum(e.minutes for e in entries if e.is_extra)
    tot_ot = sum(e.minutes for e in entries if e.is_overtime)

    body = render_template_string("""
<div class="row g-3">
  <div class="col-12">
    <div class="card p-3">
      <h5 class="mb-3">Dodaj godziny</h5>
      <form id="entryForm" class="row g-2" method="post" enctype="multipart/form-data">
        <div class="col-md-3">
          <label class="form-label">Data</label>
          <input class="form-control" type="date" name="work_date" value="{{ date.today().isoformat() }}" required>
        </div>
        <div class="col-md-3">
          <label class="form-label">Projekt</label>
          <select class="form-select" name="project_id" required>
            {% for p in projects %}
              <option value="{{ p.id }}">{{ p.name }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-2">
          <label class="form-label">Czas (HH:MM)</label>
          <input class="form-control" type="text" name="hhmm" placeholder="np. 1:30" value="1:00" required>
        </div>
        <div class="col-md-2 d-flex align-items-end gap-3">
          <div class="form-check">
            <input class="form-check-input" type="checkbox" name="is_extra" id="extra">
            <label class="form-check-label" for="extra">Extra</label>
          </div>
          <div class="form-check">
            <input class="form-check-input" type="checkbox" name="is_overtime" id="ot">
            <label class="form-check-label" for="ot">Nadgodziny</label>
          </div>
        </div>
        <div class="col-md-12">
          <label class="form-label">Notatka</label>
          <input class="form-control" type="text" name="note" placeholder="opcjonalnie">
        </div>
        <div class="col-md-12">
          <label class="form-label">Zdjęcia</label>
          <input id="imagesInput" class="form-control" type="file" name="images" accept="image/*" multiple onchange="limitFiles(this,5)">
          <div class="form-text">Możesz dodać maksymalnie 5 zdjęć (z galerii albo z aparatu).</div>
        </div>

    <div id="uploadProgress" class="col-12" style="display:none;">
      <div class="progress">
        <div id="uploadBar" class="progress-bar" role="progressbar" style="width:0%" aria-valuenow="0" aria-valuemin="0" aria-valuemax="100"></div>
      </div>
      <div class="small text-muted mt-1" id="uploadText">0%</div>
    </div>

    <div id="uploadProgressAdmin" class="col-12" style="display:none;">
      <div class="progress">
        <div id="uploadBarAdmin" class="progress-bar" role="progressbar" style="width:0%" aria-valuenow="0" aria-valuemin="0" aria-valuemax="100"></div>
      </div>
      <div class="small text-muted mt-1" id="uploadTextAdmin">0%</div>
    </div>
        <div class="col-12">
          <button class="btn btn-primary">Zapisz</button>
        </div>
      </form>
    </div>
  </div>

  <div class="col-12">
    <div class="card p-3">
      <div class="d-flex justify-content-between align-items-center">
        <h5 class="mb-0">Moje wpisy – {{ m_from.isoformat() }} → {{ m_to.isoformat() }}</h5>
      </div>
      <div class="table-responsive mt-3">
        <table class="table table-sm align-middle">
          <thead>
            <tr><th>Data</th><th>Projekt</th><th>Notatka</th><th>Zdjęcia</th><th>Godziny</th><th>Extra</th><th>OT</th><th class="text-end">Akcje</th></tr>
          </thead>
          <tbody>
            {% for e in entries %}
            <tr>
              <td>{{ e.work_date.isoformat() }}</td>
              <td>{{ e.project.name }}</td>
              <td>{{ e.note or '' }}</td>
              <td>
                {% if e.images %}
                  {% for img in e.images %}
                    <a href="{{ url_for('entry_image_view', image_id=img.id) }}" target="_blank" rel="noopener">IMG</a>{% if not loop.last %} {% endif %}
                  {% endfor %}
                  </div>
                {% else %}-{% endif %}
              </td>
              <td>{{ fmt(e.minutes) }}</td>
              <td>{% if e.is_extra %}✔{% else %}-{% endif %}</td>
              <td>{% if e.is_overtime %}✔{% else %}-{% endif %}</td>
              <td class="text-end">
                <a class="btn btn-sm btn-outline-primary" href="{{ url_for('edit_entry', entry_id=e.id) }}">Edytuj</a>
                <form class="d-inline" method="post" action="{{ url_for('delete_entry', entry_id=e.id) }}" onsubmit="return confirm('Usunąć wpis?')">
                  <button class="btn btn-sm btn-outline-danger">Usuń</button>
                </form>
              </td>
            </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
      <div class="mt-2">
        <span class="me-3">Razem: <strong>{{ fmt(tot) }}</strong></span>
        <span class="me-3">Extra: <strong>{{ fmt(tot_extra) }}</strong></span>
        <span class="me-3">Nadgodziny: <strong>{{ fmt(tot_ot) }}</strong></span>
      </div>
    </div>
  </div>

<script>
function limitFiles(input, max){
  if (!input || !input.files) return;
  if (input.files.length > max) {
    alert('Możesz dodać maksymalnie ' + max + ' zdjęć do jednego wpisu.');
    input.value = '';
  }
}

function wireUploadProgress(formId, progressId, barId, textId){
  const form = document.getElementById(formId);
  if (!form) return;

  form.addEventListener('submit', function(e){
    // jeśli brak plików, nie ma sensu AJAXować (szybszy normalny submit)
    const fileInput = form.querySelector('input[type="file"][name="images"]');
    if (!fileInput || !fileInput.files || fileInput.files.length === 0) {
      return; // normalny submit
    }

    e.preventDefault();

    const progressBox = document.getElementById(progressId);
    const bar = document.getElementById(barId);
    const text = document.getElementById(textId);

    progressBox.style.display = 'block';
    bar.style.width = '0%';
    bar.setAttribute('aria-valuenow', '0');
    if (text) text.textContent = '0%';

    const xhr = new XMLHttpRequest();
    xhr.open('POST', form.getAttribute('action') || window.location.href);

    xhr.upload.onprogress = function(evt){
      if (evt.lengthComputable) {
        const percent = Math.round((evt.loaded / evt.total) * 100);
        bar.style.width = percent + '%';
        bar.setAttribute('aria-valuenow', String(percent));
        if (text) text.textContent = percent + '%';
      }
    };

    xhr.onload = function(){
      // Po udanym zapisie, odświeżamy stronę (żeby pokazać flash i nowy wpis)
      window.location.reload();
    };

    xhr.onerror = function(){
      alert('Błąd podczas wysyłania. Spróbuj ponownie.');
      progressBox.style.display = 'none';
    };

    xhr.send(new FormData(form));
  });
}

document.addEventListener('DOMContentLoaded', function(){
  wireUploadProgress('entryForm','uploadProgress','uploadBar','uploadText');
  wireUploadProgress('adminEntryForm','uploadProgressAdmin','uploadBarAdmin','uploadTextAdmin');
});
</script>
</div>
""", projects=projects, entries=entries, fmt=fmt_hhmm, m_from=m_from, m_to=m_to, tot=tot, tot_extra=tot_extra, tot_ot=tot_ot, date=date)
    return layout("Panel", body)


# --- Plany (PDF) ---
@app.route("/plans", methods=["GET"])
@login_required
def plans():
    # Lista planów, z filtrem po projekcie
    projects = Project.query.order_by(Project.is_active.desc(), Project.name.asc()).all()
    selected_pid = request.args.get("project_id", "all")
    selected_pid_int = int(selected_pid) if str(selected_pid).isdigit() else 0

    q = Plan.query.join(Project).order_by(Plan.uploaded_at.desc(), Plan.id.desc())
    if selected_pid != "all":
        try:
            q = q.filter(Plan.project_id == int(selected_pid))
        except Exception:
            selected_pid = "all"

    rows = q.all()

    body = render_template_string("""
<div class="card p-3">
  <div class="d-flex justify-content-between align-items-center mb-2">
    <h5 class="mb-0">Plany (PDF)</h5>
  </div>

  <form class="row g-2 align-items-end mb-3" method="get">
    <div class="col-md-5">
      <label class="form-label">Projekt</label>
      <select class="form-select" name="project_id">
        <option value="all" {% if selected_pid == 'all' %}selected{% endif %}>Wszystkie projekty</option>
        {% for p in projects %}
          <option value="{{ p.id }}" {% if selected_pid|int == p.id %}selected{% endif %}>{{ p.name }}{% if not p.is_active %} (nieaktywny){% endif %}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-2">
      <button class="btn btn-outline-primary w-100">Filtruj</button>
    </div>
  </form>

  <div class="table-responsive">
    <table class="table table-sm align-middle">
      <thead>
        <tr>
          <th>Projekt</th>
          <th>Tytuł</th>
          <th>Dodano</th>
          <th class="text-end">PDF</th>
        </tr>
      </thead>
      <tbody>
        {% for pl in rows %}
          <tr>
            <td>{{ pl.project.name }}</td>
            <td>{{ pl.title or (pl.original_filename or 'Plan') }}</td>
            <td>{{ pl.uploaded_at.strftime("%Y-%m-%d %H:%M") if pl.uploaded_at else '' }}</td>
            <td class="text-end">
              <a class="btn btn-sm btn-outline-primary" target="_blank" rel="noopener"
                 href="{{ url_for('plan_view', plan_id=pl.id) }}">Otwórz</a>
            </td>
          </tr>
        {% else %}
          <tr><td colspan="4" class="text-muted">Brak planów.</td></tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</div>
""", projects=projects, rows=rows, selected_pid=selected_pid)

    return layout("Plany", body)


@app.route("/plans/<int:plan_id>/view", methods=["GET"])
@login_required
def plan_view(plan_id):
    pl = Plan.query.get_or_404(plan_id)
    path = os.path.join(PLANS_DIR, pl.stored_filename)
    if not os.path.exists(path):
        abort(404)
    return send_file(path, mimetype="application/pdf")


@app.route("/admin/plans", methods=["GET", "POST"])
@login_required
def admin_plans():
    require_admin()
    os.makedirs(PLANS_DIR, exist_ok=True)

    if request.method == "POST":
        project_id = int(request.form.get("project_id"))
        title = (request.form.get("title") or "").strip()
        category = (request.form.get("category") or "annet").strip() or "annet"
        files = request.files.getlist("pdfs")

        if not files or not any(getattr(x, "filename", "") for x in files):
            flash("Wybierz plik PDF.")
            return redirect(url_for("admin_plans"))

        count_added = 0
        for f in files:
            if not f or not getattr(f, "filename", ""):
                continue

            name = secure_filename(f.filename)
            _, ext = os.path.splitext(name)
            ext = (ext or "").lower()
            if ext not in ALLOWED_PLAN_EXTS:
                continue

            size = _file_size_bytes(f)
            if size is not None and size > MAX_PLAN_BYTES:
                continue

            stored = _safe_plan_filename(name, project_id)
            out_path = os.path.join(PLANS_DIR, stored)

            try:
                try:
                    f.stream.seek(0)
                except Exception:
                    pass
                f.save(out_path)
            except Exception:
                continue

            db.session.add(Plan(
                project_id=project_id,
                title=title or None,
                stored_filename=stored,
                original_filename=name,
                uploaded_by=current_user.id,
            ))
            count_added += 1

        if count_added == 0:
            flash("Nie dodano żadnego planu. Upewnij się, że wybierasz pliki PDF i mieszczą się w limicie.")
            return redirect(url_for("admin_plans"))

        db.session.commit()
        flash(f"Dodano plany: {count_added}")
        return redirect(url_for("admin_plans"))


    projects = Project.query.order_by(Project.is_active.desc(), Project.name.asc()).all()

    selected_pid = request.args.get("project_id", "all")
    q = Plan.query.join(Project).order_by(Plan.uploaded_at.desc(), Plan.id.desc())
    if selected_pid != "all":
        try:
            q = q.filter(Plan.project_id == int(selected_pid))
        except Exception:
            selected_pid = "all"

    rows = q.all()

    body = render_template_string("""
<div class="row g-3">
  <div class="col-12">
    <div class="card p-3">
      <h5 class="mb-3">Plany – dodaj PDF</h5>
      <form class="row g-2" method="post" enctype="multipart/form-data">
        <div class="col-md-4">
          <label class="form-label">Projekt</label>
          <select class="form-select" name="project_id" required>
            {% for p in projects %}
              <option value="{{ p.id }}">{{ p.name }}{% if not p.is_active %} (nieaktywny){% endif %}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-4">
          <label class="form-label">Tytuł (opcjonalnie)</label>
          <input class="form-control" name="title" placeholder="np. Rysunki – wersja 2">
        </div>
        <div class="col-md-4">
          <label class="form-label">PDF (możesz wybrać kilka)</label>
          <input class="form-control" type="file" name="pdfs" accept="application/pdf" multiple required>
          <div class="form-text">Maks {{ max_mb }} MB.</div>
        </div>
        <div class="col-12">
          <button class="btn btn-primary">Dodaj</button>
        </div>
      </form>
    </div>
  </div>

  <div class="col-12">
    <div class="card p-3">
      <div class="d-flex justify-content-between align-items-center mb-2">
        <h5 class="mb-0">Lista planów</h5>
      </div>

      <form class="row g-2 align-items-end mb-3" method="get">
        <div class="col-md-5">
          <label class="form-label">Projekt</label>
          <select class="form-select" name="project_id">
            <option value="all" {% if selected_pid == 'all' %}selected{% endif %}>Wszystkie projekty</option>
            {% for p in projects %}
              <option value="{{ p.id }}" {% if selected_pid|int == p.id %}selected{% endif %}>{{ p.name }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-2">
          <button class="btn btn-outline-primary w-100">Filtruj</button>
        </div>
      </form>

      <div class="table-responsive">
        <table class="table table-sm align-middle">
          <thead>
            <tr>
              <th>Projekt</th>
              <th>Tytuł</th>
              <th>Plik</th>
              <th>Dodano</th>
              <th class="text-end">Akcje</th>
            </tr>
          </thead>
          <tbody>
            {% for pl in rows %}
              <tr>
                <td>{{ pl.project.name }}</td>
                <td>{{ pl.title or '' }}</td>
                <td>{{ pl.original_filename or pl.stored_filename }}</td>
                <td>{{ pl.uploaded_at.strftime("%Y-%m-%d %H:%M") if pl.uploaded_at else '' }}</td>
                <td class="text-end text-nowrap">
                  <a class="btn btn-sm btn-outline-primary" target="_blank" rel="noopener"
                     href="{{ url_for('plan_view', plan_id=pl.id) }}">Otwórz</a>
                  <form class="d-inline" method="post" action="{{ url_for('admin_plan_delete', plan_id=pl.id) }}"
                        onsubmit="return confirm('Usunąć plan?')">
                    <button class="btn btn-sm btn-outline-danger">Usuń</button>
                  </form>
                </td>
              </tr>
            {% else %}
              <tr><td colspan="5" class="text-muted">Brak planów.</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>

    </div>
  </div>
</div>
""", projects=projects, rows=rows, selected_pid=selected_pid, max_mb=MAX_PLAN_MB)

    return layout("Plany (admin)", body)


@app.route("/admin/plans/<int:plan_id>/delete", methods=["POST"])
@login_required
def admin_plan_delete(plan_id):
    require_admin()
    pl = Plan.query.get_or_404(plan_id)

    # usuń plik
    try:
        path = os.path.join(PLANS_DIR, pl.stored_filename)
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        pass

    db.session.delete(pl)
    db.session.commit()
    flash("Usunięto plan.")
    return redirect(url_for("admin_plans"))





# --- Edit/Delete entries (user & admin) ---
@app.route("/entry/<int:entry_id>/edit", methods=["GET", "POST"])
@login_required
def edit_entry(entry_id):
    e = Entry.query.get_or_404(entry_id)
    if not (current_user.is_admin or e.user_id == current_user.id):
        abort(403)

    if request.method == "POST":
        e.work_date = datetime.strptime(request.form.get("work_date"), "%Y-%m-%d").date()
        e.project_id = int(request.form.get("project_id"))
        e.minutes = parse_hhmm(request.form.get("hhmm", "0"))
        e.is_extra = bool(request.form.get("is_extra"))
        e.is_overtime = bool(request.form.get("is_overtime"))
        e.note = request.form.get("note") or ""
        db.session.commit()
        flash("Zapisano zmiany.")
        return redirect(url_for("dashboard" if not current_user.is_admin else "admin_entries"))

    projects = Project.query.filter_by(is_active=True).order_by(Project.name).all()
    hhmm_value = fmt_hhmm(e.minutes)
    body = render_template_string("""
<div class="card p-3">
  <h5 class="mb-3">Edytuj wpis</h5>
  <form id="adminEntryForm" class="row g-2" method="post" enctype="multipart/form-data">
    <div class="col-md-3">
      <label class="form-label">Data</label>
      <input class="form-control" type="date" name="work_date" value="{{ e.work_date.isoformat() }}" required>
    </div>
    <div class="col-md-3">
      <label class="form-label">Projekt</label>
      <select class="form-select" name="project_id">
        {% for p in projects %}
          <option value="{{ p.id }}" {% if p.id == e.project_id %}selected{% endif %}>{{ p.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-2">
      <label class="form-label">Czas (HH:MM)</label>
      <input class="form-control" type="text" name="hhmm" value="{{ hhmm_value }}" required>
    </div>
    <div class="col-md-2 d-flex align-items-end gap-3">
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="is_extra" id="extra" {% if e.is_extra %}checked{% endif %}>
        <label class="form-check-label" for="extra">Extra</label>
      </div>
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="is_overtime" id="ot" {% if e.is_overtime %}checked{% endif %}>
        <label class="form-check-label" for="ot">Nadgodziny</label>
      </div>
    </div>
    <div class="col-12">
      <label class="form-label">Notatka</label>
      <input class="form-control" type="text" name="note" value="{{ e.note or '' }}">
    </div>
    <div class="col-12">
      <button class="btn btn-primary">Zapisz</button>
      <a class="btn btn-outline-secondary" href="{{ url_for('dashboard') }}">Anuluj</a>
    </div>
  </form>
</div>
""", e=e, projects=projects, hhmm_value=hhmm_value)
    return layout("Edytuj wpis", body)


@app.route("/entry/<int:entry_id>/delete", methods=["POST"])
@login_required
def delete_entry(entry_id):
    e = Entry.query.get_or_404(entry_id)
    if not (current_user.is_admin or e.user_id == current_user.id):
        abort(403)
    _delete_entry_images_files(e)
    db.session.delete(e)
    db.session.commit()
    flash("Usunięto wpis.")
    return redirect(url_for("dashboard" if not current_user.is_admin else "admin_entries"))


@app.route("/image/<int:image_id>")
@login_required
def entry_image_view(image_id):
    img = EntryImage.query.get_or_404(image_id)
    e = Entry.query.get(img.entry_id)
    if not e:
        abort(404)
    if not (current_user.is_admin or e.user_id == current_user.id):
        abort(403)
    path = os.path.join(UPLOAD_DIR, img.stored_filename)
    if not os.path.exists(path):
        abort(404)
    return send_file(path)


# --- Admin: overview (monthly totals) ---
@app.route("/admin", methods=["GET"])
@login_required
def admin_overview():
    require_admin()
    ym = request.args.get("month")
    if not ym:
        today = date.today()
        ym = f"{today.year:04d}-{today.month:02d}"
    year, month = map(int, ym.split("-"))
    m_from = date(year, month, 1)
    m_to = (m_from.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    # poprzedni miesiąc
    if month == 1:
        prev_year, prev_month = year - 1, 12
    else:
        prev_year, prev_month = year, month - 1
    prev_from = date(prev_year, prev_month, 1)
    prev_to = (prev_from.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    total = db.session.query(db.func.sum(Entry.minutes)).filter(
        Entry.work_date >= m_from, Entry.work_date <= m_to
    ).scalar() or 0

    users = User.query.filter_by(is_active_u=True).order_by(User.name).all()
    stats = []
    for u in users:
        curr_min = db.session.query(db.func.sum(Entry.minutes)).filter(
            Entry.user_id == u.id,
            Entry.work_date >= m_from,
            Entry.work_date <= m_to,
        ).scalar() or 0
        prev_min = db.session.query(db.func.sum(Entry.minutes)).filter(
            Entry.user_id == u.id,
            Entry.work_date >= prev_from,
            Entry.work_date <= prev_to,
        ).scalar() or 0
        stats.append({"user": u, "curr": curr_min, "prev": prev_min})

    body = render_template_string("""
<div class="card p-3 mb-3">
  <h5 class="mb-3">Podsumowanie miesiąca</h5>
  <form class="row g-2 mb-3" method="get">
    <div class="col-md-3">
      <label class="form-label">Miesiąc</label>
      <input class="form-control" type="month" name="month" value="{{ ym }}">
    </div>
    <div class="col-md-2 d-flex align-items-end">
      <button class="btn btn-outline-primary">Pokaż</button>
    </div>
  </form>
  <div class="display-6">{{ fmt(total) }}</div>
  <div class="text-muted">Łącznie zapisanych godzin w wybranym miesiącu</div>
</div>

<div class="card p-3">
  <h5 class="mb-3">Godziny pracowników</h5>
  <p class="small text-muted">
    Bieżący miesiąc: {{ ym }} &nbsp;&nbsp;|&nbsp;&nbsp;
    Poprzedni miesiąc: {{ prev_label }}
  </p>
  <div class="table-responsive">
    <table class="table table-sm table-striped align-middle">
      <thead>
        <tr>
          <th>Pracownik</th>
          <th>Godziny w tym miesiącu</th>
          <th>Godziny w poprzednim miesiącu</th>
        </tr>
      </thead>
      <tbody>
      {% for row in stats %}
        <tr>
          <td>{{ row.user.name }}</td>
          <td>{{ fmt(row.curr) }}</td>
          <td>{{ fmt(row.prev) }}</td>
        </tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
</div>
""", ym=ym, total=total, stats=stats, fmt=fmt_hhmm,
       prev_label=f"{prev_year:04d}-{prev_month:02d}")
    return layout("Admin", body)


# --- Admin: users ---
@app.route("/admin/users", methods=["GET", "POST"])
@login_required
def admin_users():
    require_admin()

    if request.method == "POST" and request.form.get("action") == "create":
        name = request.form.get("name","").strip()
        email = (request.form.get("email","") or "").strip().lower()
        password = request.form.get("password","")
        is_admin = bool(request.form.get("is_admin"))
        if name and email and password:
            if not User.query.filter_by(email=email).first():
                u = User(name=name, email=email, is_admin=is_admin, is_active_u=True)
                u.set_password(password)
                db.session.add(u)
                db.session.commit()
                flash("Dodano pracownika.")
            else:
                flash("Taki e-mail już istnieje.")
        else:
            flash("Uzupełnij imię, e-mail i hasło.")
        return redirect(url_for("admin_users"))

    users = User.query.order_by(User.name).all()
    body = render_template_string("""
<div class="card p-3">
  <h5>Pracownicy</h5>

  <form class="row g-2 mb-3" method="post">
    <input type="hidden" name="action" value="create">
    <div class="col-md-3"><input class="form-control" name="name" placeholder="Imię i nazwisko" required></div>
    <div class="col-md-3"><input class="form-control" type="email" name="email" placeholder="E-mail" required></div>
    <div class="col-md-3"><input class="form-control" type="text" name="password" placeholder="Hasło startowe" required></div>
    <div class="col-md-2 d-flex align-items-center gap-2">
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="is_admin" id="isadmin">
        <label class="form-check-label" for="isadmin">Admin</label>
      </div>
    </div>
    <div class="col-md-1"><button class="btn btn-primary w-100">Dodaj</button></div>
  </form>

  <div class="table-responsive">
    <table class="table table-sm align-middle">
      <thead><tr><th>Imię i nazwisko</th><th>E-mail</th><th>Rola</th><th>Status</th><th>Akcje</th></tr></thead>
      <tbody>
        {% for u in users %}
        <tr>
          <td>{{ u.name }}</td>
          <td>{{ u.email }}</td>
          <td>{% if u.is_admin %}Admin{% else %}Użytkownik{% endif %}</td>
          <td>{% if u.is_active %}Aktywny{% else %}Nieaktywny{% endif %}</td>
          <td class="text-nowrap">
            <a class="btn btn-sm btn-outline-primary" href="{{ url_for('admin_user_edit', uid=u.id) }}">Edytuj</a>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</div>
""", users=users)
    return layout("Pracownicy", body)


@app.route("/admin/users/<int:uid>", methods=["GET", "POST"])
@login_required
def admin_user_edit(uid):
    require_admin()
    u = User.query.get_or_404(uid)

    if request.method == "POST":
        action = request.form.get("action")
        if action == "save":
            u.name = request.form.get("name","").strip()
            u.email = request.form.get("email","").strip().lower()
            u.is_admin = bool(request.form.get("is_admin"))
            u.is_active_u = bool(request.form.get("is_active"))
            db.session.commit()
            flash("Zapisano.")
            return redirect(url_for("admin_users"))
        elif action == "set_password":
            pw = request.form.get("password","")
            if pw:
                u.set_password(pw)
                db.session.commit()
                flash("Zmieniono hasło.")
            else:
                flash("Hasło nie może być puste.")
            return redirect(url_for("admin_user_edit", uid=u.id))

    body = render_template_string("""
<div class="card p-3">
  <h5>Edycja pracownika</h5>
  <form class="row g-2 mb-3" method="post">
    <input type="hidden" name="action" value="save">
    <div class="col-md-4"><label class="form-label">Imię i nazwisko</label>
      <input class="form-control" name="name" value="{{ u.name }}"></div>
    <div class="col-md-4"><label class="form-label">E-mail</label>
      <input class="form-control" type="email" name="email" value="{{ u.email }}"></div>
    <div class="col-md-4 d-flex align-items-end gap-3">
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="is_admin" id="ea" {% if u.is_admin %}checked{% endif %}>
        <label class="form-check-label" for="ea">Admin</label>
      </div>
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="is_active" id="eact" {% if u.is_active %}checked{% endif %}>
        <label class="form-check-label" for="eact">Aktywny</label>
      </div>
      <button class="btn btn-primary ms-auto">Zapisz</button>
    </div>
  </form>

  <h6>Reset hasła</h6>
  <form class="row g-2" method="post" enctype="multipart/form-data">
    <input type="hidden" name="action" value="set_password">
    <div class="col-md-4"><input class="form-control" type="text" name="password" placeholder="Nowe hasło"></div>
    <div class="col-md-2"><button class="btn btn-outline-primary">Ustaw hasło</button></div>
  </form>
</div>
""", u=u)
    return layout("Edycja pracownika", body)


# --- Admin: projects (CRUD) ---
@app.route("/admin/projects", methods=["GET", "POST"])
@login_required
def admin_projects():
    require_admin()
    if request.method == "POST" and request.form.get("action") == "create":
        name = request.form.get("name").strip()
        if not name:
            flash("Nazwa nie może być pusta.")
        elif Project.query.filter_by(name=name).first():
            flash("Projekt o takiej nazwie już istnieje.")
        else:
            db.session.add(Project(name=name, is_active=True))
            db.session.commit()
            flash("Dodano projekt.")
        return redirect(url_for("admin_projects"))

    projs = Project.query.order_by(Project.is_active.desc(), Project.name.asc()).all()
    body = render_template_string("""
<div class="card p-3">
  <div class="d-flex justify-content-between align-items-center">
    <h5 class="mb-0">Projekty</h5>
    <form class="d-flex" method="post">
      <input type="hidden" name="action" value="create">
      <input class="form-control me-2" name="name" placeholder="Nazwa projektu" required>
      <button class="btn btn-primary">Dodaj</button>
    </form>
  </div>
  <div class="table-responsive mt-3">
    <table class="table table-sm align-middle">
      <thead><tr><th style="width:55%">Nazwa</th><th style="width:20%">Aktywny</th><th class="text-end" style="width:25%">Akcje</th></tr></thead>
      <tbody>
      {% for p in projs %}
        <tr>
          <td>
            <form class="d-flex" method="post" action="{{ url_for('admin_project_update', pid=p.id) }}">
              <input class="form-control form-control-sm me-2" name="name" value="{{ p.name }}" required>
              <button class="btn btn-sm btn-outline-primary">Zmień nazwę</button>
            </form>
          </td>
          <td>
            <form method="post" action="{{ url_for('admin_project_toggle', pid=p.id) }}">
              <input type="hidden" name="is_active" value="{{ 1 if not p.is_active else 0 }}">
              <button class="btn btn-sm {% if p.is_active %}btn-success{% else %}btn-secondary{% endif %}">
                {% if p.is_active %}Aktywny{% else %}Nieaktywny{% endif %}
              </button>
            </form>
          </td>
          <td class="text-end">
            <form class="d-inline" method="post" action="{{ url_for('admin_project_delete', pid=p.id) }}" onsubmit="return confirm('Usunąć projekt? (wpisy pozostaną)')">
              <button class="btn btn-sm btn-outline-danger">Usuń</button>
            </form>
          </td>
        </tr>
      {% endfor %}
      </tbody>
    </table>
  </div>
</div>
""", projs=projs)
    return layout("Projekty", body)

@app.route("/admin/projects/<int:pid>/update", methods=["POST"])
@login_required
def admin_project_update(pid):
    require_admin()
    p = Project.query.get_or_404(pid)
    new_name = (request.form.get("name") or "").strip()
    if not new_name:
        flash("Nazwa nie może być pusta.")
    elif Project.query.filter(Project.id != pid, Project.name == new_name).first():
        flash("Projekt o takiej nazwie już istnieje.")
    else:
        p.name = new_name
        db.session.commit()
        flash("Zmieniono nazwę projektu.")
    return redirect(url_for("admin_projects"))

@app.route("/admin/projects/<int:pid>/toggle", methods=["POST"])
@login_required
def admin_project_toggle(pid):
    require_admin()
    p = Project.query.get_or_404(pid)
    want_active = request.form.get("is_active")
    if want_active is not None:
        p.is_active = True if str(want_active) == "1" else False
    else:
        p.is_active = not p.is_active
    db.session.commit()
    return redirect(url_for("admin_projects"))

@app.route("/admin/projects/<int:pid>/delete", methods=["POST"])
@login_required
def admin_project_delete(pid):
    require_admin()
    p = Project.query.get_or_404(pid)
    db.session.delete(p)
    db.session.commit()
    flash("Usunięto projekt.")
    return redirect(url_for("admin_projects"))


# --- Admin: entries (full add/edit/delete + filter) ---
def _all_users_ordered():
    return User.query.order_by(User.name.asc()).all()

@app.route("/admin/entries", methods=["GET", "POST"])
@login_required
def admin_entries():
    require_admin()

    if request.method == "POST":
        uid = int(request.form.get("user_id"))
        pid = int(request.form.get("project_id"))
        work_date = datetime.strptime(request.form.get("work_date"), "%Y-%m-%d").date()
        minutes = parse_hhmm(request.form.get("hhmm", "0"))
        is_extra = bool(request.form.get("is_extra"))
        is_ot = bool(request.form.get("is_overtime"))
        note = request.form.get("note") or ""
        images_files = request.files.getlist("images")

        e = Entry(
            user_id=uid, project_id=pid, work_date=work_date,
            minutes=minutes, is_extra=is_extra, is_overtime=is_ot, note=note
        )
        db.session.add(e)
        db.session.commit()

        # zapis zdjęć (opcjonalnie)
        try:
            _save_entry_images(e, images_files)
            db.session.commit()
        except Exception:
            db.session.rollback()
        flash("Dodano wpis.")
        return redirect(url_for("admin_entries"))

    ym = request.args.get("month")
    if not ym:
        today = date.today()
        ym = f"{today.year:04d}-{today.month:02d}"
    year, month = map(int, ym.split("-"))
    m_from = date(year, month, 1)
    m_to = (m_from.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)

    selected_uid = request.args.get("user_id", "all")
    q = Entry.query.join(User).join(Project).filter(
        and_(Entry.work_date >= m_from, Entry.work_date <= m_to)
    )
    if selected_uid != "all":
        q = q.filter(Entry.user_id == int(selected_uid))

    entries = q.order_by(Entry.work_date.desc(), Entry.id.desc()).all()
    users = _all_users_ordered()
    projects = Project.query.order_by(Project.name).all()

    tot = sum(e.minutes for e in entries)
    tot_ex = sum(e.minutes for e in entries if e.is_extra)
    tot_ot = sum(e.minutes for e in entries if e.is_overtime)

    body = render_template_string("""
<div class="card p-3">
  <h5 class="mb-3">Dodaj godziny (admin)</h5>
  <form class="row g-2" method="post" enctype="multipart/form-data">
    <div class="col-md-3">
      <label class="form-label">Pracownik</label>
      <select class="form-select" name="user_id" required>
        {% for u in users %}<option value="{{ u.id }}">{{ u.name }} ({{ u.email }})</option>{% endfor %}
      </select>
    </div>
    <div class="col-md-3">
      <label class="form-label">Projekt</label>
      <select class="form-select" name="project_id" required>
        {% for p in projects %}<option value="{{ p.id }}">{{ p.name }}</option>{% endfor %}
      </select>
    </div>
    <div class="col-md-2">
      <label class="form-label">Data</label>
      <input class="form-control" type="date" name="work_date" value="{{ date.today().isoformat() }}" required>
    </div>
    <div class="col-md-2">
      <label class="form-label">Czas (HH:MM)</label>
      <input class="form-control" type="text" name="hhmm" placeholder="np. 1:30" value="1:00" required>
    </div>
    <div class="col-md-2 d-flex align-items-end gap-3">
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="is_extra" id="aextra">
        <label class="form-check-label" for="aextra">Extra</label>
      </div>
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="is_overtime" id="aot">
        <label class="form-check-label" for="aot">Nadgodziny</label>
      </div>
    </div>
    <div class="col-12">
      <label class="form-label">Notatka</label>
      <input class="form-control" type="text" name="note" placeholder="opcjonalnie">
    </div>
    <div class="col-12">
      <label class="form-label">Zdjęcia</label>
      <input id="adminImagesInput" class="form-control" type="file" name="images" accept="image/*" multiple onchange="limitFiles(this,5)">
      <div class="form-text">Opcjonalne zdjęcia do wpisu (maksymalnie 5).</div>
    </div>
    <div class="col-12">
      <button class="btn btn-primary">Zapisz</button>
    </div>
  </form>
</div>

<div class="card p-3 mt-3">
  <form class="row g-2 align-items-end" method="get">
    <div class="col-md-3">
      <label class="form-label">Miesiąc</label>
      <input class="form-control" type="month" name="month" value="{{ ym }}">
    </div>
    <div class="col-md-5">
      <label class="form-label">Pracownik</label>
      <select class="form-select" name="user_id">
        <option value="all" {% if selected_uid == 'all' %}selected{% endif %}>Wszyscy</option>
        {% for u in users %}
          <option value="{{ u.id }}" {% if selected_uid|int == u.id %}selected{% endif %}>{{ u.name }} ({{ u.email }})</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-2">
      <button class="btn btn-outline-primary w-100">Filtruj</button>
    </div>
  </form>

  <div class="table-responsive mt-3">
    <table class="table table-sm align-middle">
      <thead>
        <tr><th>Data</th><th>Pracownik</th><th>Projekt</th><th>Notatka</th><th>Zdjęcia</th><th>Godziny</th><th>Extra</th><th>OT</th><th></th></tr>
      </thead>
      <tbody>
        {% for e in entries %}
        <tr>
          <td>{{ e.work_date.isoformat() }}</td>
          <td>{{ e.user.name }}</td>
          <td>{{ e.project.name }}</td>
          <td>{{ e.note or '' }}</td>
          <td>
            {% if e.images %}
              {% for img in e.images %}
                <a href="{{ url_for('entry_image_view', image_id=img.id) }}" target="_blank" rel="noopener">IMG</a>{% if not loop.last %} {% endif %}
              {% endfor %}
            {% else %}-{% endif %}
          </td>
          <td>{{ fmt(e.minutes) }}</td>
          <td>{% if e.is_extra %}✔{% else %}-{% endif %}</td>
          <td>{% if e.is_overtime %}✔{% else %}-{% endif %}</td>
          <td class="text-nowrap">
            <a class="btn btn-sm btn-outline-primary" href="{{ url_for('admin_entry_edit', entry_id=e.id) }}">Edytuj</a>
            <form class="d-inline" method="post" action="{{ url_for('admin_entry_delete', entry_id=e.id) }}" onsubmit="return confirm('Usunąć wpis?')">
              <button class="btn btn-sm btn-outline-danger">Usuń</button>
            </form>
          </td>
        </tr>
        {% endfor %}
      </tbody>
    </table>
  </div>

  <div class="mt-2">
    <span class="me-3">Razem: <strong>{{ fmt(tot) }}</strong></span>
    <span class="me-3">Extra: <strong>{{ fmt(tot_ex) }}</strong></span>
    <span class="me-3">Nadgodziny: <strong>{{ fmt(tot_ot) }}</strong></span>
  </div>
</div>

<script>
function limitFiles(input, max){
  if (!input || !input.files) return;
  if (input.files.length > max) {
    alert('Możesz dodać maksymalnie ' + max + ' zdjęć do jednego wpisu.');
    input.value = '';
  }
}

function wireUploadProgress(formId, progressId, barId, textId){
  const form = document.getElementById(formId);
  if (!form) return;

  form.addEventListener('submit', function(e){
    // jeśli brak plików, nie ma sensu AJAXować (szybszy normalny submit)
    const fileInput = form.querySelector('input[type="file"][name="images"]');
    if (!fileInput || !fileInput.files || fileInput.files.length === 0) {
      return; // normalny submit
    }

    e.preventDefault();

    const progressBox = document.getElementById(progressId);
    const bar = document.getElementById(barId);
    const text = document.getElementById(textId);

    progressBox.style.display = 'block';
    bar.style.width = '0%';
    bar.setAttribute('aria-valuenow', '0');
    if (text) text.textContent = '0%';

    const xhr = new XMLHttpRequest();
    xhr.open('POST', form.getAttribute('action') || window.location.href);

    xhr.upload.onprogress = function(evt){
      if (evt.lengthComputable) {
        const percent = Math.round((evt.loaded / evt.total) * 100);
        bar.style.width = percent + '%';
        bar.setAttribute('aria-valuenow', String(percent));
        if (text) text.textContent = percent + '%';
      }
    };

    xhr.onload = function(){
      // Po udanym zapisie, odświeżamy stronę (żeby pokazać flash i nowy wpis)
      window.location.reload();
    };

    xhr.onerror = function(){
      alert('Błąd podczas wysyłania. Spróbuj ponownie.');
      progressBox.style.display = 'none';
    };

    xhr.send(new FormData(form));
  });
}

document.addEventListener('DOMContentLoaded', function(){
  wireUploadProgress('entryForm','uploadProgress','uploadBar','uploadText');
  wireUploadProgress('adminEntryForm','uploadProgressAdmin','uploadBarAdmin','uploadTextAdmin');
});
</script>
""", users=users, projects=projects, entries=entries, fmt=fmt_hhmm,
       ym=ym, selected_uid=selected_uid, tot=tot, tot_ex=tot_ex, tot_ot=tot_ot, date=date)
    return layout("Godziny (admin)", body)

@app.route("/admin/entries/<int:entry_id>/edit", methods=["GET", "POST"])
@login_required
def admin_entry_edit(entry_id):
    require_admin()
    e = Entry.query.get_or_404(entry_id)
    users = _all_users_ordered()
    projects = Project.query.order_by(Project.name).all()

    if request.method == "POST":
        e.user_id = int(request.form.get("user_id"))
        e.project_id = int(request.form.get("project_id"))
        e.work_date = datetime.strptime(request.form.get("work_date"), "%Y-%m-%d").date()
        e.minutes = parse_hhmm(request.form.get("hhmm", "0"))
        e.is_extra = bool(request.form.get("is_extra"))
        e.is_overtime = bool(request.form.get("is_overtime"))
        e.note = request.form.get("note") or ""
        db.session.commit()
        flash("Zapisano zmiany.")
        return redirect(url_for("admin_entries"))

    body = render_template_string("""
<div class="card p-3">
  <h5 class="mb-3">Edytuj wpis</h5>
  <form class="row g-2" method="post">
    <div class="col-md-3">
      <label class="form-label">Pracownik</label>
      <select class="form-select" name="user_id" required>
        {% for u in users %}<option value="{{ u.id }}" {% if u.id == e.user_id %}selected{% endif %}>{{ u.name }}</option>{% endfor %}
      </select>
    </div>
    <div class="col-md-3">
      <label class="form-label">Projekt</label>
      <select class="form-select" name="project_id" required>
        {% for p in projects %}<option value="{{ p.id }}" {% if p.id == e.project_id %}selected{% endif %}>{{ p.name }}</option>{% endfor %}
      </select>
    </div>
    <div class="col-md-2">
      <label class="form-label">Data</label>
      <input class="form-control" type="date" name="work_date" value="{{ e.work_date.isoformat() }}" required>
    </div>
    <div class="col-md-2">
      <label class="form-label">Czas (HH:MM)</label>
      <input class="form-control" type="text" name="hhmm" value="{{ fmt(e.minutes) }}" required>
    </div>
    <div class="col-md-2 d-flex align-items-end gap-3">
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="is_extra" id="eextra" {% if e.is_extra %}checked{% endif %}>
        <label class="form-check-label" for="eextra">Extra</label>
      </div>
      <div class="form-check">
        <input class="form-check-input" type="checkbox" name="is_overtime" id="eot" {% if e.is_overtime %}checked{% endif %}>
        <label class="form-check-label" for="eot">Nadgodziny</label>
      </div>
    </div>
    <div class="col-12">
      <label class="form-label">Notatka</label>
      <input class="form-control" type="text" name="note" value="{{ e.note or '' }}">
    </div>
    <div class="col-12">
      <button class="btn btn-primary">Zapisz</button>
      <a class="btn btn-outline-secondary" href="{{ url_for('admin_entries') }}">Anuluj</a>
    </div>
  </form>
</div>
""", e=e, users=users, projects=projects, fmt=fmt_hhmm)
    return layout("Edytuj wpis", body)

@app.route("/admin/entries/<int:entry_id>/delete", methods=["POST"])
@login_required
def admin_entry_delete(entry_id):
    require_admin()
    e = Entry.query.get_or_404(entry_id)
    _delete_entry_images_files(e)
    db.session.delete(e)
    db.session.commit()
    flash("Usunięto wpis.")
    return redirect(url_for("admin_entries"))


# --- Backup / Restore ---
def _add_uploads_to_zip(z: zipfile.ZipFile):
    """Dodaje folder uploads do archiwum. Wspiera przywracanie zdjęć."""
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    # marker, żeby folder istniał nawet gdy nie ma zdjęć
    keep_path = os.path.join(UPLOAD_DIR, ".keep")
    if not os.path.exists(keep_path):
        try:
            open(keep_path, "a").close()
        except Exception:
            pass
    for root, _, files in os.walk(UPLOAD_DIR):
        for fn in files:
            full = os.path.join(root, fn)
            rel = os.path.relpath(full, UPLOAD_DIR)
            arc = os.path.join("uploads", rel).replace("\\", "/")
            try:
                z.write(full, arcname=arc)
            except Exception:
                pass

def _add_plans_to_zip(z: zipfile.ZipFile):
    """Dodaje folder plans (PDF) do archiwum backupu."""
    os.makedirs(PLANS_DIR, exist_ok=True)

    # marker, żeby folder istniał nawet gdy nie ma PDF
    keep_path = os.path.join(PLANS_DIR, ".keep")
    if not os.path.exists(keep_path):
        try:
            open(keep_path, "a").close()
        except Exception:
            pass

    for root, _, files in os.walk(PLANS_DIR):
        for fn in files:
            full = os.path.join(root, fn)
            rel = os.path.relpath(full, PLANS_DIR)
            arc = os.path.join("plans", rel).replace("\\", "/")
            try:
                z.write(full, arcname=arc)
            except Exception:
                pass

def _make_zip_bytes(path)->bytes:
    ensure_db_file()
    if not os.path.exists(path):
        open(path, "a").close()
        ensure_db_file()
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as z:
        z.write(path, arcname="app.db")
        _add_uploads_to_zip(z)
        _add_plans_to_zip(z)
    mem.seek(0)
    return mem.read()

def _replace_db_from_zipfileobj(fileobj):
    """Podmienia plik bazy danymi z archiwum ZIP (app.db w środku).

    1. Zamyka aktualne połączenia SQLAlchemy.
    2. Nadpisuje fizyczny plik DB_FILE zawartością app.db z backupu.
    3. Woła ensure_db_file(), aby upewnić się, że struktura tabel istnieje.
    """
    try:
        fileobj.seek(0)
    except Exception:
        pass

    # Ścieżka do aktualnej bazy
    target_path = DB_FILE
    target_dir = os.path.dirname(target_path) or "."
    os.makedirs(target_dir, exist_ok=True)

    # Zamykanie połączeń z bazą
    try:
        db.session.remove()
    except Exception:
        pass
    try:
        db.engine.dispose()
    except Exception:
        pass

    # Nadpisanie pliku bazy danymi z kopii zapasowej + odtworzenie zdjęć
    with zipfile.ZipFile(fileobj, "r") as z:
        names = z.namelist()
        if "app.db" not in names:
            raise RuntimeError("Brak pliku 'app.db' w archiwum.")
        with z.open("app.db") as src, open(target_path, "wb") as dst:
            shutil.copyfileobj(src, dst, length=1024 * 1024)

        # Zdjęcia: folder uploads/*
        upload_names = [n for n in names if n.startswith("uploads/")]
        # Czyścimy lokalny folder i odtwarzamy z backupu
        try:
            if os.path.exists(UPLOAD_DIR):
                shutil.rmtree(UPLOAD_DIR)
        except Exception:
            pass
        os.makedirs(UPLOAD_DIR, exist_ok=True)
        for n in upload_names:
            if n.endswith("/"):
                continue
            rel = n[len("uploads/"):]
            if not rel or rel == ".keep":
                continue
            out_path = os.path.join(UPLOAD_DIR, rel)
            os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
            try:
                with z.open(n) as src, open(out_path, "wb") as dst:
                    shutil.copyfileobj(src, dst, length=1024 * 1024)
            except Exception:
                pass

    
        # --- ODTWARZANIE PLANÓW PDF ---
        plan_names = [n for n in names if n.startswith("plans/")]

        try:
            if os.path.exists(PLANS_DIR):
                shutil.rmtree(PLANS_DIR)
        except Exception:
            pass

        os.makedirs(PLANS_DIR, exist_ok=True)
        for n in plan_names:
            if n.endswith("/"):
                continue
            rel = n[len("plans/"):]
            if not rel or rel == ".keep":
                continue
            out_path = os.path.join(PLANS_DIR, rel)
            os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
            try:
                with z.open(n) as src, open(out_path, "wb") as dst:
                    shutil.copyfileobj(src, dst, length=1024 * 1024)
            except Exception:
                pass

# Odtworzenie struktury (jeśli trzeba), bez kasowania danych
    ensure_db_file()

@app.route("/admin/backup", methods=["GET"])
@login_required
def admin_backup():
    require_admin()
    base = os.path.dirname(DB_FILE)
    bdir = os.path.join(base, "backups") if base else "backups"
    os.makedirs(bdir, exist_ok=True)
    files = sorted([f for f in os.listdir(bdir) if f.endswith(".zip")])

    # Proste statystyki bieżącej bazy
    db_path = DB_FILE
    users = projects = entries = None
    try:
        users = User.query.count()
        projects = Project.query.count()
        entries = Entry.query.count()
    except Exception:
        pass

    body = render_template_string("""
<div class="card p-3">
  <h5 class="mb-3">Kopie zapasowe</h5>
  <p class="small text-muted">
    Baza danych: <code>{{ db_path }}</code><br>
    Użytkownicy: {{ users if users is not none else "?" }},
    Projekty: {{ projects if projects is not none else "?" }},
    Wpisy: {{ entries if entries is not none else "?" }}
  </p>
  <form class="d-inline" method="post" action="{{ url_for('admin_backup_create') }}">
    <button class="btn btn-primary">Utwórz i pobierz kopię teraz</button>
  </form>
  <form class="d-inline ms-2" method="post" action="{{ url_for('admin_backup_create_save') }}">
    <button class="btn btn-outline-primary">Zapisz kopię na dysku serwera</button>
  </form>
  <form class="d-inline ms-2" method="post" action="{{ url_for('admin_backup_email') }}">
    <button class="btn btn-outline-success">Wyślij kopię zapasową na e-mail</button>
  </form>
  <hr class="my-3">
  <h6>Przywracanie z pliku (.zip)</h6>
  <form method="post" action="{{ url_for('admin_backup_restore') }}" enctype="multipart/form-data" onsubmit="return confirm('Zastąpić bieżącą bazę?')">
    <input class="form-control mb-2" type="file" name="file" accept=".zip" required>
    <button class="btn btn-danger">Przywróć</button>
  </form>
  <hr class="my-3">
  <h6>Pliki na dysku serwera</h6>
  {% if files %}
    <ul class="list-group">
      {% for f in files %}
        <li class="list-group-item d-flex justify-content-between align-items-center">
          <span>{{ f }}</span>
          <span>
            <a class="btn btn-sm btn-outline-success" href="{{ url_for('admin_backup_download', fname=f) }}">Pobierz</a>
            <a class="btn btn-sm btn-outline-danger" href="{{ url_for('admin_backup_restore_saved', fname=f) }}" onclick="return confirm('Przywrócić z tej kopii?')">Przywróć</a>
            <form class="d-inline ms-1" method="post" action="{{ url_for('admin_backup_delete', fname=f) }}" onsubmit="return confirm('Usunąć ten backup z serwera?')">
              <button class="btn btn-sm btn-danger">Usuń</button>
            </form>
          </span>
        </li>
      {% endfor %}
    </ul>
  {% else %}
    <div class="text-muted">Brak zapisanych kopii.</div>
  {% endif %}
</div>
""", files=files, db_path=db_path, users=users, projects=projects, entries=entries)
    return layout("Kopie zapasowe", body)

@app.route("/admin/backup/create", methods=["POST"])
@login_required
def admin_backup_create():
    require_admin()
    data = _make_zip_bytes(DB_FILE)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return send_file(io.BytesIO(data), as_attachment=True, download_name=f"app_backup_{ts}.zip", mimetype="application/zip")

@app.route("/admin/backup/create_save", methods=["POST"])
@login_required
def admin_backup_create_save():
    require_admin()
    base = os.path.dirname(DB_FILE)
    bdir = os.path.join(base, "backups") if base else "backups"
    os.makedirs(bdir, exist_ok=True)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    zip_path = os.path.join(bdir, f"app_backup_{ts}.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as z:
        if not os.path.exists(DB_FILE):
            open(DB_FILE, "a").close()
            ensure_db_file()
        z.write(DB_FILE, arcname="app.db")
        _add_uploads_to_zip(z)
    flash(f"Zapisano: {os.path.basename(zip_path)}")
    return redirect(url_for("admin_backup"))



@app.route("/admin/backup/email", methods=["POST"])
@login_required
def admin_backup_email():
    require_admin()
    # Konfiguracja SMTP z zmiennych środowiskowych
    smtp_host = os.environ.get("SMTP_HOST")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER")
    smtp_password = os.environ.get("SMTP_PASSWORD")
    backup_to = os.environ.get("BACKUP_EMAIL_TO")
    backup_from = os.environ.get("BACKUP_EMAIL_FROM") or smtp_user

    if not (smtp_host and smtp_port and smtp_user and smtp_password and backup_to):
        flash("Brak konfiguracji SMTP lub adresu docelowego BACKUP_EMAIL_TO.", "danger")
        return redirect(url_for("admin_backup"))

    # Przygotowanie danych kopii zapasowej w pamięci
    data = _make_zip_bytes(DB_FILE)
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    fname = f"app_backup_{ts}.zip"

    msg = EmailMessage()
    msg["Subject"] = f"Kopia zapasowa EKKO NOR – {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}"
    msg["From"] = backup_from
    msg["To"] = backup_to
    msg.set_content(
        "Kopia zapasowa bazy danych aplikacji EKKO NOR.\n"
        "Ta wiadomość została wygenerowana automatycznie przez system."
    )
    msg.add_attachment(data, maintype="application", subtype="zip", filename=fname)

    try:
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_password)
            server.send_message(msg)
        flash(f"Wysłano kopię zapasową na adres: {backup_to}", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Nie udało się wysłać kopii zapasowej: {e}", "danger")

    return redirect(url_for("admin_backup"))

@app.route("/admin/backup/download/<path:fname>")
@login_required
def admin_backup_download(fname):
    require_admin()
    base = os.path.dirname(DB_FILE)
    bdir = os.path.join(base, "backups") if base else "backups"
    path = os.path.join(bdir, secure_filename(fname))
    if not os.path.exists(path):
        abort(404)
    return send_file(path, as_attachment=True, download_name=os.path.basename(path), mimetype="application/zip")

@app.route("/admin/backup/delete/<path:fname>", methods=["POST"])
@login_required
def admin_backup_delete(fname):
    require_admin()
    base = os.path.dirname(DB_FILE)
    bdir = os.path.join(base, "backups") if base else "backups"
    os.makedirs(bdir, exist_ok=True)

    fname_only = os.path.basename(fname)
    path = os.path.join(bdir, fname_only)

    # zabezpieczenie: tylko pliki .zip w katalogu backups
    if not path.endswith(".zip"):
        abort(400)

    if os.path.exists(path):
        try:
            os.remove(path)
            flash("Usunięto backup z serwera.")
        except Exception as e:
            flash(f"Nie udało się usunąć backupu: {e}")
    else:
        flash("Nie znaleziono takiego backupu.")
    return redirect(url_for("admin_backup"))


@app.route("/admin/backup/restore", methods=["POST"])
@login_required
def admin_backup_restore():
    require_admin()
    f = request.files.get("file")
    if not f:
        flash("Nie wybrano pliku.")
        return redirect(url_for("admin_backup"))
    try:
        mem = io.BytesIO(f.read())
        _replace_db_from_zipfileobj(mem)
        # Statystyki po przywróceniu – żeby było widać, że dane są
        users = User.query.count()
        projects = Project.query.count()
        entries = Entry.query.count()
        flash(f"Przywrócono bazę z załączonego pliku. Użytkownicy: {users}, Projekty: {projects}, Wpisy: {entries}")
    except Exception as e:
        flash(f"Błąd przywracania: {e}")
    return redirect(url_for("admin_backup"))

@app.route("/admin/backup/restore_saved/<path:fname>")
@login_required
def admin_backup_restore_saved(fname):
    require_admin()
    base = os.path.dirname(DB_FILE)
    bdir = os.path.join(base, "backups") if base else "backups"
    path = os.path.join(bdir, secure_filename(fname))
    if not os.path.exists(path):
        abort(404)
    try:
        with open(path, "rb") as fp:
            _replace_db_from_zipfileobj(fp)
        # Statystyki po przywróceniu – żeby było widać, że dane są
        users = User.query.count()
        projects = Project.query.count()
        entries = Entry.query.count()
        flash(f"Przywrócono bazę z {fname}. Użytkownicy: {users}, Projekty: {projects}, Wpisy: {entries}")
    except Exception as e:
        flash(f"Błąd przywracania: {e}")
    return redirect(url_for("admin_backup"))


# --- Reports (with Excel export) ---


@app.route("/admin/reports", methods=["GET"])
@login_required

def admin_reports():
    require_admin()
    d_from = request.args.get("from")
    d_to = request.args.get("to")
    user_id = request.args.get("user_id")
    project_id = request.args.get("project_id")
    # Domyślnie pokazuj bieżący miesiąc (jeśli nie podano zakresu dat)
    if not d_from and not d_to:
        from datetime import date, timedelta
        today = date.today()
        first_day = today.replace(day=1)
        # pierwszy dzień następnego miesiąca
        if first_day.month == 12:
            next_month = first_day.replace(year=first_day.year + 1, month=1)
        else:
            next_month = first_day.replace(month=first_day.month + 1)
        last_day = next_month - timedelta(days=1)
        d_from = first_day.isoformat()
        d_to = last_day.isoformat()


    q = Entry.query.join(User).join(Project)
    if d_from:
        q = q.filter(Entry.work_date >= d_from)
    if d_to:
        q = q.filter(Entry.work_date <= d_to)
    if user_id and user_id != "all":
        q = q.filter(Entry.user_id == int(user_id))
    if project_id and project_id != "all":
        q = q.filter(Entry.project_id == int(project_id))

    rows = q.order_by(Entry.work_date.asc(), Entry.id.asc()).all()
    total_minutes = sum(e.minutes for e in rows)
    users = User.query.order_by(User.name).all()
    projects = Project.query.order_by(Project.name).all()

    body = render_template_string("""
<div class="card p-3">
  <h5 class="mb-3">Raport</h5>

  <form class="row g-2 mb-3" method="get">
    <div class="col-md-3">
      <label class="form-label">Od</label>
      <input class="form-control" type="date" name="from" value="{{ d_from or '' }}">
    </div>
    <div class="col-md-3">
      <label class="form-label">Do</label>
      <input class="form-control" type="date" name="to" value="{{ d_to or '' }}">
    </div>
    <div class="col-md-3">
      <label class="form-label">Pracownik</label>
      <select class="form-select" name="user_id">
        <option value="all">Wszyscy</option>
        {% for u in users %}
          <option value="{{ u.id }}" {% if request.args.get('user_id')|int == u.id %}selected{% endif %}>{{ u.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-3">
      <label class="form-label">Projekt</label>
      <select class="form-select" name="project_id">
        <option value="all">Wszystkie</option>
        {% for p in projects %}
          <option value="{{ p.id }}" {% if request.args.get('project_id')|int == p.id %}selected{% endif %}>{{ p.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-12 d-flex gap-2">
      <button class="btn btn-primary">Pokaż</button>
    </div>
  </form>

  <p class="small text-muted">Łącznie rekordów: {{ rows|length }}</p>

  {% if rows %}
    <div class="mb-2 d-flex gap-2">
      <a class="btn btn-outline-success btn-sm"
         href="{{ url_for('admin_reports_export') }}?from={{ d_from or '' }}&to={{ d_to or '' }}&user_id={{ request.args.get('user_id','all') }}&project_id={{ request.args.get('project_id','all') }}">
        Eksport prosty (Excel)
      </a>
      <a class="btn btn-outline-primary btn-sm"
         href="{{ url_for('admin_reports_payroll') }}?from={{ d_from or '' }}&to={{ d_to or '' }}&user_id={{ request.args.get('user_id','all') }}&project_id={{ request.args.get('project_id','all') }}">
        Lista płac (Excel)
      </a>
    </div>
    <div class="table-responsive">
      <table class="table table-sm table-striped align-middle">
        <thead>
          <tr>
            <th>Data</th>
            <th>Pracownik</th>
            <th>Projekt</th>
            <th>Godziny</th>
            <th>Extra</th>
            <th>Nadgodziny</th>
            <th>Notatka</th>
          </tr>
        </thead>
        <tbody>
        {% for entry in rows %}
          <tr>
            <td>{{ entry.work_date }}</td>
            <td>{{ entry.user.name }}</td>
            <td>{{ entry.project.name }}</td>
            <td>{{ fmt(entry.minutes) }}</td>
            <td>{% if entry.is_extra %}tak{% else %}-{% endif %}</td>
            <td>{% if entry.is_overtime %}tak{% else %}-{% endif %}</td>
            <td>{{ entry.note }}</td>
          </tr>
        {% endfor %}
        </tbody>
      </table>
    </div>
    <div class="mt-2 fw-bold">
      Suma godzin: {{ fmt(total_minutes) }}
    </div>
  {% else %}
    <div class="text-muted">Brak wpisów.</div>
  {% endif %}
</div>
    """, rows=rows, users=users, projects=projects, fmt=fmt_hhmm, total_minutes=total_minutes, d_from=d_from, d_to=d_to)
    return layout("Raport", body)

@app.route("/admin/reports/export", methods=["GET"])
@login_required
def admin_reports_export():
    require_admin()
    try:
        from openpyxl import Workbook
    except Exception:
        abort(500, "Brak pakietu openpyxl (sprawdź requirements.txt)")

    d_from = request.args.get("from")
    d_to = request.args.get("to")
    user_id = request.args.get("user_id", "all")
    project_id = request.args.get("project_id", "all")
    today = date.today().isoformat()
    if not d_from or not d_to:
        abort(400)

    d_from_dt = datetime.strptime(d_from, "%Y-%m-%d").date()
    d_to_dt = datetime.strptime(d_to, "%Y-%m-%d").date()
    q = Entry.query.join(User).join(Project).filter(
        Entry.work_date >= d_from_dt,
        Entry.work_date <= d_to_dt
    )
    if user_id != "all":
        q = q.filter(Entry.user_id == int(user_id))
    if project_id != "all":
        q = q.filter(Entry.project_id == int(project_id))
    rows = q.order_by(Entry.work_date.asc(), Entry.id.asc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Raport"
    ws.append(["Data", "Pracownik", "Projekt", "Godziny (HH:MM)", "Extra", "Nadgodziny", "Notatka"])
    for it in rows:
        ws.append([
            it.work_date.isoformat(),
            it.user.name,
            it.project.name,
            fmt_hhmm(it.minutes),
            "TAK" if it.is_extra else "",
            "TAK" if it.is_overtime else "",
            it.note or ""
        ])

    # podsumowanie
    total_min = sum(r.minutes for r in rows)
    ws.append([])
    ws.append(["Razem", "", "", fmt_hhmm(total_min), "", "", ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"raport_{d_from}_{d_to}.xlsx"
    return send_file(buf, as_attachment=True, download_name=fname, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
@app.route("/admin/reports/payroll", methods=["GET"])
@login_required
def admin_reports_payroll():
    require_admin()
    try:
        from openpyxl import Workbook
    except Exception:
        abort(500, "Brak pakietu openpyxl (sprawdź requirements.txt)")

    d_from = request.args.get("from")
    d_to = request.args.get("to")
    user_id = request.args.get("user_id", "all")
    project_id = request.args.get("project_id", "all")
    if not d_from or not d_to:
        abort(400)

    d_from_dt = datetime.strptime(d_from, "%Y-%m-%d").date()
    d_to_dt = datetime.strptime(d_to, "%Y-%m-%d").date()

    q = Entry.query.join(User).join(Project).filter(
        Entry.work_date >= d_from_dt,
        Entry.work_date <= d_to_dt
    )
    if user_id != "all":
        q = q.filter(Entry.user_id == int(user_id))
    if project_id != "all":
        q = q.filter(Entry.project_id == int(project_id))

    rows = q.order_by(User.name.asc(), Entry.work_date.asc(), Entry.id.asc()).all()

    from collections import defaultdict
    per_user = defaultdict(list)
    for e in rows:
        per_user[e.user].append(e)

    wb = Workbook()
    # usuń domyślny arkusz, jeśli istnieje
    default_ws = wb.active
    wb.remove(default_ws)

    def sheet_title(user):
        base = user.name or f"Uzytkownik_{user.id}"
        for ch in '[]:*?/\\':
            base = base.replace(ch, "_")
        if len(base) > 25:
            base = base[:25]
        return base

    for user, entries in per_user.items():
        ws = wb.create_sheet(title=sheet_title(user))
        ws.append([f"Lista płac – {user.name}"])
        ws.append([f"Okres: {d_from_dt.isoformat()} – {d_to_dt.isoformat()}"])
        ws.append([])
        ws.append(["Data", "Projekt", "Godziny (HH:MM)", "Extra", "Nadgodziny", "Notatka"])

        total_minutes = 0
        extra_minutes = 0
        overtime_minutes = 0

        for e in entries:
            ws.append([
                e.work_date.isoformat(),
                e.project.name,
                fmt_hhmm(e.minutes),
                "TAK" if e.is_extra else "",
                "TAK" if e.is_overtime else "",
                e.note or "",
            ])
            total_minutes += e.minutes
            if e.is_extra:
                extra_minutes += e.minutes
            if e.is_overtime:
                overtime_minutes += e.minutes

        ws.append([])
        ws.append(["Suma godzin", "", fmt_hhmm(total_minutes), "", "", ""])
        ws.append(["Suma extra", "", fmt_hhmm(extra_minutes), "", "", ""])
        ws.append(["Suma nadgodzin", "", fmt_hhmm(overtime_minutes), "", "", ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"lista_plac_{d_from_dt}_{d_to_dt}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
# --- User: podsumowanie godzin (bieżący i poprzedni miesiąc) ---
@app.route("/my-summary")
@login_required
def user_summary():
    today = date.today()
    # bieżący miesiąc
    cur_first, cur_last = month_bounds(today)
    # poprzedni miesiąc – weź dzień przed pierwszym dniem bieżącego
    prev_ref = cur_first - timedelta(days=1)
    prev_first, prev_last = month_bounds(prev_ref)

    cur_entries = (
        Entry.query
        .filter(
            Entry.user_id == current_user.id,
            Entry.work_date >= cur_first,
            Entry.work_date <= cur_last,
        )
        .order_by(Entry.work_date.asc(), Entry.id.asc())
        .all()
    )
    prev_entries = (
        Entry.query
        .filter(
            Entry.user_id == current_user.id,
            Entry.work_date >= prev_first,
            Entry.work_date <= prev_last,
        )
        .order_by(Entry.work_date.asc(), Entry.id.asc())
        .all()
    )

    cur_total = sum((e.minutes or 0) for e in cur_entries)
    prev_total = sum((e.minutes or 0) for e in prev_entries)

    cur_label = cur_first.strftime("%Y-%m")
    prev_label = prev_first.strftime("%Y-%m")

    body = render_template_string("""
<div class="row">
  <div class="col-md-12">
    <div class="card p-3">
      <h5 class="mb-3">Moje godziny – {{ current_user.name }}</h5>
      <p class="text-muted small mb-3">
        Zestawienie czasu pracy za bieżący ({{ cur_label }}) i poprzedni ({{ prev_label }}) miesiąc.
      </p>
      <div class="row">
        <div class="col-md-6">
          <h6>Aktualny miesiąc ({{ cur_label }})</h6>
          <table class="table table-sm align-middle">
            <thead>
              <tr>
                <th>Data</th>
                <th>Projekt</th>
                <th>Notatka</th>
                <th>Czas</th>
                <th>Extra</th>
                <th>OT</th>
              </tr>
            </thead>
            <tbody>
              {% for e in cur_entries %}
              <tr>
                <td>{{ e.work_date.isoformat() }}</td>
                <td>{{ e.project.name }}</td>
                <td>{{ e.note or '' }}</td>
                <td>{{ fmt(e.minutes) }}</td>
                <td>{% if e.is_extra %}✔{% else %}-{% endif %}</td>
                <td>{% if e.is_overtime %}✔{% else %}-{% endif %}</td>
              </tr>
              {% else %}
              <tr><td colspan="7" class="text-muted">Brak wpisów w tym miesiącu.</td></tr>
              {% endfor %}
            </tbody>
          </table>
          <div class="mt-2 fw-bold">Suma: {{ fmt(cur_total) }}</div>
        </div>
        <div class="col-md-6">
          <h6>Poprzedni miesiąc ({{ prev_label }})</h6>
          <table class="table table-sm align-middle">
            <thead>
              <tr>
                <th>Data</th>
                <th>Projekt</th>
                <th>Notatka</th>
                <th>Czas</th>
                <th>Extra</th>
                <th>OT</th>
              </tr>
            </thead>
            <tbody>
              {% for e in prev_entries %}
              <tr>
                <td>{{ e.work_date.isoformat() }}</td>
                <td>{{ e.project.name }}</td>
                <td>{{ e.note or '' }}</td>
                <td>{{ fmt(e.minutes) }}</td>
                <td>{% if e.is_extra %}✔{% else %}-{% endif %}</td>
                <td>{% if e.is_overtime %}✔{% else %}-{% endif %}</td>
              </tr>
              {% else %}
              <tr><td colspan="7" class="text-muted">Brak wpisów w poprzednim miesiącu.</td></tr>
              {% endfor %}
            </tbody>
          </table>
          <div class="mt-2 fw-bold">Suma: {{ fmt(prev_total) }}</div>
        </div>
      </div>
    </div>
  </div>
</div>
""", cur_entries=cur_entries, prev_entries=prev_entries, fmt=fmt_hhmm,
       cur_total=cur_total, prev_total=prev_total,
       cur_label=cur_label, prev_label=prev_label, date=date)
    return layout("Moje godziny", body)


# --- User: koszty ---
@app.route("/costs", methods=["GET", "POST"])
@login_required
def user_costs():
    if request.method == "POST":
        cost_date_str = request.form.get("cost_date")
        amount = (request.form.get("amount") or "").strip()
        description = request.form.get("description") or ""

        try:
            cost_date = datetime.strptime(cost_date_str, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            flash("Nieprawidłowa data kosztu.")
            return redirect(url_for("user_costs"))

        if not amount:
            flash("Podaj kwotę kosztu.")
            return redirect(url_for("user_costs"))

        db.session.add(Cost(
            user_id=current_user.id,
            cost_date=cost_date,
            amount=amount,
            description=description,
        ))
        db.session.commit()
        flash("Dodano koszt.")
        return redirect(url_for("user_costs"))

    today = date.today()
    cur_first, cur_last = month_bounds(today)
    prev_ref = cur_first - timedelta(days=1)
    prev_first, prev_last = month_bounds(prev_ref)

    current_costs = (
        Cost.query
        .filter(
            Cost.user_id == current_user.id,
            Cost.cost_date >= cur_first,
            Cost.cost_date <= cur_last,
        )
        .order_by(Cost.cost_date.asc(), Cost.id.asc())
        .all()
    )
    previous_costs = (
        Cost.query
        .filter(
            Cost.user_id == current_user.id,
            Cost.cost_date >= prev_first,
            Cost.cost_date <= prev_last,
        )
        .order_by(Cost.cost_date.asc(), Cost.id.asc())
        .all()
    )

    cur_label = cur_first.strftime("%Y-%m")
    prev_label = prev_first.strftime("%Y-%m")

    body = render_template_string("""
<div class="row">
  <div class="col-md-12">
    <div class="card p-3">
      <div class="d-flex justify-content-between align-items-center mb-3">
        <h5 class="mb-0">Moje koszty – {{ current_user.name }}</h5>
        <div class="text-end">
          <a class="btn btn-sm btn-outline-secondary" href="{{ url_for('user_costs_export_xlsx') }}">Eksport Excel</a>
          <a class="btn btn-sm btn-outline-secondary" target="_blank" href="{{ url_for('user_costs_print') }}">Druk</a>
        </div>
      </div>
      <form class="row g-2 mb-3" method="post">
        <div class="col-md-3">
          <label class="form-label">Data kosztu</label>
          <input class="form-control" type="date" name="cost_date" value="{{ date.today().isoformat() }}" required>
        </div>
        <div class="col-md-3">
          <label class="form-label">Kwota</label>
          <input class="form-control" type="text" name="amount" placeholder="np. 1234,50" required>
        </div>
        <div class="col-md-4">
          <label class="form-label">Opis</label>
          <input class="form-control" type="text" name="description" placeholder="np. paliwo, narzędzia">
        </div>
        <div class="col-md-2 d-flex align-items-end">
          <button class="btn btn-primary w-100">Dodaj koszt</button>
        </div>
      </form>

      <div class="row">
        <div class="col-md-6">
          <h6>Aktualny miesiąc ({{ cur_label }})</h6>
          <table class="table table-sm align-middle">
            <thead>
              <tr>
                <th>Data</th>
                <th>Kwota</th>
                <th>Opis</th>
              </tr>
            </thead>
            <tbody>
              {% for c in current_costs %}
              <tr>
                <td>{{ c.cost_date.isoformat() }}</td>
                <td>{{ c.amount }}</td>
                <td>{{ c.description or '' }}</td>
              </tr>
              {% else %}
              <tr><td colspan="3" class="text-muted">Brak kosztów w tym miesiącu.</td></tr>
              {% endfor %}
            </tbody>
          </table>
        </div>
        <div class="col-md-6">
          <h6>Poprzedni miesiąc ({{ prev_label }})</h6>
          <table class="table table-sm align-middle">
            <thead>
              <tr>
                <th>Data</th>
                <th>Kwota</th>
                <th>Opis</th>
              </tr>
            </thead>
            <tbody>
              {% for c in previous_costs %}
              <tr>
                <td>{{ c.cost_date.isoformat() }}</td>
                <td>{{ c.amount }}</td>
                <td>{{ c.description or '' }}</td>
              </tr>
              {% else %}
              <tr><td colspan="3" class="text-muted">Brak kosztów w poprzednim miesiącu.</td></tr>
              {% endfor %}
            </tbody>
          </table>
        </div>
      </div>
      <p class="small text-muted mt-2 mb-0">
        Po zapisaniu kosztu nie możesz go już edytować – w razie potrzeby skontaktuj się z administratorem.
      </p>
    </div>
  </div>
</div>
""", current_costs=current_costs, previous_costs=previous_costs,
       cur_label=cur_label, prev_label=prev_label, date=date)
    return layout("Moje koszty", body)



@app.route("/costs/export.xlsx")
@login_required
def user_costs_export_xlsx():
    # eksport tylko swoich kosztów
    costs = (
        Cost.query.filter_by(user_id=current_user.id)
        .order_by(Cost.cost_date.desc(), Cost.id.desc())
        .all()
    )

    data_rows = []
    for c in costs:
        data_rows.append([
            c.cost_date.isoformat(),
            (c.amount or "").strip(),
            (c.description or "").strip(),
            (c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""),
        ])

    headers = ["Data", "Kwota", "Opis", "Utworzono"]
    bio_xlsx = _make_xlsx_bytes(headers, data_rows, sheet_name="Koszty")
    filename = f"koszty_{current_user.name}_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        bio_xlsx,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/costs/print")
@login_required
def user_costs_print():
    costs = (
        Cost.query.filter_by(user_id=current_user.id)
        .order_by(Cost.cost_date.desc(), Cost.id.desc())
        .all()
    )

    body = render_template_string(
        """<!doctype html>
<html lang="pl">
<head>
  <meta charset="utf-8">
  <title>Koszty – wydruk</title>
  <style>
    body { font-family: Arial, sans-serif; font-size: 12px; margin: 24px; }
    h2 { margin: 0 0 8px 0; }
    .meta { color: #555; margin-bottom: 14px; }
    table { width: 100%; border-collapse: collapse; }
    th, td { border: 1px solid #999; padding: 6px 8px; vertical-align: top; }
    th { background: #f2f2f2; text-align: left; }
    .small { color:#666; font-size:11px; }
    @media print { .noprint { display:none; } }
  </style>
</head>
<body>
  <div class="noprint" style="margin-bottom:10px;">
    <button onclick="window.print()">Drukuj</button>
  </div>

  <h2>Koszty – {{ user.name }}</h2>
  <div class="meta">Wygenerowano: {{ now }}</div>

  <table>
    <thead>
      <tr>
        <th>Data</th>
        <th>Kwota</th>
        <th>Opis</th>
      </tr>
    </thead>
    <tbody>
      {% for c in costs %}
      <tr>
        <td>{{ c.cost_date.isoformat() }}</td>
        <td>{{ c.amount }}</td>
        <td>{{ c.description or '' }}</td>
      </tr>
      {% endfor %}
      {% if not costs %}
      <tr><td colspan="3" class="small">Brak danych.</td></tr>
      {% endif %}
    </tbody>
  </table>

  <script>window.onload = () => { window.print(); };</script>
</body>
</html>""",
        costs=costs,
        user=current_user,
        now=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    return body




# --- Admin: koszty wszystkich użytkowników ---
@app.route("/admin/costs", methods=["GET", "POST"])
@login_required
def admin_costs():
    require_admin()

    if request.method == "POST":
        user_id = int(request.form.get("user_id"))
        cost_date_str = request.form.get("cost_date")
        amount = (request.form.get("amount") or "").strip()
        description = request.form.get("description") or ""

        try:
            cost_date = datetime.strptime(cost_date_str, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            flash("Nieprawidłowa data kosztu.")
            return redirect(url_for("admin_costs"))

        if not amount:
            flash("Podaj kwotę kosztu.")
            return redirect(url_for("admin_costs"))

        db.session.add(Cost(
            user_id=user_id,
            cost_date=cost_date,
            amount=amount,
            description=description,
        ))
        db.session.commit()
        flash("Dodano koszt.")
        return redirect(url_for("admin_costs"))

    users = _all_users_ordered()
    costs = (
        Cost.query
        .join(User)
        .order_by(Cost.cost_date.desc(), Cost.id.desc())
        .all()
    )

    body = render_template_string("""
<div class="row">
  <div class="col-md-12">
    <div class="card p-3">
      <div class="d-flex justify-content-between align-items-center mb-3">
        <h5 class="mb-0">Koszty – wszyscy użytkownicy</h5>
        <div class="text-end">
          <a class="btn btn-sm btn-outline-secondary" href="{{ url_for('admin_costs_export_xlsx') }}">Eksport Excel</a>
          <a class="btn btn-sm btn-outline-secondary" target="_blank" href="{{ url_for('admin_costs_print') }}">Druk</a>
        </div>
      </div>
      <form class="row g-2 mb-3" method="post">
        <div class="col-md-3">
          <label class="form-label">Pracownik</label>
          <select class="form-select" name="user_id" required>
            {% for u in users %}
              <option value="{{ u.id }}">{{ u.name }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-3">
          <label class="form-label">Data kosztu</label>
          <input class="form-control" type="date" name="cost_date" value="{{ date.today().isoformat() }}" required>
        </div>
        <div class="col-md-3">
          <label class="form-label">Kwota</label>
          <input class="form-control" type="text" name="amount" placeholder="np. 1234,50" required>
        </div>
        <div class="col-md-3">
          <label class="form-label">Opis</label>
          <input class="form-control" type="text" name="description" placeholder="np. paliwo, narzędzia">
        </div>
        <div class="col-md-12 d-flex justify-content-end mt-2">
          <button class="btn btn-primary">Dodaj koszt</button>
        </div>
      </form>

      <div class="table-responsive">
        <table class="table table-sm align-middle">
          <thead>
            <tr>
              <th>Data</th>
              <th>Pracownik</th>
              <th>Kwota</th>
              <th>Opis</th>
              <th class="text-end">Akcje</th>
            </tr>
          </thead>
          <tbody>
            {% for c in costs %}
            <tr>
              <td>{{ c.cost_date.isoformat() }}</td>
              <td>{{ c.user.name }}</td>
              <td>{{ c.amount }}</td>
              <td>{{ c.description or '' }}</td>
              <td class="text-end">
                <a class="btn btn-sm btn-outline-primary" href="{{ url_for('admin_cost_edit', cost_id=c.id) }}">Edytuj</a>
                <form class="d-inline" method="post" action="{{ url_for('admin_cost_delete', cost_id=c.id) }}" onsubmit="return confirm('Na pewno usunąć ten koszt?');">
                  <button class="btn btn-sm btn-outline-danger">Usuń</button>
                </form>
              </td>
            </tr>
            {% else %}
            <tr><td colspan="5" class="text-muted">Brak kosztów.</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</div>
""", users=users, costs=costs, date=date)
    return layout("Koszty (admin)", body)



@app.route("/admin/costs/export.xlsx")
@login_required
def admin_costs_export_xlsx():
    require_admin()

    costs = (
        Cost.query.join(User)
        .order_by(Cost.cost_date.desc(), Cost.id.desc())
        .all()
    )

    data_rows = []
    for c in costs:
        data_rows.append([
            c.user.name,
            c.cost_date.isoformat(),
            (c.amount or "").strip(),
            (c.description or "").strip(),
            (c.created_at.strftime("%Y-%m-%d %H:%M") if c.created_at else ""),
        ])

    headers = ["Użytkownik", "Data", "Kwota", "Opis", "Utworzono"]
    bio_xlsx = _make_xlsx_bytes(headers, data_rows, sheet_name="Koszty")
    filename = f"koszty_admin_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        bio_xlsx,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/costs/print")
@login_required
def admin_costs_print():
    require_admin()

    costs = (
        Cost.query.join(User)
        .order_by(Cost.cost_date.desc(), Cost.id.desc())
        .all()
    )

    body = render_template_string(
        """<!doctype html>
<html lang="pl">
<head>
  <meta charset="utf-8">
  <title>Koszty – wydruk (admin)</title>
  <style>
    body { font-family: Arial, sans-serif; font-size: 12px; margin: 24px; }
    h2 { margin: 0 0 8px 0; }
    .meta { color: #555; margin-bottom: 14px; }
    table { width: 100%; border-collapse: collapse; }
    th, td { border: 1px solid #999; padding: 6px 8px; vertical-align: top; }
    th { background: #f2f2f2; text-align: left; }
    .small { color:#666; font-size:11px; }
    @media print { .noprint { display:none; } }
  </style>
</head>
<body>
  <div class="noprint" style="margin-bottom:10px;">
    <button onclick="window.print()">Drukuj</button>
  </div>

  <h2>Koszty – zestawienie (admin)</h2>
  <div class="meta">Wygenerowano: {{ now }}</div>

  <table>
    <thead>
      <tr>
        <th>Użytkownik</th>
        <th>Data</th>
        <th>Kwota</th>
        <th>Opis</th>
      </tr>
    </thead>
    <tbody>
      {% for c in costs %}
      <tr>
        <td>{{ c.user.name }}</td>
        <td>{{ c.cost_date.isoformat() }}</td>
        <td>{{ c.amount }}</td>
        <td>{{ c.description or '' }}</td>
      </tr>
      {% endfor %}
      {% if not costs %}
      <tr><td colspan="4" class="small">Brak danych.</td></tr>
      {% endif %}
    </tbody>
  </table>

  <script>window.onload = () => { window.print(); };</script>
</body>
</html>""",
        costs=costs,
        now=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    return body




@app.route("/admin/costs/<int:cost_id>/edit", methods=["GET", "POST"])
@login_required
def admin_cost_edit(cost_id):
    require_admin()
    cost = Cost.query.get_or_404(cost_id)
    users = _all_users_ordered()

    if request.method == "POST":
        cost.user_id = int(request.form.get("user_id"))
        cost_date_str = request.form.get("cost_date")
        cost.amount = (request.form.get("amount") or "").strip()
        cost.description = request.form.get("description") or ""

        try:
            cost.cost_date = datetime.strptime(cost_date_str, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            flash("Nieprawidłowa data kosztu.")
            return redirect(url_for("admin_cost_edit", cost_id=cost.id))

        if not cost.amount:
            flash("Podaj kwotę kosztu.")
            return redirect(url_for("admin_cost_edit", cost_id=cost.id))

        db.session.commit()
        flash("Zapisano zmiany.")
        return redirect(url_for("admin_costs"))

    body = render_template_string("""
<div class="row justify-content-center">
  <div class="col-md-6">
    <div class="card p-3">
      <h5 class="mb-3">Edytuj koszt</h5>
      <form class="row g-2" method="post">
        <div class="col-md-6">
          <label class="form-label">Pracownik</label>
          <select class="form-select" name="user_id" required>
            {% for u in users %}
              <option value="{{ u.id }}" {% if u.id == cost.user_id %}selected{% endif %}>{{ u.name }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-6">
          <label class="form-label">Data kosztu</label>
          <input class="form-control" type="date" name="cost_date" value="{{ cost.cost_date.isoformat() }}" required>
        </div>
        <div class="col-md-6">
          <label class="form-label">Kwota</label>
          <input class="form-control" type="text" name="amount" value="{{ cost.amount }}" required>
        </div>
        <div class="col-md-12">
          <label class="form-label">Opis</label>
          <input class="form-control" type="text" name="description" value="{{ cost.description or '' }}">
        </div>
        <div class="col-md-12 d-flex justify-content-end">
          <button class="btn btn-primary">Zapisz</button>
        </div>
      </form>
    </div>
  </div>
</div>
""", users=users, cost=cost)
    return layout("Edytuj koszt", body)


@app.route("/admin/costs/<int:cost_id>/delete", methods=["POST"])
@login_required
def admin_cost_delete(cost_id):
    require_admin()
    cost = Cost.query.get_or_404(cost_id)
    db.session.delete(cost)
    db.session.commit()
    flash("Usunięto koszt.")
    return redirect(url_for("admin_costs"))


# --- Urlopy (Leave requests) ---
def _leave_status_pl(s: str) -> str:
    s = (s or "").upper()
    if s == "DRAFT":
        return "Szkic"
    if s == "SUBMITTED":
        return "Wysłane"
    if s == "APPROVED":
        return "Zaakceptowane"
    return s or "-"




def _make_xlsx_bytes(headers, rows, sheet_name="Dane"):
    """headers: list[str], rows: iterable[iterable]"""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] if sheet_name else "Dane"

    # Header
    ws.append(list(headers))

    # Rows
    for r in rows:
        ws.append(list(r))

    # Basic formatting: bold header + autosize
    try:
        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)
    except Exception:
        pass

    for col_idx in range(1, len(headers) + 1):
        max_len = 0
        for cell in ws[get_column_letter(col_idx)]:
            v = cell.value
            if v is None:
                continue
            s = str(v)
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, max_len + 2), 60)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio

@app.route("/leaves", methods=["GET", "POST"])
@login_required
def leaves():
    # Użytkownik: tworzy szkic urlopu
    if not current_user.is_admin and request.method == "POST":
        df = request.form.get("date_from")
        dt = request.form.get("date_to")
        reason = request.form.get("reason") or ""

        try:
            date_from = datetime.strptime(df, "%Y-%m-%d").date()
            date_to = datetime.strptime(dt, "%Y-%m-%d").date()
        except Exception:
            flash("Nieprawidłowa data.")
            return redirect(url_for("leaves"))

        if date_to < date_from:
            flash("Data 'do' nie może być wcześniejsza niż 'od'.")
            return redirect(url_for("leaves"))

        lr = LeaveRequest(
            user_id=current_user.id,
            date_from=date_from,
            date_to=date_to,
            reason=reason,
            status="DRAFT",
        )
        db.session.add(lr)
        db.session.commit()
        flash("Dodano prośbę o urlop (szkic).")
        return redirect(url_for("leaves"))


    # Admin: dodaje urlop wybranemu użytkownikowi (od razu zaakceptowany)
    if current_user.is_admin and request.method == "POST" and request.form.get("action") == "admin_add":
        uid = request.form.get("user_id")
        df = request.form.get("date_from")
        dt = request.form.get("date_to")
        reason = request.form.get("reason") or ""

        try:
            user_id = int(uid)
            date_from = datetime.strptime(df, "%Y-%m-%d").date()
            date_to = datetime.strptime(dt, "%Y-%m-%d").date()
        except Exception:
            flash("Nieprawidłowe dane formularza.", "danger")
            return redirect(url_for("leaves"))

        if date_to < date_from:
            flash("Data 'Do' nie może być wcześniejsza niż 'Od'.", "danger")
            return redirect(url_for("leaves"))

        u = User.query.get(user_id)
        if not u:
            flash("Nie znaleziono użytkownika.", "danger")
            return redirect(url_for("leaves"))

        now = datetime.utcnow()
        lr = LeaveRequest(
            user_id=user_id,
            date_from=date_from,
            date_to=date_to,
            reason=reason,
            status="APPROVED",
            submitted_at=now,
            decided_at=now,
            decided_by=current_user.id,
        )
        db.session.add(lr)
        db.session.commit()
        flash("Urlop został dodany i zaakceptowany.", "success")
        return redirect(url_for("leaves"))

    # Admin: lista wszystkich
    if current_user.is_admin:
        users = User.query.order_by(User.name.asc(), User.id.asc()).all()
        rows = (
            LeaveRequest.query
            .join(User, LeaveRequest.user_id == User.id)
            .order_by(LeaveRequest.created_at.desc(), LeaveRequest.id.desc())
            .all()
        )

        body = render_template_string("""
<div class="card p-3">
  <div class="d-flex justify-content-between align-items-center mb-3">
    <h5 class="mb-0">Urlopy – wszystkie prośby</h5>
    <div class="text-end">
      <a class="btn btn-sm btn-outline-secondary" href="{{ url_for('admin_leaves_export_xlsx') }}">Eksport Excel</a>
      <a class="btn btn-sm btn-outline-secondary" target="_blank" href="{{ url_for('admin_leaves_print') }}">Druk</a>
    </div>
  </div>
  <form method="post" class="row g-2 mb-3">
    <input type="hidden" name="action" value="admin_add">
    <div class="col-12 col-md-3">
      <select class="form-select form-select-sm" name="user_id" required>
        <option value="" disabled selected>Wybierz użytkownika</option>
        {% for u in users %}
          <option value="{{ u.id }}">{{ u.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-6 col-md-2">
      <input class="form-control form-control-sm" type="date" name="date_from" required>
    </div>
    <div class="col-6 col-md-2">
      <input class="form-control form-control-sm" type="date" name="date_to" required>
    </div>
    <div class="col-12 col-md-3">
      <input class="form-control form-control-sm" type="text" name="reason" placeholder="Uzasadnienie (opcjonalnie)">
    </div>
    <div class="col-12 col-md-2 text-md-end">
      <button class="btn btn-sm btn-primary" type="submit">Dodaj urlop</button>
    </div>
  </form>
  <div class="table-responsive">
    <table class="table table-sm align-middle">
      <thead>
        <tr>
          <th>Pracownik</th>
          <th>Od</th>
          <th>Do</th>
          <th>Status</th>
          <th>Uzasadnienie</th>
          <th class="text-end">Akcje</th>
        </tr>
      </thead>
      <tbody>
        {% for r in rows %}
        <tr>
          <td>{{ r.user.name }}</td>
          <td>{{ r.date_from.isoformat() }}</td>
          <td>{{ r.date_to.isoformat() }}</td>
          <td>
            <span class="badge badge-soft">{{ status_pl(r.status) }}</span>
          </td>
          <td style="max-width:420px;">{{ (r.reason or '')[:250] }}{% if r.reason and r.reason|length > 250 %}...{% endif %}</td>
          <td class="text-end text-nowrap">
            {% if r.status != 'APPROVED' %}
              <form class="d-inline" method="post" action="{{ url_for('leave_approve', leave_id=r.id) }}" onsubmit="return confirm('Zaakceptować ten urlop?')">
                <button class="btn btn-sm btn-outline-success">Akceptuj</button>
              </form>
            {% endif %}
            <form class="d-inline" method="post" action="{{ url_for('leave_delete', leave_id=r.id) }}" onsubmit="return confirm('Usunąć tę prośbę?')">
              <button class="btn btn-sm btn-outline-danger">Usuń</button>
            </form>
          </td>
        </tr>
        {% else %}
          <tr><td colspan="7" class="text-muted">Brak próśb o urlop.</td></tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</div>
""", rows=rows, users=users, status_pl=_leave_status_pl)
        return layout("Urlopy (admin)", body)

    # User: lista swoich
    rows = (
        LeaveRequest.query
        .filter(LeaveRequest.user_id == current_user.id)
        .order_by(LeaveRequest.created_at.desc(), LeaveRequest.id.desc())
        .all()
    )

    body = render_template_string("""
<div class="row g-3">
  <div class="col-12">
    <div class="card p-3">
      <h5 class="mb-3">Zgłoś urlop</h5>
      <form class="row g-2" method="post">
        <div class="col-md-3">
          <label class="form-label">Od</label>
          <input class="form-control" type="date" name="date_from" value="{{ date.today().isoformat() }}" required>
        </div>
        <div class="col-md-3">
          <label class="form-label">Do</label>
          <input class="form-control" type="date" name="date_to" value="{{ date.today().isoformat() }}" required>
        </div>
        <div class="col-md-6">
          <label class="form-label">Uzasadnienie</label>
          <input class="form-control" type="text" name="reason" placeholder="np. wyjazd, sprawy rodzinne">
        </div>
        <div class="col-12">
          <button class="btn btn-primary">Dodaj (szkic)</button>
        </div>
      </form>
      <div class="small text-muted mt-2">
        Najpierw dodajesz szkic, potem przy prośbie klikasz „Wyślij do akceptacji”.
      </div>
    </div>
  </div>

  <div class="col-12">
    <div class="card p-3">
      <h5 class="mb-0">Moje urlopy</h5>
      <div class="table-responsive mt-3">
        <table class="table table-sm align-middle">
          <thead>
            <tr>
              <th>Od</th>
              <th>Do</th>
              <th>Status</th>
              <th>Uzasadnienie</th>
              <th class="text-end">Akcje</th>
            </tr>
          </thead>
          <tbody>
            {% for r in rows %}
            <tr>
              <td>{{ r.date_from.isoformat() }}</td>
              <td>{{ r.date_to.isoformat() }}</td>
              <td><span class="badge badge-soft">{{ status_pl(r.status) }}</span></td>
              <td style="max-width:520px;">{{ r.reason or '' }}</td>
              <td class="text-end text-nowrap">
                {% if r.status != 'APPROVED' %}
                  <a class="btn btn-sm btn-outline-primary" href="{{ url_for('leave_edit', leave_id=r.id) }}">Edytuj</a>
                  <form class="d-inline" method="post" action="{{ url_for('leave_delete', leave_id=r.id) }}" onsubmit="return confirm('Usunąć tę prośbę?')">
                    <button class="btn btn-sm btn-outline-danger">Usuń</button>
                  </form>
                  {% if r.status == 'DRAFT' %}
                    <form class="d-inline" method="post" action="{{ url_for('leave_submit', leave_id=r.id) }}" onsubmit="return confirm('Wysłać do akceptacji?')">
                      <button class="btn btn-sm btn-outline-success">Wyślij do akceptacji</button>
                    </form>
                  {% endif %}
                {% else %}
                  <span class="text-muted">Zaakceptowane – bez zmian</span>
                {% endif %}
              </td>
            </tr>
            {% else %}
              <tr><td colspan="5" class="text-muted">Brak próśb o urlop.</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</div>
""", rows=rows, status_pl=_leave_status_pl, date=date)
    return layout("Urlopy", body)



@app.route("/admin/leaves/export.xlsx")
@login_required
def admin_leaves_export_xlsx():
    require_admin()
    rows = (
        LeaveRequest.query.join(User, LeaveRequest.user_id == User.id)
        .order_by(LeaveRequest.created_at.desc())
        .all()
    )

    data_rows = []
    for r in rows:
        days = None
        try:
            days = (r.date_to - r.date_from).days + 1
        except Exception:
            pass
        data_rows.append([
            r.user.name,
            r.date_from.isoformat(),
            r.date_to.isoformat(),
            days,
            _leave_status_pl(r.status),
            (r.reason or "").strip(),
            (r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else ""),
            (r.submitted_at.strftime("%Y-%m-%d %H:%M") if getattr(r, "submitted_at", None) else ""),
            (r.decided_at.strftime("%Y-%m-%d %H:%M") if getattr(r, "decided_at", None) else ""),
        ])

    headers = ["Użytkownik", "Od", "Do", "Dni", "Status", "Uzasadnienie", "Utworzono", "Wysłano", "Zaakceptowano"]
    bio_xlsx = _make_xlsx_bytes(headers, data_rows, sheet_name="Urlopy")
    filename = f"urlopy_admin_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(
        bio_xlsx,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/leaves/print")
@login_required
def admin_leaves_print():
    require_admin()
    rows = (
        LeaveRequest.query.join(User, LeaveRequest.user_id == User.id)
        .order_by(LeaveRequest.created_at.desc())
        .all()
    )

    body = render_template_string(
        """<!doctype html>
<html lang="pl">
<head>
  <meta charset="utf-8">
  <title>Urlopy – wydruk</title>
  <style>
    body { font-family: Arial, sans-serif; font-size: 12px; margin: 24px; }
    h2 { margin: 0 0 8px 0; }
    .meta { color: #555; margin-bottom: 14px; }
    table { width: 100%; border-collapse: collapse; }
    th, td { border: 1px solid #999; padding: 6px 8px; vertical-align: top; }
    th { background: #f2f2f2; text-align: left; }
    .small { color:#666; font-size:11px; }
    @media print { .noprint { display:none; } }
  </style>
</head>
<body>
  <div class="noprint" style="margin-bottom:10px;">
    <button onclick="window.print()">Drukuj</button>
  </div>

  <h2>Urlopy – zestawienie (admin)</h2>
  <div class="meta">Wygenerowano: {{ now }}</div>

  <table>
    <thead>
      <tr>
        <th>Użytkownik</th>
        <th>Od</th>
        <th>Do</th>
        <th>Dni</th>
        <th>Status</th>
        <th>Uzasadnienie</th>
      </tr>
    </thead>
    <tbody>
      {% for r in rows %}
      <tr>
        <td>{{ r.user.name }}</td>
        <td>{{ r.date_from.isoformat() }}</td>
        <td>{{ r.date_to.isoformat() }}</td>
        <td>
          {% set d = (r.date_to - r.date_from).days + 1 %}
          {{ d }}
        </td>
        <td>{{ status_pl(r.status) }}</td>
        <td>{{ r.reason or '' }}</td>
      </tr>
      {% endfor %}
      {% if not rows %}
      <tr><td colspan="7" class="small">Brak danych.</td></tr>
      {% endif %}
    </tbody>
  </table>

  <script>window.onload = () => { window.print(); };</script>
</body>
</html>""",
        rows=rows,
        status_pl=_leave_status_pl,
        now=datetime.now().strftime("%Y-%m-%d %H:%M"),
    )
    return body




@app.route("/leaves/<int:leave_id>/edit", methods=["GET", "POST"])
@login_required
def leave_edit(leave_id):
    lr = LeaveRequest.query.get_or_404(leave_id)
    if not (current_user.is_admin or lr.user_id == current_user.id):
        abort(403)

    if lr.status == "APPROVED":
        flash("Zaakceptowanej prośby nie można edytować.")
        return redirect(url_for("leaves"))

    if request.method == "POST":
        df = request.form.get("date_from")
        dt = request.form.get("date_to")
        reason = request.form.get("reason") or ""

        try:
            date_from = datetime.strptime(df, "%Y-%m-%d").date()
            date_to = datetime.strptime(dt, "%Y-%m-%d").date()
        except Exception:
            flash("Nieprawidłowa data.")
            return redirect(url_for("leave_edit", leave_id=leave_id))

        if date_to < date_from:
            flash("Data 'do' nie może być wcześniejsza niż 'od'.")
            return redirect(url_for("leave_edit", leave_id=leave_id))

        lr.date_from = date_from
        lr.date_to = date_to
        lr.reason = reason
        db.session.commit()
        flash("Zapisano zmiany.")
        return redirect(url_for("leaves"))

    body = render_template_string("""
<div class="row justify-content-center">
  <div class="col-md-7">
    <div class="card p-3">
      <h5 class="mb-3">Edytuj prośbę o urlop</h5>
      <form class="row g-2" method="post">
        <div class="col-md-4">
          <label class="form-label">Od</label>
          <input class="form-control" type="date" name="date_from" value="{{ lr.date_from.isoformat() }}" required>
        </div>
        <div class="col-md-4">
          <label class="form-label">Do</label>
          <input class="form-control" type="date" name="date_to" value="{{ lr.date_to.isoformat() }}" required>
        </div>
        <div class="col-md-12">
          <label class="form-label">Uzasadnienie</label>
          <input class="form-control" type="text" name="reason" value="{{ lr.reason or '' }}">
        </div>
        <div class="col-12 d-flex gap-2">
          <button class="btn btn-primary">Zapisz</button>
          <a class="btn btn-outline-secondary" href="{{ url_for('leaves') }}">Anuluj</a>
        </div>
      </form>
      {% if lr.status == 'SUBMITTED' %}
        <div class="small text-muted mt-2">Ta prośba jest już wysłana do akceptacji. Nadal możesz ją edytować/usunąć, dopóki nie zostanie zaakceptowana.</div>
      {% endif %}
    </div>
  </div>
</div>
""", lr=lr)
    return layout("Edytuj urlop", body)


@app.route("/leaves/<int:leave_id>/delete", methods=["POST"])
@login_required
def leave_delete(leave_id):
    lr = LeaveRequest.query.get_or_404(leave_id)

    if current_user.is_admin:
        db.session.delete(lr)
        db.session.commit()
        flash("Usunięto prośbę o urlop.")
        return redirect(url_for("leaves"))

    if lr.user_id != current_user.id:
        abort(403)

    if lr.status == "APPROVED":
        flash("Zaakceptowanej prośby nie można usunąć.")
        return redirect(url_for("leaves"))

    db.session.delete(lr)
    db.session.commit()
    flash("Usunięto prośbę o urlop.")
    return redirect(url_for("leaves"))


@app.route("/leaves/<int:leave_id>/submit", methods=["POST"])
@login_required
def leave_submit(leave_id):
    lr = LeaveRequest.query.get_or_404(leave_id)
    if lr.user_id != current_user.id and not current_user.is_admin:
        abort(403)

    if lr.status == "APPROVED":
        flash("Ta prośba jest już zaakceptowana.")
        return redirect(url_for("leaves"))

    if lr.status != "DRAFT":
        flash("Ta prośba jest już wysłana do akceptacji.")
        return redirect(url_for("leaves"))

    lr.status = "SUBMITTED"
    lr.submitted_at = datetime.utcnow()
    db.session.commit()
    flash("Wysłano do akceptacji.")
    return redirect(url_for("leaves"))


@app.route("/leaves/<int:leave_id>/approve", methods=["POST"])
@login_required
def leave_approve(leave_id):
    require_admin()
    lr = LeaveRequest.query.get_or_404(leave_id)

    if lr.status == "APPROVED":
        flash("Ta prośba jest już zaakceptowana.")
        return redirect(url_for("leaves"))

    lr.status = "APPROVED"
    lr.decided_at = datetime.utcnow()
    lr.decided_by = current_user.id
    db.session.commit()
    flash("Zaakceptowano prośbę o urlop.")
    return redirect(url_for("leaves"))




# --- Dodatki (extra godziny) ---

@app.route("/dodatki", methods=["GET", "POST"])
@login_required
def extras():
    if request.method == "POST":
        work_date_s = request.form.get("work_date") or date.today().isoformat()
        project_id = int(request.form.get("project_id") or "0")
        hhmm = request.form.get("hhmm") or "0:00"
        desc = request.form.get("description") or ""
        minutes = parse_hhmm(hhmm)

        try:
            d = datetime.strptime(work_date_s, "%Y-%m-%d").date()
        except Exception:
            d = date.today()

        if project_id <= 0:
            flash("Wybierz projekt.", "danger")
            return redirect(url_for("extras"))

        req_obj = ExtraRequest(
            user_id=current_user.id,
            project_id=project_id,
            work_date=d,
            minutes=minutes,
            description=desc.strip() or None,
            status="NEW",
        )
        db.session.add(req_obj)
        db.session.commit()

        images_files = request.files.getlist("images")
        try:
            _save_extra_images(req_obj, images_files)
            db.session.commit()
        except Exception:
            db.session.rollback()

        flash("Dodano zgłoszenie dodatków.", "success")
        return redirect(url_for("extras"))

    projects = Project.query.filter_by(is_active=True).order_by(Project.name).all()
    my = ExtraRequest.query.filter_by(user_id=current_user.id).order_by(ExtraRequest.created_at.desc(), ExtraRequest.id.desc()).limit(50).all()

    body = render_template_string("""
<div class="row g-3">
  <div class="col-12">
    <div class="card p-3">
      <h5 class="mb-3">Dodatki – zgłoś extra godziny</h5>
      <form id="extrasForm" class="row g-2" method="post" enctype="multipart/form-data">
        <div class="col-md-3">
          <label class="form-label">Data</label>
          <input class="form-control" type="date" name="work_date" value="{{ date.today().isoformat() }}" required>
        </div>
        <div class="col-md-4">
          <label class="form-label">Projekt</label>
          <select class="form-select" name="project_id" required>
            {% for p in projects %}
              <option value="{{ p.id }}">{{ p.name }}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-2">
          <label class="form-label">Czas (HH:MM)</label>
          <input class="form-control" type="text" name="hhmm" value="1:00" required>
        </div>
        <div class="col-12">
          <label class="form-label">Opis</label>
          <input class="form-control" type="text" name="description" placeholder="Co było extra? (opcjonalnie)">
        </div>
        <div class="col-12">
          <label class="form-label">Zdjęcia</label>
          <input class="form-control" type="file" name="images" accept="image/*" multiple onchange="limitFiles(this,5)">
          <div class="form-text">Maks 5 zdjęć na zgłoszenie.</div>
        </div>
        <div class="col-12">
          <button class="btn btn-primary">Dodaj zgłoszenie</button>
        </div>
      </form>
    </div>
  </div>

  <div class="col-12">
    <div class="card p-3">
      <h5 class="mb-0">Moje zgłoszenia (ostatnie 50)</h5>
      <div class="table-responsive mt-3">
        <table class="table table-sm align-middle">
          <thead>
            <tr>
              <th>Data</th><th>Projekt</th><th>Opis</th><th>Zdjęcia</th><th>Godziny</th><th>Status</th><th class="text-end">Akcje</th>
            </tr>
          </thead>
          <tbody>
            {% for r in my %}
              <tr>
                <td>{{ r.work_date.isoformat() }}</td>
                <td>{{ r.project.name }}</td>
                <td>{{ r.description or '' }}</td>
                <td>
                  {% if r.images %}
                    {% for img in r.images %}
                      <a href="{{ url_for('extra_image_view', image_id=img.id) }}" target="_blank" rel="noopener">IMG</a>{% if not loop.last %} {% endif %}
                    {% endfor %}
                  {% else %}-{% endif %}
                </td>
                <td>{{ fmt(r.minutes) }}</td>
                <td><span class="badge bg-light text-dark border">{{ r.status }}</span></td>
              
                <td class="text-end text-nowrap">
                  {% if r.status == 'NEW' %}
                    <a class="btn btn-sm btn-outline-primary" href="{{ url_for('user_extra_request_edit', req_id=r.id) }}">Edytuj</a>
                    <form method="post" action="/dodatki/request/{{ r.id }}/delete" style="display:inline;" onsubmit="return confirm('Usunąć zgłoszenie?');">
                      <button class="btn btn-sm btn-outline-danger">Usuń</button>
                    </form>
                  {% else %}
                    <span class="text-muted">-</span>
                  {% endif %}
                </td>
              </tr>

            {% else %}
              <tr><td colspan="7" class="text-muted">Brak zgłoszeń.</td></tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
    </div>
  </div>
</div>
""", projects=projects, my=my, fmt=fmt_hhmm, date=date, categories=EXTRA_CATEGORIES)

    return layout("Dodatki", body)



@app.route("/dodatki/request/<int:req_id>/edit", methods=["GET", "POST"])
@login_required
def user_extra_request_edit(req_id):
    r = ExtraRequest.query.get_or_404(req_id)
    if r.user_id != current_user.id:
        abort(403)
    if r.status != "NEW":
        flash("Nie można edytować zgłoszenia, które zostało już wysłane do raportu.", "warning")
        return redirect(url_for("extras"))

    projects = Project.query.filter_by(is_active=True).order_by(Project.name).all()

    if request.method == "POST":
        try:
            r.work_date = datetime.strptime(request.form.get("work_date"), "%Y-%m-%d").date()
        except Exception:
            pass

        try:
            pid = int(request.form.get("project_id") or r.project_id)
            if pid > 0:
                r.project_id = pid
        except Exception:
            pass

        r.minutes = parse_hhmm(request.form.get("hhmm") or fmt_hhmm(r.minutes or 0))
        r.description = (request.form.get("description") or "").strip() or None

        # dodaj nowe zdjęcia (max 5 łącznie)
        files = request.files.getlist("images")
        if files:
            try:
                existing = len(r.images or [])
                if existing >= 5:
                    flash("Masz już 5 zdjęć w tym zgłoszeniu. Usuń jakieś zdjęcie, aby dodać nowe.", "warning")
                else:
                    _save_extra_images(r, files[: max(0, 5 - existing)])
            except Exception:
                pass

        db.session.commit()
        flash("Zapisano zmiany.", "success")
        return redirect(url_for("extras"))

    body = render_template_string(r"""
<div class="card p-3">
  <div class="d-flex justify-content-between align-items-center">
    <h5 class="mb-0">Edytuj zgłoszenie dodatków</h5>
    <a class="btn btn-outline-secondary" href="{{ url_for('extras') }}">Wróć</a>
  </div>

  <form class="row g-2 mt-3" method="post" enctype="multipart/form-data">
    <div class="col-md-3">
      <label class="form-label">Data</label>
      <input class="form-control" type="date" name="work_date" value="{{ r.work_date.isoformat() }}" required>
    </div>
    <div class="col-md-4">
      <label class="form-label">Projekt</label>
      <select class="form-select" name="project_id" required>
        {% for p in projects %}
          <option value="{{ p.id }}" {% if p.id == r.project_id %}selected{% endif %}>{{ p.name }}</option>
        {% endfor %}
      </select>
    </div>
    <div class="col-md-2">
      <label class="form-label">Godziny (HH:MM)</label>
      <input class="form-control" name="hhmm" value="{{ fmt(r.minutes) }}" placeholder="np. 02:30" required>
    </div>
    <div class="col-md-12">
      <label class="form-label">Opis</label>
      <textarea class="form-control" name="description" rows="3">{{ r.description or '' }}</textarea>
    </div>

    <div class="col-md-12">
      <label class="form-label">Zdjęcia (max 5 łącznie)</label>
      <input class="form-control" type="file" name="images" accept="image/*" multiple>
      <div class="small text-muted mt-1">Dodane zdjęcia:</div>
      <div class="mt-1">
        {% if r.images %}
          {% for img in r.images %}
            <a class="btn btn-sm btn-outline-secondary" href="{{ url_for('extra_image_view', image_id=img.id) }}" target="_blank" rel="noopener">Podgląd</a>
            <form method="post" action="{{ url_for('user_extra_image_delete', image_id=img.id) }}" style="display:inline;" onsubmit="return confirm('Usunąć to zdjęcie?');">
              <button class="btn btn-sm btn-outline-danger">Usuń zdjęcie</button>
            </form>
          {% endfor %}
        {% else %}
          <span class="text-muted">brak</span>
        {% endif %}
      </div>
    </div>

    <div class="col-md-12 mt-2">
      <button class="btn btn-primary">Zapisz</button>
      <a class="btn btn-outline-secondary" href="{{ url_for('extras') }}">Anuluj</a>
    </div>
  </form>
</div>
""", r=r, projects=projects, fmt=fmt_hhmm)

    return layout("Edytuj dodatki", body)


@app.route("/dodatki/image/<int:image_id>/delete", methods=["POST"])
@login_required
def user_extra_image_delete(image_id):
    img = ExtraRequestImage.query.get_or_404(image_id)
    req_obj = ExtraRequest.query.get(img.request_id)
    if not req_obj or req_obj.user_id != current_user.id:
        abort(403)
    if req_obj.status != "NEW":
        flash("Nie można usuwać zdjęć ze zgłoszenia, które zostało już wysłane do raportu.", "warning")
        return redirect(url_for("extras"))

    try:
        path = extra_image_view_path(img.stored_filename)
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        pass

    db.session.delete(img)
    db.session.commit()
    flash("Usunięto zdjęcie.", "success")
    return redirect(url_for("user_extra_request_edit", req_id=req_obj.id))





@app.route("/dodatki/request/<int:req_id>/delete", methods=["POST"])
@login_required
def user_extra_request_delete(req_id):
    r = ExtraRequest.query.get_or_404(req_id)
    if r.user_id != current_user.id:
        abort(403)
    if r.status != "NEW":
        flash("Nie można usunąć zgłoszenia, które zostało już wysłane do raportu.", "warning")
        return redirect(url_for("extras"))

    # usuń zdjęcia
    try:
        for img in list(r.images or []):
            try:
                path = extra_image_view_path(img.stored_filename)
                if os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass
    except Exception:
        pass

    db.session.delete(r)
    db.session.commit()
    flash("Zgłoszenie zostało usunięte.", "success")
    return redirect(url_for("extras"))


@app.route("/dodatki/image/<int:image_id>", methods=["GET"])
@login_required
def extra_image_view(image_id):
    img = ExtraRequestImage.query.get_or_404(image_id)
    path = extra_image_view_path(img.stored_filename)
    if not os.path.exists(path):
        abort(404)
    return send_file(path, mimetype="image/jpeg")


@app.route("/admin/dodatki", methods=["GET", "POST"])
@login_required
def admin_extras():
    require_admin()

    today = date.today()

    # Ustawienie maila kontaktowego per projekt (opcjonalnie)
    if request.method == "POST" and request.form.get("action") == "save_contact":
        pid = int(request.form.get("project_id") or "0")
        email = (request.form.get("contact_email") or "").strip()
        name = (request.form.get("contact_name") or "").strip()
        if pid and email:
            _upsert_project_contact(pid, email, name or None)
            db.session.commit()
            flash("Zapisano kontakt do projektu.", "success")
        return redirect(url_for("admin_extras", project_id=pid))


    # Dodawanie dodatku przez admina (jakby pracownik)
    if request.method == "POST" and request.form.get("action") == "admin_add_request":
        pid = int(request.form.get("project_id") or "0")
        uid = int(request.form.get("user_id") or "0")
        work_date_str = (request.form.get("work_date") or "").strip()
        minutes = parse_hhmm(request.form.get("minutes") or "0")
        desc = (request.form.get("description") or "").strip()

        if not pid or not uid:
            flash("Wybierz projekt i pracownika.", "warning")
            return redirect(url_for("admin_extras", project_id=pid or "all"))

        # data (YYYY-MM-DD)
        if work_date_str:
            try:
                work_date = datetime.strptime(work_date_str, "%Y-%m-%d").date()
            except Exception:
                flash("Nieprawidłowa data.", "warning")
                return redirect(url_for("admin_extras", project_id=pid))
        else:
            work_date = date.today()

        if minutes <= 0:
            flash("Podaj czas (np. 01:30).", "warning")
            return redirect(url_for("admin_extras", project_id=pid))

        req = ExtraRequest(
            user_id=uid,
            project_id=pid,
            work_date=work_date,
            minutes=minutes,
            description=desc,
            status="NEW",
        )
        db.session.add(req)
        db.session.commit()

        # zdjęcia (opcjonalnie)
        try:
            files = request.files.getlist("images") if "images" in request.files else []
            _save_extra_images(req, files)
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            flash(f"Nie udało się zapisać zdjęć: {e}", "warning")

        flash("Dodatek został dodany.", "success")
        return redirect(url_for("admin_extras", project_id=pid))

    projects = Project.query.order_by(Project.is_active.desc(), Project.name.asc()).all()
    selected_pid = request.args.get("project_id", "all")
    selected_pid_int = None

    q = ExtraRequest.query.join(User).join(Project).filter(ExtraRequest.status != "CANCELED")
    if selected_pid != "all":
        try:
            selected_pid_int = int(selected_pid)
            q = q.filter(ExtraRequest.project_id == selected_pid_int)
        except Exception:
            selected_pid = "all"
            selected_pid_int = None

    rows = q.order_by(ExtraRequest.created_at.desc(), ExtraRequest.id.desc()).limit(300).all()

    # kontakty
    contact_email = None
    contact_name = None
    if selected_pid != "all":
        try:
            c = ProjectContact.query.filter_by(project_id=int(selected_pid)).order_by(ProjectContact.is_default.desc(), ProjectContact.id.asc()).first()
            if c:
                contact_email = c.email
                contact_name = c.name
        except Exception:
            pass

    
    # lista pracowników do dodawania dodatków przez admina
    employees = User.query.order_by(User.name.asc()).all()
    body = render_template_string("""
<div class="row g-3">
  <div class="col-12">
    <div class="card p-3">
      <h5 class="mb-2">Dodatki (admin)</h5>
      <form class="row g-2 align-items-end" method="get">
        <div class="col-md-5">
          <label class="form-label">Projekt</label>
          <select class="form-select" name="project_id">
            <option value="all" {% if selected_pid == 'all' %}selected{% endif %}>Wszystkie</option>
            {% for p in projects %}
              <option value="{{ p.id }}" {% if selected_pid|int == p.id %}selected{% endif %}>{{ p.name }}{% if not p.is_active %} (nieaktywny){% endif %}</option>
            {% endfor %}
          </select>
        </div>
        <div class="col-md-2">
          <button class="btn btn-outline-primary w-100">Filtruj</button>
        </div>
        <div class="col-md-5 text-end">
          <a class="btn btn-outline-secondary" href="{{ url_for('admin_extra_reports') }}">Lista raportów</a>
        </div>
      </form>

      {% if selected_pid != 'all' %}
      <hr class="my-3">
      <form class="row g-2" method="post">
        <input type="hidden" name="action" value="save_contact">
        <div class="col-md-4">
          <label class="form-label">E-mail odpowiedzialnej osoby (domyślny)</label>
          <input class="form-control" name="contact_email" value="{{ contact_email or '' }}" placeholder="np. pm@firma.no">
        </div>
        <div class="col-md-4">
          <label class="form-label">Imię/Nazwa (opcjonalnie)</label>
          <input class="form-control" name="contact_name" value="{{ contact_name or '' }}" placeholder="np. Jan Kowalski">
        </div>
        <div class="col-md-2">
          <input type="hidden" name="project_id" value="{{ selected_pid }}">
          <button class="btn btn-outline-success w-100 mt-4">Zapisz</button>
        </div>
      </form>
      {% endif %}
    </div>
  </div>

  <div class="col-12">
    <div class="card p-3">
      <h6 class="mb-2">
        <div class="card mb-3 p-3">
          <h5>Dodaj dodatek (admin)</h5>
          <form method="post" enctype="multipart/form-data" class="mt-2">
            <input type="hidden" name="action" value="admin_add_request">
            <div class="row g-2">
              <div class="col-md-3">
                <label class="form-label">Pracownik</label>
                <select name="user_id" class="form-select" required>
                  <option value="">-- wybierz --</option>
                  {% for u in employees %}
                    <option value="{{u.id}}">{{u.name}}</option>
                  {% endfor %}
                </select>
              </div>
              <div class="col-md-3">
                <label class="form-label">Projekt</label>
                <select name="project_id" class="form-select" required>
                  <option value="">-- wybierz --</option>
                  {% for p in projects %}
                    <option value="{{p.id}}" {% if p.id==selected_pid_int %}selected{% endif %}>{{p.name}}</option>
                  {% endfor %}
                </select>
              </div>
              <div class="col-md-2">
                <label class="form-label">Data</label>
                <input type="date" name="work_date" class="form-control" value="{{today}}">
              </div>
              <div class="col-md-2">
                <label class="form-label">Godziny</label>
                <input type="text" name="minutes" class="form-control" placeholder="np. 01:30" required>
              </div>
              <div class="col-md-2">
                <label class="form-label">Zdjęcia</label>
                <input type="file" name="images" class="form-control" multiple accept="image/*">
              </div>
            </div>
            <div class="mt-2">
              <label class="form-label">Opis</label>
              <input type="text" name="description" class="form-control" placeholder="Opis (opcjonalnie)">
            </div>
            <button class="btn btn-success mt-2" type="submit">Dodaj</button>
          </form>
        </div>

Zgłoszenia (zaznacz i utwórz raport)</h6>
      <form method="post" action="{{ url_for('admin_extra_report_create') }}">
        <div class="row g-2 align-items-end mb-2">
          <div class="col-md-5">
            <label class="form-label">E-mail do wysyłki (możesz zmienić)</label>
            <input class="form-control" name="recipient_email" value="{{ contact_email or '' }}" placeholder="np. pm@firma.no">
          </div>
          <div class="col-md-5">
            <label class="form-label">Tekst w raporcie (opcjonalnie)</label>
            <input class="form-control" name="report_text" placeholder="Krótki opis dodatków / zakresu">
          </div>
          <div class="col-md-2">
            <button class="btn btn-primary w-100">Utwórz raport</button>
          </div>
        </div>

        <div class="table-responsive">
          <table class="table table-sm align-middle">
            <thead>
              <tr>
                <th></th><th>Data</th><th>Pracownik</th><th>Projekt</th><th>Godziny</th><th>Opis</th><th>Zdjęcia</th><th>Status</th><th class="text-end">Akcje</th>
              </tr>
            </thead>
            <tbody>
              {% for r in rows %}
                <tr>
                  <td><input class="form-check-input" type="checkbox" name="req_id" value="{{ r.id }}" {% if r.status == 'INCLUDED' %}disabled{% endif %}></td>
                  <td>{{ r.work_date.isoformat() }}</td>
                  <td>{{ r.user.name }}</td>
                  <td>{{ r.project.name }}</td>
                  <td>{{ fmt(r.minutes) }}</td>
                  <td>{{ r.description or '' }}</td>
                  <td>
                    {% if r.images %}
                      {% for img in r.images %}
                        <a href="{{ url_for('extra_image_view', image_id=img.id) }}" target="_blank" rel="noopener">IMG</a>{% if not loop.last %} {% endif %}
                      {% endfor %}
                    {% else %}-{% endif %}
                  </td>
                  <td><span class="badge bg-light text-dark border">{{ r.status }}</span></td>
                  <td class="text-end text-nowrap">
                    <a class="btn btn-sm btn-outline-primary" href="{{ url_for('admin_extra_request_edit', req_id=r.id) }}">Edytuj</a>
                    <button class="btn btn-sm btn-outline-danger" type="submit" formmethod="post" formaction="{{ url_for('admin_extra_request_delete', req_id=r.id) }}" onclick="return confirm('Usunąć zgłoszenie?');">Usuń</button>
                  </td>
                </tr>
              {% else %}
                <tr><td colspan="9" class="text-muted">Brak zgłoszeń.</td></tr>
              {% endfor %}
            </tbody>
          </table>
        </div>

        <input type="hidden" name="project_id" value="{{ selected_pid }}">
      </form>
    </div>
  </div>
</div>
""", projects=projects, rows=rows, selected_pid=selected_pid, fmt=fmt_hhmm, contact_email=contact_email, contact_name=contact_name, employees=employees, today=today, selected_pid_int=selected_pid_int)

    return layout("Dodatki (admin)", body)


@app.route("/admin/dodatki/request/<int:req_id>/edit", methods=["GET", "POST"])
@login_required
def admin_extra_request_edit(req_id):
    require_admin()
    r = ExtraRequest.query.get_or_404(req_id)

    if request.method == "POST":
        r.work_date = datetime.strptime(request.form.get("work_date"), "%Y-%m-%d").date()
        r.minutes = parse_hhmm(request.form.get("hhmm") or "0:00")
        r.description = (request.form.get("description") or "").strip() or None
        db.session.commit()
        flash("Zapisano zmiany.", "success")
        return redirect(url_for("admin_extras", project_id=r.project_id))

    body = render_template_string("""
<div class="card p-3">
  <h5 class="mb-3">Edytuj zgłoszenie dodatków</h5>
  <form id="adminExtraEditForm" class="row g-2" method="post">
    <div class="col-md-3">
      <label class="form-label">Data</label>
      <input class="form-control" type="date" name="work_date" value="{{ r.work_date.isoformat() }}" required>
    </div>
    <div class="col-md-4">
      <label class="form-label">Projekt</label>
      <input class="form-control" value="{{ r.project.name }}" disabled>
    </div>
    <div class="col-md-2">
      <label class="form-label">Czas (HH:MM)</label>
      <input class="form-control" type="text" name="hhmm" value="{{ fmt(r.minutes) }}" required>
    </div>
    <div class="col-12">
      <label class="form-label">Opis</label>
      <input class="form-control" type="text" name="description" value="{{ r.description or '' }}">
    </div>
    <div class="col-12">
      <button class="btn btn-primary">Zapisz</button>
      <a class="btn btn-outline-secondary" href="{{ url_for('admin_extras', project_id=r.project_id) }}">Wróć</a>
    </div>
  </form>
</div>
""", r=r, fmt=fmt_hhmm)
    return layout("Edytuj dodatki", body)


@app.route("/admin/dodatki/report/create", methods=["POST"])
@login_required
def admin_extra_report_create():
    require_admin()
    project_id = request.form.get("project_id")
    recipient_email = (request.form.get("recipient_email") or "").strip() or None
    report_text = (request.form.get("report_text") or "").strip() or None

    ids = request.form.getlist("req_id")
    if not ids:
        flash("Zaznacz przynajmniej jedno zgłoszenie.", "danger")
        return redirect(url_for("admin_extras", project_id=project_id or "all"))

    # Ustal projekt: jeśli wybrano filtr, wymuszamy spójność
    pid = None
    if project_id and project_id != "all":
        try:
            pid = int(project_id)
        except Exception:
            pid = None

    reqs = ExtraRequest.query.filter(ExtraRequest.id.in_([int(x) for x in ids])).all()
    if not reqs:
        flash("Nie znaleziono zgłoszeń.", "danger")
        return redirect(url_for("admin_extras", project_id=project_id or "all"))

    if pid is None:
        pid = reqs[0].project_id

    # pilnujemy, żeby raport był dla jednego projektu
    reqs = [r for r in reqs if r.project_id == pid and r.status != "INCLUDED"]
    if not reqs:
        flash("Wybrane zgłoszenia są już dodane do raportu albo nie pasują do projektu.", "warning")
        return redirect(url_for("admin_extras", project_id=pid))

    # domyślny mail jeśli pusty
    if not recipient_email:
        recipient_email = _default_project_contact_email(pid)

    rep = ExtraReport(
        project_id=pid,
        created_by=current_user.id,
        recipient_email=recipient_email,
        report_text=report_text,
        status="DRAFT",
    )
    db.session.add(rep)
    db.session.commit()

    # dodaj pozycje jako snapshot
    for r in reqs:
        it = ExtraReportItem(
            report_id=rep.id,
            request_id=r.id,
            user_name=r.user.name,
            work_date=r.work_date,
            minutes=r.minutes,
            description=r.description,
        )
        db.session.add(it)
        r.status = "INCLUDED"
    db.session.commit()

    flash("Utworzono raport (szkic).", "success")
    return redirect(url_for("admin_extra_report_view", report_id=rep.id))



@app.route("/admin/dodatki/request/<int:req_id>/delete", methods=["POST"])
@login_required
def admin_extra_request_delete(req_id):
    require_admin()
    r = ExtraRequest.query.get_or_404(req_id)

    # Jeśli zgłoszenie było już użyte w raporcie, usuń też powiązania (żeby admin mógł faktycznie skasować wpis)
    linked_items = ExtraReportItem.query.filter_by(request_id=r.id).all()
    if linked_items:
        for it in linked_items:
            try:
                db.session.delete(it)
            except Exception:
                pass

    # usuń pliki zdjęć
    try:
        for img in list(r.images or []):
            try:
                path = extra_image_view_path(img.stored_filename)
                if os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass
    except Exception:
        pass

    db.session.delete(r)
    db.session.commit()
    flash("Usunięto zgłoszenie.", "success")
    return redirect(url_for("admin_extras", project_id=r.project_id))



@app.route("/admin/dodatki/reports", methods=["GET"])
@login_required
def admin_extra_reports():
    require_admin()
    q = ExtraReport.query.join(Project).order_by(ExtraReport.created_at.desc(), ExtraReport.id.desc()).limit(200).all()
    # auto-accept na widoku listy, żeby admin widział status od razu
    for rep in q:
        try:
            _auto_accept_if_due(rep)
        except Exception:
            pass

    body = render_template_string("""
<div class="card p-3">
  <div class="d-flex justify-content-between align-items-center">
    <h5 class="mb-0">Raporty dodatków</h5>
    <a class="btn btn-outline-secondary" href="{{ url_for('admin_extras') }}">Wróć do zgłoszeń</a>
  </div>

  <div class="table-responsive mt-3">
    <table class="table table-sm align-middle">
      <thead>
        <tr>
          <th>ID</th><th>Projekt</th><th>Status</th><th>Utworzono</th><th>Wysłano</th><th>E-mail</th><th>Suma</th><th class="text-end">Akcje</th>
        </tr>
      </thead>
      <tbody>
        {% for r in reps %}
          <tr>
            <td>#{{ r.id }}</td>
            <td>{{ r.project.name }}</td>
            <td><span class="badge bg-light text-dark border">{{ r.status }}</span></td>
            <td>{{ r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else '' }}</td>
            <td>{{ r.sent_at.strftime("%Y-%m-%d %H:%M") if r.sent_at else '' }}</td>
            <td>{{ r.recipient_email or '' }}</td>
            <td>{{ fmt(total(r)) }}</td>
            <td class="text-end text-nowrap">
              <a class="btn btn-sm btn-outline-primary" href="{{ url_for('admin_extra_report_view', report_id=r.id) }}">Otwórz</a>
              <form method="post" action="{{ url_for('admin_extra_report_delete', report_id=r.id) }}" style="display:inline;" onsubmit="return confirm('Usunąć raport?');">
                <button class="btn btn-sm btn-outline-danger">Usuń</button>
              </form>
              {% if r.status in ['APPROVED','APPROVED_AUTO','REJECTED','COMMENTED'] %}
                <a class="btn btn-sm btn-outline-success" href="{{ url_for('admin_extra_report_pdf', report_id=r.id) }}">PDF</a>
              {% endif %}
            </td>
          </tr>
        {% else %}
          <tr><td colspan="8" class="text-muted">Brak raportów.</td></tr>
        {% endfor %}
      </tbody>
    </table>
  </div>
</div>
""", reps=q, fmt=fmt_hhmm, total=_extra_report_total_minutes)

    return layout("Raporty dodatków", body)




@app.route("/admin/dodatki/report/<int:report_id>/delete", methods=["POST"])
@login_required
def admin_extra_report_delete(report_id):
    require_admin()
    rep = ExtraReport.query.get_or_404(report_id)
    audit = ExtraReportAudit.query.filter_by(report_id=rep.id).order_by(ExtraReportAudit.created_at.desc()).limit(100).all()
    dec = ExtraReportDecision.query.filter_by(report_id=rep.id).first()

    # Zbierz ID zgłoszeń zanim usuniemy raport (i jego items)
    req_ids = []
    try:
        req_ids = [it.request_id for it in (rep.items or []) if it and it.request_id]
    except Exception:
        req_ids = []

    # Cofnij status zgłoszeń na NEW (żeby checkbox znowu działał)
    if req_ids:
        try:
            ExtraRequest.query.filter(ExtraRequest.id.in_(req_ids)).update(
                {"status": "NEW"},
                synchronize_session=False
            )
        except Exception:
            # fallback (bez bulk update)
            try:
                for rid in req_ids:
                    req = ExtraRequest.query.get(rid)
                    if req:
                        req.status = "NEW"
            except Exception:
                pass

    # Usuń pliki załączników raportu (admin attachments)
    try:
        for att in list(rep.attachments or []):
            try:
                path = os.path.join(EXTRA_REPORT_ATTACH_DIR, att.stored_filename)
                if os.path.exists(path):
                    os.remove(path)
            except Exception:
                pass
    except Exception:
        pass

    # Usuń raport (cascade usunie items i attachment rekordy)
    db.session.delete(rep)
    db.session.commit()

    flash("Usunięto raport. Zgłoszenia wróciły do statusu NEW.", "success")
    # wróć do listy zgłoszeń dla tego projektu (wygodniej)
    try:
        return redirect(url_for("admin_extras", project_id=rep.project_id))
    except Exception:
        return redirect(url_for("admin_extras"))


@app.route("/admin/dodatki/report/<int:report_id>", methods=["GET", "POST"])
@login_required
def admin_extra_report_view(report_id):
    require_admin()
    rep = ExtraReport.query.get_or_404(report_id)
    decisions = _extra_report_get_decisions(rep.id)
    admin_atts = ExtraReportAttachment.query.filter_by(report_id=rep.id).order_by(ExtraReportAttachment.id.desc()).all()
    audit = ExtraReportAudit.query.filter_by(report_id=rep.id).order_by(ExtraReportAudit.created_at.desc()).all()
    audit = ExtraReportAudit.query.filter_by(report_id=rep.id).order_by(ExtraReportAudit.created_at.desc()).limit(100).all()
    dec = ExtraReportDecision.query.filter_by(report_id=rep.id).first()

    # auto accept jeśli minęło 7 dni
    _auto_accept_if_due(rep)

    if request.method == "POST":
        action = request.form.get("action") or "save"
        rep.report_text = (request.form.get("report_text") or "").strip() or None
        rep.recipient_email = (request.form.get("recipient_email") or "").strip() or None

        # override sumy (opcjonalnie)
        override = (request.form.get("total_override") or "").strip()
        if override:
            try:
                rep.total_minutes_override = parse_hhmm(override)
            except Exception:
                rep.total_minutes_override = None
        else:
            rep.total_minutes_override = None

        db.session.commit()

        # zapis załączników admina do raportu
        try:
            files = request.files.getlist("attachments")
            if files:
                saved_cnt = _save_extra_report_attachments(rep, files)
                if saved_cnt:
                    flash(f"Dodano załączniki: {saved_cnt} szt.", "success")
                    return redirect(url_for("admin_extra_report_view", report_id=rep.id))
        except Exception as e:
            flash(f"Nie udało się dodać załączników: {e}", "danger")

        if action == "send":
            if not rep.recipient_email:
                flash("Podaj e-mail odbiorcy przed wysyłką.", "warning")
                return redirect(url_for("admin_extra_report_view", report_id=rep.id))

            if not rep.token:
                rep.token = secrets.token_hex(32)

            rep.status = "SENT"
            rep.sent_at = datetime.utcnow()
            rep.updated_at = datetime.utcnow()
            db.session.commit()

            link = url_for("extra_report_public", token=rep.token, _external=True)
            subject = "Tilleggsrapport fra EKKO NOR AS"

            # informacja o auto-akceptacji po 7 dniach (w treści maila)

            auto_deadline = None

            if rep.sent_at:

                try:

                    auto_deadline = (rep.sent_at + timedelta(days=7)).date().isoformat()

                except Exception:

                    auto_deadline = None


            base_url = link.split("/dodatki/r/")[0] if "/dodatki/r/" in link else link.rsplit("/", 1)[0]

            logo_url = base_url.rstrip("/") + "/static/img/logo.png"

            deadline_txt = ("innen " + auto_deadline) if auto_deadline else "innen 7 dager"


            text_body = (

                "Hei!\n\n"

                "I lenken nedenfor sender vi dere rapporten.\n\n"

                f"Vennligst godkjenn {deadline_txt}. Dersom vi ikke mottar tilbakemelding innen fristen, vil rapporten bli automatisk godkjent.\n\n"

                "Åpne rapporten her:\n"

                f"{link}\n\n"

                "Ta gjerne kontakt dersom dere har spørsmål eller merknader.\n\n"

                "Med vennlig hilsen\nEKKO NOR AS\n"

            )


            html_body = f'''<!doctype html>
<html><body style=\"font-family:Arial,Helvetica,sans-serif;line-height:1.5;color:#111;\">
  <div style=\"max-width:640px;margin:0 auto;\">
    <img src=\"{logo_url}\" alt=\"EKKO NOR AS\" style=\"max-width:220px;height:auto;display:block;margin:0 0 16px 0;\">
    <p>Hei!</p>
    <p>I lenken nedenfor sender vi dere rapporten.</p>
    <p><strong>Vennligst godkjenn {deadline_txt}</strong>. Dersom vi ikke mottar tilbakemelding innen fristen, vil rapporten bli automatisk godkjent.</p>
    <p><a href=\"{link}\" style=\"display:inline-block;padding:10px 14px;background:#0d6efd;color:#fff;text-decoration:none;border-radius:6px;\">Åpne rapport</a></p>
    <p>Hvis knappen ikke fungerer, bruk denne lenken:<br><a href=\"{link}\">{link}</a></p>
    <p>Ta gjerne kontakt dersom dere har spørsmål eller merknader.</p>
    <p>Med vennlig hilsen<br><strong>EKKO NOR AS</strong></p>
  </div>
</body></html>'''


            try:

                _send_email_smtp(rep.recipient_email, subject, {"text": text_body, "html": html_body})
                flash("Wysłano raport. Link został wysłany e-mailem.", "success")
            except Exception as e:
                flash(f"Nie udało się wysłać maila: {e}", "danger")

            return redirect(url_for("admin_extra_report_view", report_id=rep.id))

        flash("Zapisano.", "success")
        return redirect(url_for("admin_extra_report_view", report_id=rep.id))



    link = url_for("extra_report_public", token=rep.token, _external=True) if rep.token else None

    body = render_template_string("""
<div class="row g-3">
  <div class="col-12">
    <div class="card p-3">
      <div class="d-flex justify-content-between align-items-center">
        <h5 class="mb-0">Raport dodatków #{{ rep.id }}</h5>
        <a class="btn btn-outline-secondary" href="{{ url_for('admin_extra_reports') }}">Lista raportów</a>
      </div>
      <div class="small text-muted mt-2">
        Projekt: <strong>{{ rep.project.name }}</strong><br>
        Status: <strong>{{ rep.status }}</strong>
      {% if rep.attachments %}<br>Załączniki: {% for a in rep.attachments %}<a href="{{ url_for('extra_report_public_attachment', token=rep.token, att_id=a.id) if rep.token else url_for('admin_extra_report_attachment_download', report_id=rep.id, att_id=a.id) }}" target="_blank" rel="noopener">{{ a.original_filename or "plik" }}</a>{% if not loop.last %}, {% endif %}{% endfor %}{% endif %}
        {% if rep.sent_at %} | Wysłano: {{ rep.sent_at.strftime("%Y-%m-%d %H:%M") }}{% endif %}
        {% if rep.decided_at %} | Decyzja: {{ rep.decided_at.strftime("%Y-%m-%d %H:%M") }}{% endif %}
      </div>
      {% if rep.decided_note %}
        <div class="alert alert-info mt-2 mb-0">{{ rep.decided_note }}</div>
      {% endif %}

      <hr class="my-3">
      <h6 class="mb-2">Historia zmian (log)</h6>
      <div class="row g-2">
        <div class="col-md-6">
          <div class="small text-muted mb-1">Decyzje klienta</div>
          <div class="table-responsive">
            <table class="table table-sm align-middle mb-0">
              <thead><tr><th>Tid</th><th>Navn</th><th>Merknad</th><th>Signatur</th></tr></thead>
              <tbody>
                {% for d in decisions %}
                  <tr>
                    <td>{{ d.decided_at.strftime("%Y-%m-%d %H:%M") if d.decided_at else "" }}</td>
                    <td>{{ d.user_name or "" }}</td>
                    <td>{{ d.decided_note or "" }}</td>
                    <td>{% if d.signature_png %}<a href="{{ url_for('admin_extra_report_signature_download', report_id=rep.id, dec_id=d.id) }}" target="_blank" rel="noopener">Vis</a>{% else %}-{% endif %}</td>
                  </tr>
                {% endfor %}
                {% if decisions|length == 0 %}<tr><td colspan="4" class="text-muted">Brak.</td></tr>{% endif %}
              </tbody>
            </table>
          </div>
        </div>
        <div class="col-md-6">
          <div class="small text-muted mb-1">Operasjoner i systemet</div>
          <div class="table-responsive">
            <table class="table table-sm align-middle mb-0">
              <thead><tr><th>Tid</th><th>Kto</th><th>Co</th></tr></thead>
              <tbody>
                {% for a in audit %}
                  <tr>
                    <td>{{ a.created_at.strftime("%Y-%m-%d %H:%M") if a.created_at else "" }}</td>
                    <td>{{ a.user_name or "" }}</td>
                    <td>{{ a.action or "" }}</td>
                  </tr>
                {% endfor %}
                {% if audit|length == 0 %}<tr><td colspan="3" class="text-muted">Brak.</td></tr>{% endif %}
              </tbody>
            </table>
          </div>
        </div>
      </div>
    </div>
  </div>

  <div class="col-12">
    <div class="card p-3">
      <form method="post" enctype="multipart/form-data">
        <div class="row g-2">
          <div class="col-md-5">
            <label class="form-label">E-mail odbiorcy</label>
            <input class="form-control" name="recipient_email" value="{{ rep.recipient_email or '' }}" placeholder="np. pm@firma.no">
          </div>
          <div class="col-md-3">
            <label class="form-label">Suma (override, opcjonalnie)</label>
            <input class="form-control" name="total_override" value="{{ fmt(total(rep)) if rep.total_minutes_override is not none else '' }}" placeholder="np. 12:30">
          </div>
          <div class="col-12">
            <label class="form-label">Treść raportu (opcjonalnie)</label>
            <textarea class="form-control" name="report_text" rows="3" placeholder="Opis dodatków...">{{ rep.report_text or '' }}</textarea>
            <div class="form-text">Na raporcie zawsze pokażemy informację o auto-akceptacji po 7 dniach od wysłania.</div>

            <div class="mt-2">
              <label class="form-label">Załączniki (pliki/zdjęcia do odbiorcy)</label>
              <input class="form-control" type="file" name="attachments" multiple>
              <div class="form-text">Limit: {{ max_attach_count }} plików, max {{ max_attach_mb }} MB / plik. Dozwolone: PDF, PNG/JPG/WEBP, DOC/DOCX, XLS/XLSX, TXT.</div>

              {% if rep.attachments %}
                <div class="mt-2">
                  <div class="small text-muted mb-1">Dodane pliki:</div>
                  <ul class="mb-0">
                    {% for a in rep.attachments %}
                      <li class="d-flex justify-content-between align-items-center gap-2">
                        <a href="{{ url_for('admin_extra_report_attachment_download', report_id=rep.id, att_id=a.id) }}" target="_blank" rel="noopener">{{ a.original_filename or a.stored_filename }}</a>
                        <form method="post" action="{{ url_for('admin_extra_report_attachment_delete', report_id=rep.id, att_id=a.id) }}" onsubmit="return confirm('Usunąć ten załącznik?');">
                          <button class="btn btn-sm btn-outline-danger">Usuń</button>
                        </form>
                      </li>
                    {% endfor %}
                  </ul>
                </div>
              {% endif %}
            </div>
          </div>

          <div class="col-12 d-flex gap-2">
            <button class="btn btn-primary" name="action" value="save">Zapisz</button>
            <button class="btn btn-outline-success" name="action" value="send">Zapisz i wyślij</button>
            {% if link %}
              <a class="btn btn-outline-secondary" href="{{ link }}" target="_blank" rel="noopener">Podgląd linku</a>
            {% endif %}
          </div>
        </div>
      </form>

      {% if link %}
        <hr class="my-3">
        <div class="small">
          Link do raportu: <a href="{{ link }}" target="_blank" rel="noopener">{{ link }}</a>
        </div>
      {% endif %}
    </div>
  </div>

  <div class="col-12">
    <div class="card p-3">
      <h6 class="mb-2">Pozycje w raporcie</h6>
      <div class="table-responsive">
        <table class="table table-sm align-middle">
          <thead><tr><th>Data</th><th>Pracownik</th><th>Godziny</th><th>Opis</th><th>Zdjęcia</th></tr></thead>
          <tbody>
            {% for it in rep.items %}
              <tr>
                <td>{{ it.work_date.isoformat() }}</td>
                <td>{{ it.user_name }}</td>
                <td>{{ fmt(it.minutes) }}</td>
                <td>{{ it.description or '' }}</td>
                <td>
                  {% if it.request and it.request.images %}
                    {% for img in it.request.images %}
                      <a href="{{ url_for('extra_image_view', image_id=img.id) }}" target="_blank" rel="noopener">IMG</a>{% if not loop.last %} {% endif %}
                    {% endfor %}
                  {% else %}-{% endif %}
                </td>
              </tr>
            {% endfor %}
          </tbody>
        </table>
      </div>
      <div class="mt-2 fw-bold">Suma: {{ fmt(total(rep)) }}</div>
    </div>
  </div>
</div>
""", rep=rep, audit=audit, decisions=decisions, fmt=fmt_hhmm, total=_extra_report_total_minutes, link=link)

    return layout("Raport dodatków", body)


@app.route("/dodatki/r/<token>/att/<int:att_id>", methods=["GET"])
def extra_report_public_attachment(token, att_id):
    rep = ExtraReport.query.filter_by(token=token).first_or_404()
    att = ExtraReportAttachment.query.filter_by(id=att_id, report_id=rep.id).first_or_404()
    path = os.path.join(EXTRA_REPORT_ATTACH_DIR, att.stored_filename)
    return send_file(path, as_attachment=True, download_name=(att.original_filename or att.stored_filename))


@app.route("/dodatki/r/<token>/img/<int:image_id>")
def extra_report_public_image(token, image_id):
    rep = ExtraReport.query.filter_by(token=token).first_or_404()
    _auto_accept_if_due(rep)

    img = ExtraRequestImage.query.get_or_404(image_id)

    ok = False
    try:
        for it in rep.items or []:
            if it.request_id == img.request_id:
                ok = True
                break
    except Exception:
        ok = False

    if not ok:
        abort(404)

    path = extra_image_view_path(img.stored_filename)
    if not os.path.exists(path):
        abort(404)
    return send_file(path)




@app.route("/dodatki/r/<token>/signature.png")
def extra_signature_public(token):
    rep = ExtraReport.query.filter_by(token=token).first_or_404()
    dec = ExtraReportDecision.query.filter_by(report_id=rep.id).first()
    if not dec or not dec.signature_png:
        abort(404)
    path = os.path.join(EXTRA_SIGNATURE_DIR, dec.signature_png)
    if not os.path.exists(path):
        abort(404)
    return send_file(path)


    img = ExtraRequestImage.query.get_or_404(image_id)

    ok = False
    try:
        for it in rep.items or []:
            if it.request_id == img.request_id:
                ok = True
                break
    except Exception:
        ok = False

    if not ok:
        abort(404)

    path = extra_image_view_path(img.stored_filename)
    if not os.path.exists(path):
        abort(404)
    return send_file(path)



@app.route("/dodatki/r/<token>", methods=["GET", "POST"])
def extra_report_public(token):
    rep = ExtraReport.query.filter_by(token=token).first_or_404()

    # language: default Norwegian (no). Optional Polish: ?lang=pl
    lang = (request.args.get("lang") or "no").lower().strip()
    if lang not in ("no", "pl"):
        lang = "no"

    def tr(no_txt, pl_txt=None):
        if lang == "pl":
            return pl_txt if pl_txt is not None else no_txt
        return no_txt

    # auto accept jeśli minęło 7 dni
    _auto_accept_if_due(rep)

    decisions = _extra_report_get_decisions(rep.id)
    admin_atts = ExtraReportAttachment.query.filter_by(report_id=rep.id).order_by(ExtraReportAttachment.id.desc()).all()
    audit = ExtraReportAudit.query.filter_by(report_id=rep.id).order_by(ExtraReportAudit.created_at.desc()).all()

    if request.method == "POST":
        if rep.status not in ("SENT",):
            flash(tr("Denne rapporten venter ikke lenger på en beslutning.",
                     "Ten raport nie oczekuje już na decyzję."), "warning")
            return redirect(url_for("extra_report_public", token=token, lang=lang))

        action = request.form.get("action")
        note = (request.form.get("note") or "").strip() or None
        sign_name = (request.form.get("sign_name") or "").strip() or None
        signature_data = (request.form.get("signature_data") or "").strip() or None

        rep.decided_at = datetime.utcnow()
        rep.decided_note = note

        if action == "approve":
            rep.status = "APPROVED"
            if not rep.decided_note:
                rep.decided_note = tr("Godkjent.", "Zaakceptowano.")
        elif action == "reject":
            rep.status = "REJECTED"
            if not rep.decided_note:
                rep.decided_note = tr("Avvist.", "Odrzucono.")
        else:
            rep.status = "COMMENTED"
            if not rep.decided_note:
                rep.decided_note = tr("Kommentarer lagt til.", "Dodano uwagi.")

        # zapisz decyzję + podpis do osobnej tabeli (bez ryzykownych migracji)
        try:
            dec = ExtraReportDecision.query.filter_by(report_id=rep.id).first()
            if not dec:
                dec = ExtraReportDecision(report_id=rep.id)
                db.session.add(dec)
            dec.decided_at = rep.decided_at
            dec.decided_note = rep.decided_note
            dec.decided_name = sign_name or ""
            dec.user_name = (sign_name or "Klient")
            dec.work_date = date.today()
            dec.minutes = 0
            fn = _save_signature_png(signature_data) if signature_data else None
            if fn:
                dec.signature_png = fn
        except Exception:
            pass

        db.session.commit()

        try:
            _extra_audit(rep, rep.status.lower(), actor_type="public", actor_name=sign_name, details=rep.decided_note)
        except Exception:
            pass

        try:
            _notify_extra_report_status(rep, "status change")
        except Exception:
            pass

        flash(tr("Takk. Beslutningen er lagret.", "Dziękujemy. Zapisano decyzję."), "success")
        return redirect(url_for("extra_report_public", token=token, lang=lang))

    auto_date = None
    if rep.sent_at:
        auto_date = (rep.sent_at + timedelta(days=7)).date()

    base_no = url_for("extra_report_public", token=rep.token, lang="no")
    base_pl = url_for("extra_report_public", token=rep.token, lang="pl")

    # fetch decision for display (name/signature)
    dec = ExtraReportDecision.query.filter_by(report_id=rep.id).first()

    body = render_template_string(r"""
<div class="container-narrow">
  <style>
    .report-head{display:flex;gap:16px;align-items:flex-start;margin-bottom:12px;}
    .report-head img{max-width:220px;height:auto;}
    .thumbs img{width:90px;height:70px;object-fit:cover;border-radius:6px;border:1px solid #ddd;}
  </style>
  <div class="card p-3">
    <div class="report-head">
      <div><img src="{{ url_for('static', filename='img/logo.png') }}" alt="EKKO NOR AS"></div>
      <div class="flex-grow-1">
        <div class="small text-muted mb-1">EKKO NOR AS</div>
      </div>
    </div>
    <div class="d-flex justify-content-between align-items-start">
      <div>
        <h5 class="mb-1">{{ tr("Tilleggsrapport", "Raport dodatków") }}</h5>
        <div class="small text-muted">
          {{ tr("Prosjekt", "Projekt") }}: <strong>{{ rep.project.name }}</strong><br>
          {{ tr("Status", "Status") }}: <strong>{{ rep.status }}</strong>
          {% if rep.sent_at %}<br>{{ tr("Sendt", "Wysłano") }}: {{ rep.sent_at.strftime("%Y-%m-%d %H:%M") }}{% endif %}
          {% if auto_date %}<br><strong>{{ tr("Merk", "Uwaga") }}:</strong> {{ tr("rapporten blir automatisk godkjent 7 dager etter sending", "raport zostanie automatycznie zatwierdzony po 7 dniach od wysłania") }} ({{ auto_date.isoformat() }}).{% endif %}
        </div>
      </div>
      <div class="text-end">
        <div class="small text-muted mb-1">{{ tr("Språk", "Język") }}:</div>
        <a class="btn btn-sm {% if lang=='no' %}btn-primary{% else %}btn-outline-primary{% endif %}" href="{{ base_no }}">NO</a>
        <a class="btn btn-sm {% if lang=='pl' %}btn-primary{% else %}btn-outline-primary{% endif %}" href="{{ base_pl }}">PL</a>
      </div>
    </div>

    {% if rep.report_text %}
      <hr class="my-3">
      <div style="white-space:pre-wrap">{{ rep.report_text }}</div>
    {% endif %}

    <hr class="my-3">
    <h6 class="mb-2">{{ tr("Linjer", "Pozycje") }}</h6>
    <div class="table-responsive">
      <table class="table table-sm align-middle">
        <thead>
          <tr>
            <th>{{ tr("Dato", "Data") }}</th>
            <th>{{ tr("Ansatt", "Pracownik") }}</th>
            <th>{{ tr("Timer", "Godziny") }}</th>
            <th>{{ tr("Beskrivelse", "Opis") }}</th>
            <th>{{ tr("Bilder", "Zdjęcia") }}</th>
          </tr>
        </thead>
        <tbody>
          {% for it in rep.items %}
            <tr>
              <td>{{ it.work_date.isoformat() }}</td>
              <td>{{ it.user_name }}</td>
              <td>{{ fmt(it.minutes) }}</td>
              <td>{{ it.description or '' }}</td>
              <td>
                {% if it.request and it.request.images %}
                  {% for img in it.request.images %}
                    <a href="{{ url_for('extra_report_public_image', token=rep.token, image_id=img.id) }}" target="_blank" rel="noopener" style="display:inline-block;margin-right:6px;"><img src="{{ url_for('extra_report_public_image', token=rep.token, image_id=img.id) }}" alt="img" style="width:90px;height:70px;object-fit:cover;border-radius:6px;border:1px solid #ddd;"></a>
                  {% endfor %}
                  </div>
                {% else %}-{% endif %}
              </td>
            </tr>
          {% endfor %}
        </tbody>
      </table>
    </div>

    {% if rep.attachments %}
      <hr class="my-3">
      <h6 class="mb-2">{{ tr("Vedlegg til rapporten", "Załączniki do raportu") }}</h6>
      <div class="list-group">
        {% for a in rep.attachments %}
          <a class="list-group-item list-group-item-action d-flex justify-content-between align-items-center"
             href="{{ url_for('extra_report_public_attachment', token=rep.token, att_id=a.id) if rep.token else url_for('admin_extra_report_attachment_download', report_id=rep.id, att_id=a.id) }}" target="_blank" rel="noopener">
            <span>{{ a.original_filename or tr("fil", "plik") }}</span>
            <span class="badge bg-light text-dark border">{{ tr("Last ned", "Pobierz") }}</span>
          </a>
        {% endfor %}
      </div>
    {% endif %}

    <div class="mt-3 fw-bold">{{ tr("Sum", "Suma") }}: {{ fmt(total_minutes(rep)) }}</div>

    {% if rep.status == 'SENT' %}
      <hr class="my-3">
      <form method="post" class="row g-2" onsubmit="storeSig()">
        <div class="col-12">
          <label class="form-label">{{ tr("Kommentar (valgfritt)", "Uwagi (opcjonalnie)") }}</label>
          <textarea class="form-control" name="note" rows="3" placeholder="{{ tr('Hvis du vil legge til noe...', 'Jeśli chcesz coś dopisać...') }}"></textarea>
        </div>
        <div class="col-12">
          <label class="form-label">{{ tr("Navn (signatur)", "Imię i nazwisko (podpis)") }}</label>
          <input class="form-control" name="sign_name" placeholder="{{ tr('Skriv ditt navn', 'Wpisz swoje imię i nazwisko') }}">
        </div>
        <div class="col-12">
          <label class="form-label">{{ tr("Håndskrevet signatur (valgfritt)", "Podpis odręczny (opcjonalnie)") }}</label>
          <div class="border rounded p-2" style="background:#fff;">
            <canvas id="sigpad" width="520" height="140" style="width:100%;max-width:520px;touch-action:none;"></canvas>
          </div>
          <div class="mt-2 d-flex gap-2">
            <button type="button" class="btn btn-sm btn-outline-secondary" onclick="sigClear()">{{ tr("Tøm", "Wyczyść") }}</button>
          </div>
          <input type="hidden" name="signature_data" id="signature_data">
          <div class="small text-muted mt-1">{{ tr("Hvis du ikke signerer her, kan du bare skrive navnet ditt.", "Jeśli nie podpiszesz się tutaj, wystarczy wpisać imię i nazwisko.") }}</div>
        </div>

        <div class="col-12 d-flex gap-2">
          <button class="btn btn-success" name="action" value="approve">{{ tr("Godkjenn", "Zatwierdź") }}</button>
          <button class="btn btn-danger" name="action" value="reject">{{ tr("Avvis", "Odrzuć") }}</button>
          <button class="btn btn-outline-primary" name="action" value="comment">{{ tr("Legg til kommentar", "Dodaj uwagi") }}</button>
        </div>
      </form>

      <script>
        (function(){
          const c = document.getElementById('sigpad');
          if(!c) return;
          const ctx = c.getContext('2d');
          ctx.lineWidth = 2;
          let drawing = false;
          let last = null;

          function pos(evt){
            const r = c.getBoundingClientRect();
            const p = (evt.touches && evt.touches[0]) ? evt.touches[0] : evt;
            return {x: (p.clientX - r.left) * (c.width / r.width), y: (p.clientY - r.top) * (c.height / r.height)};
          }

          function start(evt){ drawing = true; last = pos(evt); evt.preventDefault(); }
          function move(evt){
            if(!drawing) return;
            const p = pos(evt);
            ctx.beginPath();
            ctx.moveTo(last.x, last.y);
            ctx.lineTo(p.x, p.y);
            ctx.stroke();
            last = p;
            evt.preventDefault();
          }
          function end(evt){ drawing = false; last = null; evt.preventDefault(); }

          c.addEventListener('mousedown', start);
          c.addEventListener('mousemove', move);
          window.addEventListener('mouseup', end);

          c.addEventListener('touchstart', start, {passive:false});
          c.addEventListener('touchmove', move, {passive:false});
          window.addEventListener('touchend', end, {passive:false});

          window.sigClear = function(){ ctx.clearRect(0,0,c.width,c.height); };

          window.storeSig = function(){
            try{
              document.getElementById('signature_data').value = c.toDataURL('image/png');
            }catch(e){}
            return true;
          }
        })();
      </script>
    {% else %}
      {% if rep.decided_note %}
        <hr class="my-3">
        <div class="alert alert-info mb-2" style="white-space:pre-wrap">{{ rep.decided_note }}</div>
      {% endif %}

      {% if dec and (dec.decided_name or dec.signature_png) %}
        <div class="small text-muted">{{ tr("Signert av", "Podpis") }}: <strong>{{ dec.decided_name or "-" }}</strong> {% if dec.decided_at %}({{ dec.decided_at.strftime("%Y-%m-%d %H:%M") }}){% endif %}</div>
        {% if dec.signature_png %}
          <div class="mt-2">
            <img alt="signature" src="{{ url_for('extra_signature_public', token=rep.token) }}" style="max-width:520px;width:100%;border:1px solid #ddd;border-radius:8px;background:#fff;">
          </div>
        {% endif %}
      {% endif %}
    {% endif %}
  </div>
</div>
""", rep=rep, decisions=decisions, audit=audit, admin_atts=admin_atts,
       fmt=fmt_hhmm, total_minutes=_extra_report_total_minutes, auto_date=auto_date,
       lang=lang, tr=tr, base_no=base_no, base_pl=base_pl)

    return layout(tr("Tilleggsrapport", "Raport dodatków"), body)


@app.route("/admin/dodatki/report/<int:report_id>/att/<int:att_id>", methods=["GET"])
@login_required
def admin_extra_report_attachment_download(report_id, att_id):
    require_admin()
    rep = ExtraReport.query.get_or_404(report_id)
    att = ExtraReportAttachment.query.filter_by(id=att_id, report_id=rep.id).first_or_404()
    path = os.path.join(EXTRA_ATTACH_DIR, att.stored_filename)
    if not os.path.exists(path):
        abort(404)
    # Download with original filename
    try:
        return send_file(path, as_attachment=True, download_name=att.original_filename)
    except TypeError:
        # Older Werkzeug/Flask fallback
        return send_file(path, as_attachment=True)

@app.route("/admin/dodatki/report/<int:report_id>/sig/<int:dec_id>")
@login_required
def admin_extra_report_signature_download(report_id, dec_id):
    require_admin()
    rep = ExtraReport.query.get_or_404(report_id)
    dec = ExtraReportDecision.query.filter_by(id=dec_id, report_id=rep.id).first_or_404()
    if not dec.signature_png:
        abort(404)
    path = os.path.join(EXTRA_SIG_DIR, dec.signature_png)
    if not os.path.exists(path):
        abort(404)
    return send_file(path, as_attachment=False)

def admin_extra_report_attachment_download(report_id, att_id):
    require_admin()
    rep = ExtraReport.query.get_or_404(report_id)
    audit = ExtraReportAudit.query.filter_by(report_id=rep.id).order_by(ExtraReportAudit.created_at.desc()).limit(100).all()
    dec = ExtraReportDecision.query.filter_by(report_id=rep.id).first()
    att = ExtraReportAttachment.query.filter_by(id=att_id, report_id=rep.id).first_or_404()
    path = os.path.join(EXTRA_REPORT_ATTACH_DIR, att.stored_filename)
    return send_file(path, as_attachment=True, download_name=(att.original_filename or att.stored_filename))


@app.route("/admin/dodatki/report/<int:report_id>/att/<int:att_id>/delete", methods=["POST"])
@login_required
def admin_extra_report_attachment_delete(report_id, att_id):
    require_admin()
    rep = ExtraReport.query.get_or_404(report_id)
    audit = ExtraReportAudit.query.filter_by(report_id=rep.id).order_by(ExtraReportAudit.created_at.desc()).limit(100).all()
    dec = ExtraReportDecision.query.filter_by(report_id=rep.id).first()
    att = ExtraReportAttachment.query.filter_by(id=att_id, report_id=rep.id).first_or_404()
    try:
        path = os.path.join(EXTRA_REPORT_ATTACH_DIR, att.stored_filename)
        if os.path.exists(path):
            os.remove(path)
    except Exception:
        pass
    db.session.delete(att)
    db.session.commit()
    flash("Usunięto załącznik.", "success")
    return redirect(url_for("admin_extra_report_view", report_id=rep.id))


@app.route("/admin/dodatki/report/<int:report_id>/pdf", methods=["GET"])
@login_required
def admin_extra_report_pdf(report_id):
    require_admin()
    rep = ExtraReport.query.get_or_404(report_id)
    audit = ExtraReportAudit.query.filter_by(report_id=rep.id).order_by(ExtraReportAudit.created_at.desc()).limit(100).all()
    dec = ExtraReportDecision.query.filter_by(report_id=rep.id).first()

    if rep.status not in ("APPROVED", "APPROVED_AUTO", "REJECTED", "COMMENTED"):
        flash("PDF jest dostępny po decyzji (zatwierdzenie/odrzucenie/uwagi).", "warning")
        return redirect(url_for("admin_extra_report_view", report_id=rep.id))

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
    except Exception:
        abort(500, "Brak pakietu reportlab (dodaj do requirements).")

    mem = io.BytesIO()
    c = canvas.Canvas(mem, pagesize=A4)
    w, h = A4

    # Logo (duże, lewy górny róg)
    logo_path = os.path.join(BASE_DIR, "static", "img", "logo.png")
    logo_w = 140
    logo_h = 45
    text_x = 50
    if os.path.exists(logo_path):
        try:
            from reportlab.lib.utils import ImageReader
            c.drawImage(ImageReader(logo_path), 50, h - 70, width=logo_w, height=logo_h, mask='auto', preserveAspectRatio=True)
            text_x = 50 + logo_w + 20
        except Exception:
            text_x = 50

    y = h - 60
    c.setFont("Helvetica-Bold", 14)
    c.drawString(text_x, y, f"Raport dodatków #{rep.id}")
    y -= 22
    c.setFont("Helvetica", 10)
    c.drawString(50, y, f"Projekt: {rep.project.name}")
    y -= 14
    c.drawString(50, y, f"Status: {rep.status}")
    y -= 14
    if rep.sent_at:
        c.drawString(50, y, f"Wysłano: {rep.sent_at.strftime('%Y-%m-%d %H:%M')}")
        y -= 14
        auto_d = (rep.sent_at + timedelta(days=7)).date().isoformat()
        c.drawString(50, y, f"Auto-akceptacja po 7 dniach: {auto_d}")
        y -= 14
    if rep.decided_at:
        c.drawString(50, y, f"Decyzja: {rep.decided_at.strftime('%Y-%m-%d %H:%M')}")
        y -= 14

    if rep.report_text:
        y -= 8
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, y, "Treść:")
        y -= 14
        c.setFont("Helvetica", 10)
        for line in rep.report_text.splitlines():
            if y < 80:
                c.showPage()
                y = h - 60
                c.setFont("Helvetica", 10)
            c.drawString(50, y, line[:110])
            y -= 12

    y -= 8
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, y, "Pozycje:")
    y -= 14
    c.setFont("Helvetica", 10)

    for it in rep.items:
        line = f"{it.work_date.isoformat()} | {it.user_name} | {fmt_hhmm(it.minutes)} | {(it.description or '')}"
        # łamanie proste
        chunks = [line[i:i+110] for i in range(0, len(line), 110)] or [line]
        for ch in chunks:
            if y < 80:
                c.showPage()
                y = h - 60
                c.setFont("Helvetica", 10)
            c.drawString(50, y, ch)
            y -= 12

    y -= 10
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, y, f"Suma: {fmt_hhmm(_extra_report_total_minutes(rep))}")

    if rep.decided_note:
        y -= 18
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, y, "Uwagi/Decyzja:")
        y -= 14
        c.setFont("Helvetica", 10)
        for line in rep.decided_note.splitlines():
            if y < 80:
                c.showPage()
                y = h - 60
                c.setFont("Helvetica", 10)
            c.drawString(50, y, line[:110])
            y -= 12

    c.showPage()
    c.save()
    mem.seek(0)
    return send_file(mem, as_attachment=True, download_name=f"raport_dodatki_{rep.id}.pdf", mimetype="application/pdf")




def _save_extra_report_attachments(rep: ExtraReport, files) -> int:
    """Zapisuje załączniki admina do raportu (pliki/zdjęcia). Zwraca liczbę zapisanych plików."""
    if not files:
        return 0
    os.makedirs(EXTRA_REPORT_ATTACH_DIR, exist_ok=True)

    existing = len(rep.attachments or [])
    saved = 0

    for f in files:
        if not f or not getattr(f, "filename", ""):
            continue
        if existing + saved >= MAX_ATTACH_COUNT:
            break

        original = f.filename
        ext = os.path.splitext(original)[1].lower()
        if ext not in ALLOWED_ATTACH_EXTS:
            continue

        # limit rozmiaru (jeśli da się odczytać)
        try:
            pos = f.stream.tell()
            f.stream.seek(0, os.SEEK_END)
            size = f.stream.tell()
            f.stream.seek(pos, os.SEEK_SET)
        except Exception:
            size = None

        if size is not None and size > MAX_ATTACH_BYTES:
            continue

        stored = f"eratt_{rep.id}_{uuid.uuid4().hex}{ext}"
        out_path = os.path.join(EXTRA_REPORT_ATTACH_DIR, stored)
        f.save(out_path)

        att = ExtraReportAttachment(
            report_id=rep.id,
            stored_filename=stored,
            original_filename=original,
        )
        db.session.add(att)
        saved += 1

    if saved:
        db.session.commit()
    return saved


def _notify_extra_report_status(rep: ExtraReport, reason: str) -> None:
    """Wysyła maila do admina o zmianie statusu raportu (np. podpis/odrzucenie/uwagi/auto-akcept)."""
    notify_to = (os.getenv("REPORT_STATUS_NOTIFY_TO") or "").strip() or None
    fallback = (os.getenv("SMTP_USER") or "").strip() or None


EXTRA_SIGNATURE_DIR = os.path.join(UPLOAD_DIR, "extra_signatures")
os.makedirs(EXTRA_SIGNATURE_DIR, exist_ok=True)

def _extra_audit(rep, action, actor_type="system", actor_name=None, details=None):
    try:
        a = ExtraReportAudit(
            report_id=rep.id,
            actor_type=actor_type,
            actor_name=actor_name,
            action=action,
            ip=request.remote_addr if request else None,
            user_agent=(request.headers.get("User-Agent") if request else None),
            details=details,
        )
        db.session.add(a)
        db.session.commit()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass

def _save_signature_png(data_url):
    """
    Accepts a data URL like 'data:image/png;base64,...' and stores it to EXTRA_SIGNATURE_DIR.
    Returns stored filename or None.
    """
    if not data_url or "," not in data_url:
        return None
    try:
        header, b64 = data_url.split(",", 1)
        if "image/png" not in header:
            return None
        import base64, uuid
        raw = base64.b64decode(b64.encode("utf-8"))
        name = f"sig_{uuid.uuid4().hex}.png"
        path = os.path.join(EXTRA_SIGNATURE_DIR, name)
        with open(path, "wb") as f:
            f.write(raw)
        # naive "blank signature" filter: very small png likely means empty
        if os.path.getsize(path) < 1200:
            try:
                os.remove(path)
            except Exception:
                pass
            return None
        return name
    except Exception:
        return None


    recipients = []
    for r in (notify_to, fallback):
        if r and r not in recipients:
            recipients.append(r)

    if not recipients:
        return

    admin_link = url_for("admin_extra_report_view", report_id=rep.id, _external=True)
    public_link = url_for("extra_report_public", token=rep.token, _external=True) if rep.token else None

    lines = [
        f"Zmiana statusu raportu dodatków #{rep.id}",
        f"Projekt: {rep.project.name}",
        f"Status: {rep.status}",
    ]
    if rep.recipient_email:
        lines.append(f"Odbiorca: {rep.recipient_email}")
    if rep.sent_at:
        lines.append(f"Wysłano: {rep.sent_at.strftime('%Y-%m-%d %H:%M UTC')}")
    if rep.decided_at:
        lines.append(f"Decyzja: {rep.decided_at.strftime('%Y-%m-%d %H:%M UTC')}")
    if rep.decided_note:
        lines.append(f"Uwagi/nota: {rep.decided_note}")
    lines.append(f"Powód: {reason}")
    lines.append("")
    lines.append(f"Panel admina: {admin_link}")
    if public_link:
        lines.append(f"Link publiczny: {public_link}")

    subject = f"[Dodatki] Status raportu #{rep.id}: {rep.status}"
    body = "\n".join(lines)

    for to in recipients:
        try:
            _send_smtp_email(to, subject, body)
        except Exception:
            pass



# --- Init DB after all models/routes are defined ---
init_db()



@app.route("/dodatki/r/<token>/pdf", methods=["GET"])
def extra_report_public_pdf(token):
    rep = ExtraReport.query.filter_by(token=token).first_or_404()
    # only after client decision (accepted/rejected) or when admin already sent it
    if rep.status not in ("accepted", "rejected", "sent"):
        flash("PDF będzie dostępny po akceptacji/odrzuceniu.", "warning")
        return redirect(url_for("extra_report_public", token=token))

    lang = (request.args.get("lang") or rep.lang or "no").strip().lower()
    if lang not in ("no", "pl"):
        lang = "no"

    # reuse the same PDF generator as admin
    # we call the admin function body logic by duplicating minimal parts
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    from reportlab.lib.units import mm
    from reportlab.lib.utils import ImageReader
    from reportlab.lib import colors
    from io import BytesIO

    buf = BytesIO()
    c = canvas.Canvas(buf, pagesize=A4)
    w, h = A4

    tr = _tr_no if lang == "no" else _tr_pl
    fmt = fmt_hhmm

    y = h - 20*mm
    c.setFont("Helvetica-Bold", 14)
    c.drawString(20*mm, y, tr("Ekstra rapport", "Raport dodatków"))
    y -= 8*mm

    c.setFont("Helvetica", 10)
    c.drawString(20*mm, y, f"{tr('Prosjekt', 'Projekt')}: {rep.project.name if rep.project else '-'}")
    y -= 5*mm
    c.drawString(20*mm, y, f"{tr('Dato', 'Data')}: {rep.created_at.date().isoformat() if rep.created_at else ''}")
    y -= 8*mm

    # items
    c.setFont("Helvetica-Bold", 10)
    c.drawString(20*mm, y, tr("Linjer", "Pozycje"))
    y -= 6*mm

    c.setFont("Helvetica", 9)
    items = ExtraReportItem.query.filter_by(report_id=rep.id).order_by(ExtraReportItem.id.asc()).all()
    total = 0
    for it in items:
        if y < 30*mm:
            c.showPage()
            y = h - 20*mm
            c.setFont("Helvetica", 9)
        line = f"{it.work_date.isoformat() if it.work_date else ''} | {it.user_name or ''} | {it.title or ''} | {fmt(it.minutes or 0)}"
        c.drawString(20*mm, y, line[:110])
        y -= 5*mm
        total += int(it.minutes or 0)

    y -= 4*mm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(20*mm, y, f"{tr('Sum', 'Suma')}: {fmt(total)}")
    y -= 10*mm

    # decision + signature
    decisions = _extra_report_get_decisions(rep.id)
    if decisions:
        d = decisions[0]
        c.setFont("Helvetica-Bold", 10)
        c.drawString(20*mm, y, tr("Signatur", "Podpis"))
        y -= 6*mm
        c.setFont("Helvetica", 9)
        c.drawString(20*mm, y, f"{tr('Navn', 'Imię i nazwisko')}: {d.decided_name or d.user_name or ''}")
        y -= 5*mm
        if d.decided_note:
            c.drawString(20*mm, y, (tr("Kommentar", "Komentarz") + ": " + d.decided_note)[:110])
            y -= 5*mm
        if d.signature_png:
            sig_path = os.path.join(EXTRA_REPORT_SIG_DIR, d.signature_png)
            if os.path.exists(sig_path):
                try:
                    img = ImageReader(sig_path)
                    c.drawImage(img, 20*mm, y-35*mm, width=80*mm, height=30*mm, preserveAspectRatio=True, mask='auto')
                    y -= 40*mm
                except Exception:
                    pass

    c.showPage()
    c.save()
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf", as_attachment=False,
                     download_name=f"ekstra_rapport_{rep.id}.pdf")


if __name__ == "__main__":
    ensure_db_file()
    app.run(debug=True, host="0.0.0.0")
# =========================
# SMTP helper (added)
# Fix for: NameError: _send_email_smtp is not defined
# =========================
def _send_email_smtp(to, subject, body, attachments=None):
    """
    Wysyłka maila SMTP (Render / ENV), obsługuje SSL (465) i STARTTLS (587).

    ENV (zalecane na Render):
      SMTP_HOST
      SMTP_PORT  (465 dla SSL lub 587 dla STARTTLS)
      SMTP_USER
      SMTP_PASSWORD
      MAIL_FROM  (opcjonalnie)
      SMTP_STARTTLS (opcjonalnie: "1" wymusza STARTTLS)
      SMTP_SSL (opcjonalnie: "1" wymusza SMTP_SSL)
    """

    smtp_host = os.getenv("SMTP_HOST", app.config.get("SMTP_HOST"))
    smtp_port = int(os.getenv("SMTP_PORT", app.config.get("SMTP_PORT", 465)))
    smtp_user = os.getenv("SMTP_USER", app.config.get("SMTP_USER"))
    smtp_pass = os.getenv("SMTP_PASSWORD", app.config.get("SMTP_PASSWORD"))
    mail_from = os.getenv("MAIL_FROM", app.config.get("MAIL_FROM", smtp_user))

    force_starttls = os.getenv("SMTP_STARTTLS", "").strip() in ("1", "true", "True", "yes", "YES")
    force_ssl = os.getenv("SMTP_SSL", "").strip() in ("1", "true", "True", "yes", "YES")

    if not smtp_host or not smtp_user or not smtp_pass or not mail_from:
        raise RuntimeError("Brak konfiguracji SMTP (SMTP_HOST/SMTP_USER/SMTP_PASSWORD/MAIL_FROM)")

    msg = EmailMessage()
    msg["From"] = mail_from
    msg["To"] = to
    msg["Subject"] = subject
        # body może być str (tekst) albo dict {'text':..., 'html':...}
    if isinstance(body, dict):
        text_body = (body.get('text') or '').strip()
        html_body = (body.get('html') or '').strip()
        msg.set_content(text_body)
        if html_body:
            msg.add_alternative(html_body, subtype='html')
    else:
        msg.set_content(body or "")

    if attachments:
        for path in attachments:
            if not path or not os.path.exists(path):
                continue
            with open(path, "rb") as f:
                data = f.read()
            filename = os.path.basename(path)
            msg.add_attachment(
                data,
                maintype="application",
                subtype="octet-stream",
                filename=filename,
            )

    # Heurystyka:
    # - port 465 => SSL
    # - port 587/25 => STARTTLS (jeśli dostępne)
    use_ssl = force_ssl or (smtp_port == 465 and not force_starttls)
    use_starttls = force_starttls or (smtp_port in (587, 25) and not use_ssl)

    if use_ssl:
        with smtplib.SMTP_SSL(smtp_host, smtp_port, timeout=20) as server:
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
    else:
        with smtplib.SMTP(smtp_host, smtp_port, timeout=20) as server:
            server.ehlo()
            if use_starttls:
                server.starttls()
                server.ehlo()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
